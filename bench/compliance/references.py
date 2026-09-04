"""Strict, provenance-carrying importers for IBSI 1 and IBSI 2 references."""

from __future__ import annotations

import csv
import io
import json
import math
import re
from collections import Counter, defaultdict
from dataclasses import replace
from pathlib import Path
from typing import Any, Iterable, Optional

from bench.benchmark_ledger import atomic_write_json, atomic_write_text, sha256_file
from bench.compliance.models import ReferenceRecord
from bench.ibsi_families import CODE_TO_FAMILY
from bench.ibsi_mapping import MIRP_PREFIX_MAP, map_mirp


REFERENCE_SCHEMA_VERSION = 1

IBSI1_WORKBOOK_URL = (
    "https://ibsi.radiomics.hevs.ch/assets/IBSI-1-submission-table.xlsx"
)
IBSI1_WORKBOOK_SHA256 = (
    "3af20577b580f2161bd546e147338ff226abb1676488e2906f2ce908b7985710"
)
IBSI1_DIGITAL_PHANTOM_IMAGE_SHA256 = (
    "83773ac2a288aa93cf819a98eaf301c18d35d5876182ec9c18733184e2b7a83a"
)
IBSI1_DIGITAL_PHANTOM_MASK_SHA256 = (
    "3032d340944b577b83bc559ea7ae9a6034b84633a840eff39570ec17aa1d6281"
)
IBSI1_WORKBOOK_SHEETS = {
    "digital phantom": (487, 482),
    "config A": (411, 406),
    "config B": (411, 406),
    "config C": (275, 270),
    "config D": (275, 270),
    "config E": (275, 270),
}
IBSI1_PROFILE = "digital_phantom_3d_merged"
IBSI1_PROFILE_ROWS = 174
IBSI1_PROFILE_STANDARDIZED = 169
IBSI1_REQUIRED_DIRECTIONAL_AGGREGATION = "3d_merge"

# IBSI 1 Table 2 defines 172 features.  The submission workbook contains 174
# benchmark instances because two parameterised IVH definitions are evaluated
# at both 10% and 90%.  Keep all three denominators explicit: definitions for
# coverage, benchmark instances for traceability, and standardized instances
# for numerical tolerance checks.
IBSI1_TABLE2_FAMILY_DEFINITIONS = {
    "morphology": 29,
    "local_intensity": 2,
    "intensity": 18,
    "histogram": 23,
    "ivh": 5,
    "glcm": 25,
    "glrlm": 16,
    "glszm": 16,
    "gldzm": 16,
    "ngtdm": 5,
    "ngldm": 17,
}
IBSI1_BENCHMARK_INSTANCES_BY_FAMILY = {
    **IBSI1_TABLE2_FAMILY_DEFINITIONS,
    "ivh": 7,
}
IBSI1_STANDARDIZED_INSTANCES_BY_FAMILY = {
    **IBSI1_BENCHMARK_INSTANCES_BY_FAMILY,
    "morphology": 25,
    "ivh": 6,
}
IBSI1_TABLE2_DEFINITION_COUNT = sum(IBSI1_TABLE2_FAMILY_DEFINITIONS.values())
IBSI1_CONFIGURATION_PROFILE_COUNTS = {
    "A": (174, 169),
    "B": (174, 169),
    "C": (174, 169),
    "D": (174, 169),
    "E": (174, 169),
}
IBSI1_CONFIGURATION_AGGREGATIONS = {
    "A": (
        "2d_average",
        "2d_slice_merge",
        "2.5d_direction_merge",
        "2.5d_merge",
    ),
    "B": (
        "2d_average",
        "2d_slice_merge",
        "2.5d_direction_merge",
        "2.5d_merge",
    ),
    "C": ("3d_average", "3d_merge"),
    "D": ("3d_average", "3d_merge"),
    "E": ("3d_average", "3d_merge"),
}

_IBSI1_DIAGNOSTIC_IMAGE_BASES = (
    "img_dim_x",
    "img_dim_y",
    "img_dim_z",
    "vox_dim_x",
    "vox_dim_y",
    "vox_dim_z",
    "mean_int",
    "min_int",
    "max_int",
)
_IBSI1_DIAGNOSTIC_ROI_BASES = (
    "int_mask_dim_x",
    "int_mask_dim_y",
    "int_mask_dim_z",
    "int_mask_bb_dim_x",
    "int_mask_bb_dim_y",
    "int_mask_bb_dim_z",
    "morph_mask_bb_dim_x",
    "morph_mask_bb_dim_y",
    "morph_mask_bb_dim_z",
    "int_mask_vox_count",
    "morph_mask_vox_count",
    "int_mask_mean_int",
    "int_mask_min_int",
    "int_mask_max_int",
)
IBSI1_DIAGNOSTIC_FAMILY_TAGS = {
    "Diagnostics-initial image": frozenset(
        f"{base}_init_img" for base in _IBSI1_DIAGNOSTIC_IMAGE_BASES
    ),
    "Diagnostics-interpolated image": frozenset(
        f"{base}_interp_img" for base in _IBSI1_DIAGNOSTIC_IMAGE_BASES
    ),
    "Diagnostics-initial ROI": frozenset(
        f"{base}_init_roi" for base in _IBSI1_DIAGNOSTIC_ROI_BASES
    ),
    "Diagnostics-interpolated ROI": frozenset(
        f"{base}_interp_roi" for base in _IBSI1_DIAGNOSTIC_ROI_BASES
    ),
    "Diagnostics-resegmented ROI": frozenset(
        f"{base}_reseg_roi" for base in _IBSI1_DIAGNOSTIC_ROI_BASES
    ),
}
IBSI1_DIAGNOSTIC_TAGS = frozenset().union(*IBSI1_DIAGNOSTIC_FAMILY_TAGS.values())
IBSI1_DIAGNOSTIC_ROWS_BY_SHEET = {
    "digital phantom": 0,
    "config A": 60,
    "config B": 60,
    "config C": 60,
    "config D": 60,
    "config E": 60,
}

IBSI_DATA_REPOSITORY = "https://github.com/theibsi/data_sets"
IBSI_DATA_COMMIT = "6da96021bc91faf4c0cb7fd7fa56a4225d2064a8"

IBSI2_REFERENCE_REPOSITORY = "https://github.com/theibsi/ibsi_2_reference_data"
IBSI2_REFERENCE_COMMIT = "5404579fb3e0d17e8db421f0e82d64ce2432ec03"
IBSI2_REFERENCE_README_SHA256 = (
    "7cd0c53d69ffd405fa72fa31f364e373687845431aa67a377239b5727f65cadc"
)
IBSI2_ANALYSIS_REPOSITORY = "https://github.com/theibsi/ibsi_2_data_analysis"
IBSI2_ANALYSIS_COMMIT = "fde70ca61819799f673c07dca776deea065d2a7d"
IBSI2_PHASE1_COMPARISON_SOURCE = "README.md"
IBSI2_PHASE1_COMPARISON_SOURCE_SHA256 = IBSI2_REFERENCE_README_SHA256
IBSI2_PHASE1_COMPARISON_RULE = (
    "every voxel absolute error <= 1% of the reference-map intensity range"
)
IBSI2_PHASE2_ANALYSIS_SOURCE = "results/analysis_with_tolerance.R"
IBSI2_PHASE2_ANALYSIS_SOURCE_SHA256 = (
    "fa1d68823adcc91adc7713015a964fab8464cc24221622f7d6c04f0cb9a840f7"
)
IBSI2_PHASE2_REFERENCE_PATH = "reference_feature_values/reference_values.csv"
IBSI2_PHASE2_REFERENCE_SHA256 = (
    "6d1a4e7b297e5e9d79ce54471e81f6f84e202fafe0322df1258ac834ce95dd45"
)
IBSI2_PHASE2_PROFILE = "phase2_defined_filters"

IBSI2_PHASE2_TAGS = (
    "stat_mean",
    "stat_var",
    "stat_skew",
    "stat_kurt",
    "stat_median",
    "stat_min",
    "stat_p10",
    "stat_p90",
    "stat_max",
    "stat_iqr",
    "stat_range",
    "stat_mad",
    "stat_rmad",
    "stat_medad",
    "stat_cov",
    "stat_qcod",
    "stat_energy",
    "stat_rms",
)
IBSI2_PHASE2_FEATURE_NAMES = {
    "stat_mean": "Mean",
    "stat_var": "Variance",
    "stat_skew": "Skewness",
    "stat_kurt": "(Excess) kurtosis",
    "stat_median": "Median",
    "stat_min": "Minimum",
    "stat_p10": "10th percentile",
    "stat_p90": "90th percentile",
    "stat_max": "Maximum",
    "stat_iqr": "Interquartile range",
    "stat_range": "Range",
    "stat_mad": "Mean absolute deviation",
    "stat_rmad": "Robust mean absolute deviation",
    "stat_medad": "Median absolute deviation",
    "stat_cov": "Coefficient of variation",
    "stat_qcod": "Quartile coefficient of dispersion",
    "stat_energy": "Energy",
    "stat_rms": "Root mean square",
}
IBSI2_PHASE2_PUBLISHED_FILTER_IDS = tuple(
    f"{number}.{dimension}" for number in range(1, 10) for dimension in ("A", "B")
)
IBSI2_PHASE2_DEFINED_FILTER_IDS = tuple(
    f"{number}.{dimension}" for number in range(1, 12) for dimension in ("A", "B")
)
# Full protocol surface.  Import of the official 324-row reference CSV uses the
# narrower IBSI2_PHASE2_PUBLISHED_FILTER_IDS constant explicitly.
IBSI2_PHASE2_FILTER_IDS = IBSI2_PHASE2_DEFINED_FILTER_IDS
IBSI2_PHASE2_DEFINED_CHECKS = 396
IBSI2_PHASE2_PUBLISHED_REFERENCE_ROWS = 324
IBSI2_PHASE2_STANDARDIZED_CHECKS = 323
IBSI2_PHASE2_NONSTANDARDIZED_CHECKS = 73
IBSI2_PHASE2_NONSTANDARDIZED_FILTER_IDS = ("10.A", "10.B", "11.A", "11.B")
IBSI2_PHASE2_NONSTANDARDIZED_PUBLISHED_PAIR = ("8.B", "stat_qcod")

IBSI2_PHASE2_SOURCE_IMAGE_SHA256 = (
    "909129fdf99d5f5fd17bc568b1834d31edce8663ea6746c6020720bfc1d7bbcb"
)
IBSI2_PHASE2_SOURCE_MASK_SHA256 = (
    "0e9243e473b1255c60e6460df465d0b07b65436d6e0c691fe3d17bca0594ae78"
)
IBSI2_PHASE1_SOURCE_IMAGE_SHA256 = frozenset(
    {
        "20feaef6e6b5970dc4ce9979a00e1fe41d5017bd2f044d1ad2eec4f74f507ccc",
        "3ba64e8a782127c2595f2f74a59820d414cf525c1f942380de587806a75ed464",
        "3f33f74dff6c9d5a58d38f999074f6b55732ea859dcdebf93b54aa440e04b27b",
        "55cb8915c22178a420904e94a74f7fca94f3a887095957ef679809bfdff7a23e",
        "60ce42614abe68acd16a2d6f853218a62666fed580ecbb49a8c3a15c914b7bb0",
        "8a8314fdcf41a55f4de3f8e7d0f78a09563ffd4d11f20d73995e8b13369a78bd",
        "ab6136aa65c6378fc680f3a3dffc9f2ac12950635b3ad8d15aa3a4896f745064",
        "e0f961ec142b597569aad1d46717fd3d8ae17b100330278a819b4e012357d3b6",
        "e2ec165e237a9fb8e017c5ef7c5472fa82d10da1433d4bb296762916f8e7570f",
    }
)
IBSI2_PHASE1_SOURCE_MASK_SHA256 = (
    "f169b2fced49fe23eee6c09914d9ad07e0aba835b46f2c74caa99a609996fc16"
)

IBSI2_PHASE1_TEST_IDS = (
    "1.a.1",
    "1.a.2",
    "1.a.3",
    "1.a.4",
    "1.b.1",
    "2.a",
    "2.b",
    "2.c",
    "3.a.1",
    "3.a.2",
    "3.a.3",
    "3.b.1",
    "3.b.2",
    "3.b.3",
    "3.c.1",
    "3.c.2",
    "3.c.3",
    "4.a.1",
    "4.a.2",
    "4.b.1",
    "4.b.2",
    "5.a.1",
    "5.a.2",
    "6.a.1",
    "6.a.2",
    "7.a.1",
    "7.a.2",
    "8.a.1",
    "8.a.2",
    "8.a.3",
    "9.a",
    "9.b.1",
    "9.b.2",
    "10.a",
    "10.b.1",
    "10.b.2",
)
IBSI2_PHASE1_NONSTANDARDIZED_IDS = ("9.b.2", "10.a", "10.b.2")
IBSI2_PHASE1_STANDARDIZED_IDS = tuple(
    value
    for value in IBSI2_PHASE1_TEST_IDS
    if value not in IBSI2_PHASE1_NONSTANDARDIZED_IDS
)

IBSI2_PHASE1_REFERENCE_SHA256 = {
    "1.a.1": "4d45bd79271d0b4034f248690dee2fb60990fb436f29b0edb76c67ea65dedda3",
    "1.a.2": "d32e92e7fd1d8833d63aa4cf516bd2f59bddf704ccd6e1f177ad2d0b06a93e22",
    "1.a.3": "8c6d6c09e2f0cb3711ebd6d5860a1f29c34197900d479e2e09d967765e60c95c",
    "1.a.4": "5903a490acf4b85601f0af64d336415e61f72ca0992c107941591017de92ba74",
    "1.b.1": "c483968d7e4a1db017497825e380342ea0433208b6ba30cb8233442ed33770ab",
    "2.a": "2b09b838cecbf9324d7ef5978ec949d5a18c33d40cb4a8044a674fe24ce2e065",
    "2.b": "654458872fd3f6574b6638820fe88593fd78b644305cf68ddcf2b3e3e4e13f51",
    "2.c": "4c0b340261f6bdb5741843a0d09dca9605509a06453c21b03a9dbd65ddf1fe98",
    "3.a.1": "653de66c165d449d621540b0c3091ded89585f8cc52606b99a7b63292f94eae2",
    "3.a.2": "e9ac127ae56a71ecf3f6c9bc2a41c4effafb10fbea2dfbc6884dbffe1aa6cc28",
    "3.a.3": "cf0569fa94ae5d1babeeb570cac7e5baa76a7894dec60c8c9c78be27bae67106",
    "3.b.1": "b0733aa5eeff197e84a61dbdd3ecce86a2615688a1fa63930bdc235ce3f80e3c",
    "3.b.2": "43318d37691a69cdf4e0995ddf34da245347bb39524fe0a3658c44aa61736580",
    "3.b.3": "ad11703127a2f173aa12e1bbae7270d794d3f2483932ced8ef8a4f1ca2e128f2",
    "3.c.1": "3983e7ba9b20ead5cba46ac791445c777d62c9439740b2d0c226496210e9c1e9",
    "3.c.2": "b4dcd5c18c149eef3076c40905e006373c19893f41c018d7f0e164886f3611c9",
    "3.c.3": "63d630767358892b1dd15fdcbec55e20e8ca4608911590d5da3b3aeeda98e199",
    "4.a.1": "72636c8b8873e1b184e7ddff14fd381781503cb63b69d57567ca77785993e8a9",
    "4.a.2": "35d04398515d654a99fca5a616018a67b9265ee614da0ded490e67df86a3eadb",
    "4.b.1": "c516a4ac1ce4b844285afff97785a3571af2ae543b44f51e9403cf71e0d85fef",
    "4.b.2": "2feab18d9958e760363c341126a4b8fa4b1b48abf18d6fefd2897f6114bd4f43",
    "5.a.1": "28ca3830658c09be1fda6fd0ce0d658a7511d1ad09c9fadfd37405ce5f3ad512",
    "5.a.2": "2f06f786d19ebe49f067f7e035212acb7bfb893022a92d79658ff982d141e33f",
    "6.a.1": "c32f12b1839f4f658325f5b5cb05363495efb0bbcefa5dbc108f06b2960e5629",
    "6.a.2": "5e35d206b3b4b9033451cc978330dc5f0c38ff87f0cb6dc32a57092411019026",
    "7.a.1": "849d625044a3f2291128d44ce72faa9ef4dc1980053064b36b72a4e32c644eba",
    "7.a.2": "b58b6f4166bebd2091d2adfb453af12e5c925467de778fa12a2eb60ea392a724",
    "8.a.1": "f26254a1dac5b51a50849b11b53ce4fd2d8a8c61228627acce06740094ccaad5",
    "8.a.2": "e5df8a75c68807935652a70b92f5c616508216cb8cb3f0df4ced6158db20e7e2",
    "8.a.3": "0affd2ee328882cf7501b6d79ac173cf5811b7fb09308bf0c4b8315851f7a75a",
    "9.a": "9ef1621d397e12a4d16079b7bb83658a1b2fb1da4706047c9c50c043daddf62b",
    "9.b.1": "fb6df2cd53d5fe9ec03756db15c27c7aaba3dec344210625c7365d6cd7747c6d",
    "10.b.1": "b02016be9a851a79a2ebc6c697781ec6ae73b54eca990f6c4d3facfc85da6726",
}

_REFERENCE_FIELDS = tuple(ReferenceRecord.__dataclass_fields__)


class ReferenceValidationError(ValueError):
    pass


def _finite_optional(value: Any) -> Optional[float]:
    if value is None or str(value).strip() == "":
        return None
    try:
        result = float(value)
    except (TypeError, ValueError, OverflowError) as exc:
        raise ReferenceValidationError(
            f"Invalid numeric reference value: {value!r}"
        ) from exc
    if not math.isfinite(result):
        raise ReferenceValidationError(f"Reference value must be finite: {value!r}")
    return result


def _canonical_configuration(sheet: str, dataset: str) -> str:
    if sheet == "digital phantom":
        return "digital_phantom"
    match = re.search(r"([A-E])$", str(dataset).strip(), flags=re.IGNORECASE)
    if match:
        return match.group(1).upper()
    return sheet.removeprefix("config ").strip().upper()


def _aggregation(family_label: str) -> str:
    label = family_label.casefold()
    patterns = (
        ("(2d, averaged)", "2d_average"),
        ("(2d, slice-merged)", "2d_slice_merge"),
        ("(2.5d, direction-merged)", "2.5d_direction_merge"),
        ("(2.5d, merged)", "2.5d_merge"),
        ("(3d, averaged)", "3d_average"),
        ("(3d, merged)", "3d_merge"),
        ("(2.5d)", "2.5d"),
        ("(2d)", "2d"),
        ("(3d)", "3d"),
    )
    for marker, value in patterns:
        if marker in label:
            return value
    return "not_applicable"


_SEMANTIC_NAMESPACES = {
    "morphology": "morphology",
    "local_intensity": "local_intensity",
    "intensity": "intensity_statistics",
    "histogram": "intensity_histogram",
    "ivh": "ivh",
    "glcm": "glcm",
    "glrlm": "glrlm",
    "glszm": "glszm",
    "gldzm": "gldzm",
    "ngtdm": "ngtdm",
    "ngldm": "ngldm",
}
_SEMANTIC_PREFIXES = {
    "morphology": "morph_",
    "local_intensity": "loc_",
    "intensity": "stat_",
    "histogram": "ih_",
    "ivh": "ivh_",
    "glcm": "cm_",
    "glrlm": "rlm_",
    "glszm": "szm_",
    "gldzm": "dzm_",
    "ngtdm": "ngt_",
    "ngldm": "ngl_",
}
_SEMANTIC_OVERRIDES = {
    "ih_p10": "intensity_histogram.percentile_10",
    "ih_p90": "intensity_histogram.percentile_90",
    "szm_sze": "glszm.small_zone_emphasis",
    "cm_sum_entr": "glcm.sum_entropy",
    "ivh_v10": "ivh.volume_at_intensity_fraction_10",
    "ivh_v90": "ivh.volume_at_intensity_fraction_90",
    "ivh_i10": "ivh.intensity_at_volume_fraction_10",
    "ivh_i90": "ivh.intensity_at_volume_fraction_90",
    "ivh_diff_v10_v90": "ivh.volume_fraction_difference_10_90",
    "ivh_diff_i10_i90": "ivh.intensity_fraction_difference_10_90",
}


def semantic_key_from_tag(feature_tag: str, family: str) -> str:
    """Create a durable semantic join key independent of reused external IDs."""

    if family == "diagnostics":
        return "diagnostics." + re.sub(
            r"[^a-z0-9]+", "_", feature_tag.casefold()
        ).strip("_")
    base = re.sub(
        r"_(?:2D|2_5D|3D)(?:_(?:avg|comb))?$",
        "",
        feature_tag,
        flags=re.IGNORECASE,
    )
    if base in _SEMANTIC_OVERRIDES:
        return _SEMANTIC_OVERRIDES[base]
    prefix = _SEMANTIC_PREFIXES[family]
    token = base[len(prefix) :] if base.startswith(prefix) else base
    token = re.sub(r"[^a-z0-9]+", "_", token.casefold()).strip("_")
    return f"{_SEMANTIC_NAMESPACES[family]}.{token}"


def ibsi1_table2_definition_key(semantic_key: str) -> str:
    """Collapse workbook operating points to their IBSI Table 2 definition.

    Only two IVH definitions are parameterised at two operating points in the
    official workbook.  Every other benchmark instance is one-to-one with a
    Table 2 definition.
    """

    if semantic_key in {
        "ivh.volume_at_intensity_fraction_10",
        "ivh.volume_at_intensity_fraction_90",
    }:
        return "ivh.volume_at_intensity_fraction"
    if semantic_key in {
        "ivh.intensity_at_volume_fraction_10",
        "ivh.intensity_at_volume_fraction_90",
    }:
        return "ivh.intensity_at_volume_fraction"
    return semantic_key


def _code_semantic_keys() -> dict[str, str]:
    """Resolve the adapter crosswalk to one durable semantic key per feature.

    Adapter-native names are classified through the source code table. That
    external identifier is retained for audit output only; this derived mapping
    immediately converts it to the semantic identity used for all joins.
    Construction fails if a code would identify two
    different features.
    """

    by_code: dict[str, set[str]] = defaultdict(set)
    for feature_tag in MIRP_PREFIX_MAP:
        code = map_mirp(feature_tag)
        family = CODE_TO_FAMILY.get(code or "")
        if code and family:
            by_code[code].add(semantic_key_from_tag(feature_tag, family))
    collisions = {
        code: sorted(keys) for code, keys in by_code.items() if len(keys) != 1
    }
    if collisions:
        raise RuntimeError(
            "IBSI source identifiers do not resolve uniquely to semantic keys: "
            + json.dumps(collisions, sort_keys=True)
        )
    resolved = {code: next(iter(keys)) for code, keys in by_code.items()}
    if len(resolved) != len(MIRP_PREFIX_MAP):
        raise RuntimeError(
            "The reviewed IBSI 1 crosswalk must resolve every one of the "
            f"{len(MIRP_PREFIX_MAP)} semantic features to a unique external identifier"
        )
    return resolved


CODE_TO_SEMANTIC_KEY = _code_semantic_keys()


def _in_ibsi1_profile(configuration: str, family: str, aggregation: str) -> bool:
    if configuration != "digital_phantom" or family == "diagnostics":
        return False
    if family in {"glcm", "glrlm"}:
        return aggregation == "3d_merge"
    if family in {"glszm", "gldzm", "ngtdm", "ngldm"}:
        return aggregation == "3d"
    return aggregation == "not_applicable"


def _record_csv(records: Iterable[ReferenceRecord]) -> str:
    stream = io.StringIO(newline="")
    writer = csv.DictWriter(stream, fieldnames=_REFERENCE_FIELDS, lineterminator="\n")
    writer.writeheader()
    for record in records:
        row = record.to_dict()
        row["reference_value"] = (
            "" if record.reference_value is None else repr(record.reference_value)
        )
        row["tolerance"] = "" if record.tolerance is None else repr(record.tolerance)
        writer.writerow(row)
    return stream.getvalue()


def load_reference_csv(path: Path) -> list[ReferenceRecord]:
    with Path(path).open("r", encoding="utf-8", newline="") as stream:
        rows = []
        for row_number, row in enumerate(csv.DictReader(stream), start=2):
            try:
                record = ReferenceRecord.from_mapping(row)
            except (TypeError, ValueError, OverflowError) as exc:
                raise ReferenceValidationError(
                    f"Malformed reference record at CSV row {row_number}"
                ) from exc
            numeric_pair = (
                record.reference_value is not None,
                record.tolerance is not None,
            )
            if numeric_pair[0] != numeric_pair[1]:
                raise ReferenceValidationError(
                    f"Partial reference/tolerance pair at CSV row {row_number}"
                )
            if record.reference_value is not None and not math.isfinite(
                record.reference_value
            ):
                raise ReferenceValidationError(
                    f"Non-finite reference value at CSV row {row_number}"
                )
            if record.tolerance is not None and (
                not math.isfinite(record.tolerance) or record.tolerance < 0.0
            ):
                raise ReferenceValidationError(
                    f"Invalid tolerance at CSV row {row_number}"
                )
            if record.standardized != all(numeric_pair):
                raise ReferenceValidationError(
                    f"Standardization flag disagrees with numerical reference at CSV row {row_number}"
                )
            if any(
                not value
                for value in (
                    record.specification,
                    record.configuration,
                    record.aggregation,
                    record.family,
                    record.feature_tag,
                    record.semantic_key,
                    record.ibsi_code,
                )
            ):
                raise ReferenceValidationError(
                    f"Reference identity is incomplete at CSV row {row_number}"
                )
            rows.append(record)
    if not rows:
        raise ReferenceValidationError(f"Reference table is empty: {path}")
    if len({record.key for record in rows}) != len(rows):
        raise ReferenceValidationError(
            "Reference table contains duplicate configuration-aware keys"
        )
    return rows


def _validate_reviewed_table_structure(
    records: list[ReferenceRecord],
    *,
    specification: str,
    phase: Optional[str],
) -> None:
    """Recheck canonical table invariants instead of trusting manifest counters."""

    if any(record.specification != specification for record in records):
        raise ReferenceValidationError(
            "Reference table contains rows from a different specification"
        )
    if specification == "IBSI 1":
        expected_total = sum(rows for rows, _ in IBSI1_WORKBOOK_SHEETS.values())
        if len(records) != expected_total:
            raise ReferenceValidationError(
                f"Reviewed IBSI 1 table requires {expected_total} rows, got {len(records)}"
            )
        observed = Counter(record.source_sheet for record in records)
        observed_standardized = Counter(
            record.source_sheet for record in records if record.standardized
        )
        for sheet, (row_count, standardized_count) in IBSI1_WORKBOOK_SHEETS.items():
            if (
                observed[sheet] != row_count
                or observed_standardized[sheet] != standardized_count
            ):
                raise ReferenceValidationError(
                    f"Reviewed IBSI 1 table has invalid counts for {sheet!r}"
                )
        for sheet, expected_count in IBSI1_DIAGNOSTIC_ROWS_BY_SHEET.items():
            diagnostics = [
                record
                for record in records
                if record.source_sheet == sheet and record.family == "diagnostics"
            ]
            tag_counts = Counter(record.feature_tag for record in diagnostics)
            expected_tags = IBSI1_DIAGNOSTIC_TAGS if expected_count else frozenset()
            if (
                len(diagnostics) != expected_count
                or set(tag_counts) != set(expected_tags)
                or any(count != 1 for count in tag_counts.values())
            ):
                raise ReferenceValidationError(
                    f"Reviewed IBSI 1 table has an invalid diagnostic surface for {sheet!r}"
                )
            if any(
                not record.standardized
                or record.aggregation != "not_applicable"
                or record.in_profile
                or record.profile
                or record.semantic_key
                != semantic_key_from_tag(record.feature_tag, "diagnostics")
                or record.ibsi_code != f"diagnostic:{record.feature_tag}"
                for record in diagnostics
            ):
                raise ReferenceValidationError(
                    f"Reviewed IBSI 1 table has malformed diagnostics for {sheet!r}"
                )
        for record in records:
            if record.family == "diagnostics":
                continue
            code = map_mirp(record.feature_tag) or ""
            family = CODE_TO_FAMILY.get(code)
            if (
                not code
                or family != record.family
                or code != record.ibsi_code
                or record.semantic_key
                != semantic_key_from_tag(record.feature_tag, record.family)
            ):
                raise ReferenceValidationError(
                    "Reviewed IBSI 1 table contains an unmapped radiomic feature row"
                )
        select_ibsi1_digital_phantom_profile(records)
        for configuration, aggregations in IBSI1_CONFIGURATION_AGGREGATIONS.items():
            for aggregation in aggregations:
                select_ibsi1_configuration_profile(
                    records,
                    configuration=configuration,
                    directional_aggregation=aggregation,
                )
        return

    if specification == "IBSI 2" and phase == "phase2":
        standardized_count = sum(record.standardized for record in records)
        if (
            len(records) != IBSI2_PHASE2_DEFINED_CHECKS
            or standardized_count != IBSI2_PHASE2_STANDARDIZED_CHECKS
        ):
            raise ReferenceValidationError(
                "Reviewed IBSI 2 Phase 2 table requires 396 defined checks and "
                "323 standardized references"
            )
        if any(record.phase != "phase2" for record in records):
            raise ReferenceValidationError(
                "IBSI 2 Phase 2 table contains another phase"
            )
        expected_pairs = {
            (filter_id, tag)
            for filter_id in IBSI2_PHASE2_FILTER_IDS
            for tag in IBSI2_PHASE2_TAGS
        }
        observed_pairs = Counter(
            (record.configuration, record.feature_tag) for record in records
        )
        if set(observed_pairs) != expected_pairs or any(
            count != 1 for count in observed_pairs.values()
        ):
            raise ReferenceValidationError(
                "Reviewed IBSI 2 Phase 2 filter/statistic grid is incomplete or duplicated"
            )
        if any(
            not record.semantic_key
            or not record.ibsi_code
            or record.aggregation != "not_applicable"
            for record in records
        ):
            raise ReferenceValidationError(
                "Reviewed IBSI 2 Phase 2 table contains an unmapped statistic or "
                "incorrect feature-aggregation label"
            )
        semantic_grid = Counter(
            (record.configuration, record.semantic_key) for record in records
        )
        if len(semantic_grid) != IBSI2_PHASE2_DEFINED_CHECKS or any(
            count != 1 for count in semantic_grid.values()
        ):
            raise ReferenceValidationError(
                "Reviewed IBSI 2 Phase 2 semantic feature grid is not one-to-one"
            )
        nonstandardized_pairs = {
            (record.configuration, record.feature_tag)
            for record in records
            if not record.standardized
        }
        expected_nonstandardized = {
            (filter_id, tag)
            for filter_id in IBSI2_PHASE2_NONSTANDARDIZED_FILTER_IDS
            for tag in IBSI2_PHASE2_TAGS
        }
        expected_nonstandardized.add(IBSI2_PHASE2_NONSTANDARDIZED_PUBLISHED_PAIR)
        if nonstandardized_pairs != expected_nonstandardized:
            raise ReferenceValidationError(
                "Reviewed IBSI 2 Phase 2 table has the wrong nonstandardized cells"
            )
        return

    raise ReferenceValidationError(
        f"No reviewed table structure is defined for {specification} {phase or ''}".strip()
    )


def validate_reference_table_manifest(
    references_csv: Path,
    manifest_path: Path,
    *,
    expected_specification: str,
    expected_phase: Optional[str] = None,
    require_reviewed_source: bool = True,
) -> dict[str, Any]:
    """Bind a canonical reference CSV to its checksummed import manifest."""

    table = Path(references_csv).expanduser().resolve()
    manifest_file = Path(manifest_path).expanduser().resolve()
    with manifest_file.open("r", encoding="utf-8") as stream:
        manifest = json.load(stream)
    if (
        not isinstance(manifest, dict)
        or manifest.get("schema_version") != REFERENCE_SCHEMA_VERSION
    ):
        raise ReferenceValidationError("Invalid compliance reference manifest schema")
    if manifest.get("specification") != expected_specification:
        raise ReferenceValidationError(
            f"Reference manifest is for {manifest.get('specification')!r}, "
            f"expected {expected_specification!r}"
        )
    if expected_phase is not None and manifest.get("phase") != expected_phase:
        raise ReferenceValidationError(
            f"Reference manifest phase is {manifest.get('phase')!r}, expected {expected_phase!r}"
        )
    table_record = manifest.get("reference_table")
    if not isinstance(table_record, dict):
        raise ReferenceValidationError(
            "Reference manifest lacks reference_table metadata"
        )
    listed_table = (manifest_file.parent / str(table_record.get("path", ""))).resolve()
    if manifest_file.parent not in listed_table.parents or listed_table != table:
        raise ReferenceValidationError(
            "Reference CSV must be the exact table named by its import manifest"
        )
    expected_hash = str(table_record.get("sha256", ""))
    if not table.is_file() or not expected_hash or sha256_file(table) != expected_hash:
        raise ReferenceValidationError(
            "Reference CSV checksum does not match its manifest"
        )

    if require_reviewed_source and expected_specification == "IBSI 1":
        source = manifest.get("source", {})
        if not isinstance(source, dict) or (
            source.get("url") != IBSI1_WORKBOOK_URL
            or source.get("sha256") != IBSI1_WORKBOOK_SHA256
            or source.get("known_reviewed_sha256") != IBSI1_WORKBOOK_SHA256
            or source.get("hash_verified") is not True
        ):
            raise ReferenceValidationError(
                "IBSI 1 execution requires the reviewed official workbook hash"
            )
    if require_reviewed_source and expected_specification == "IBSI 2":
        source = manifest.get("source", {})
        if not isinstance(source, dict) or (
            source.get("reference_repository") != IBSI2_REFERENCE_REPOSITORY
            or source.get("reference_commit") != IBSI2_REFERENCE_COMMIT
            or source.get("reference_path") != IBSI2_PHASE2_REFERENCE_PATH
            or source.get("sha256") != IBSI2_PHASE2_REFERENCE_SHA256
            or source.get("known_reviewed_sha256") != IBSI2_PHASE2_REFERENCE_SHA256
            or source.get("hash_verified") is not True
            or source.get("analysis_repository") != IBSI2_ANALYSIS_REPOSITORY
            or source.get("analysis_commit") != IBSI2_ANALYSIS_COMMIT
            or source.get("generator") != IBSI2_PHASE2_ANALYSIS_SOURCE
            or source.get("generator_sha256") != IBSI2_PHASE2_ANALYSIS_SOURCE_SHA256
        ):
            raise ReferenceValidationError(
                "IBSI 2 Phase 2 execution requires the reviewed official reference "
                "table and its pinned analysis lineage"
            )
    if require_reviewed_source:
        records = load_reference_csv(table)
        _validate_reviewed_table_structure(
            records,
            specification=expected_specification,
            phase=expected_phase,
        )
    return manifest


def select_ibsi1_digital_phantom_profile(
    records: Iterable[ReferenceRecord],
    *,
    directional_aggregation: str = IBSI1_REQUIRED_DIRECTIONAL_AGGREGATION,
) -> list[ReferenceRecord]:
    """Select the required 3D-merged digital-phantom benchmark profile."""

    if directional_aggregation != IBSI1_REQUIRED_DIRECTIONAL_AGGREGATION:
        raise ValueError("The Pictologics IBSI 1 profile requires 3d_merge")
    selected: list[ReferenceRecord] = []
    for record in records:
        if (
            record.specification != "IBSI 1"
            or record.configuration != "digital_phantom"
        ):
            continue
        include = False
        if record.family in {"glcm", "glrlm"}:
            include = record.aggregation == directional_aggregation
        elif record.family in {"glszm", "gldzm", "ngtdm", "ngldm"}:
            include = record.aggregation == "3d"
        elif record.family != "diagnostics":
            include = record.aggregation == "not_applicable"
        if include:
            selected.append(
                replace(
                    record,
                    profile=IBSI1_PROFILE,
                    in_profile=True,
                )
            )
    standardized = [record for record in selected if record.standardized]
    if (
        len(selected) != IBSI1_PROFILE_ROWS
        or len(standardized) != IBSI1_PROFILE_STANDARDIZED
    ):
        raise ReferenceValidationError(
            f"Selected IBSI 1 profile has {len(selected)}/{len(standardized)} rows, expected "
            f"{IBSI1_PROFILE_ROWS}/{IBSI1_PROFILE_STANDARDIZED}"
        )
    if len({record.semantic_key for record in selected}) != IBSI1_PROFILE_ROWS:
        raise ReferenceValidationError(
            "Selected IBSI 1 profile does not contain 174 unique workbook instances"
        )
    definition_counts = Counter(
        (record.family, ibsi1_table2_definition_key(record.semantic_key))
        for record in selected
    )
    if len(definition_counts) != IBSI1_TABLE2_DEFINITION_COUNT:
        raise ReferenceValidationError(
            "Selected IBSI 1 profile does not collapse to 172 Table 2 definitions"
        )
    observed_by_family = Counter(family for family, _ in definition_counts)
    if dict(observed_by_family) != IBSI1_TABLE2_FAMILY_DEFINITIONS:
        raise ReferenceValidationError(
            "Selected IBSI 1 profile does not match Table 2 family denominators"
        )
    return selected


def select_ibsi1_configuration_profile(
    records: Iterable[ReferenceRecord],
    *,
    configuration: str,
    directional_aggregation: str,
) -> list[ReferenceRecord]:
    """Select the exact feature rows for one official IBSI 1 CT configuration."""

    config = str(configuration).strip().upper()
    if config not in IBSI1_CONFIGURATION_PROFILE_COUNTS:
        raise ValueError(f"IBSI 1 configuration must be A-E, got {configuration!r}")
    allowed_aggregations = IBSI1_CONFIGURATION_AGGREGATIONS[config]
    if directional_aggregation not in allowed_aggregations:
        raise ValueError(
            f"IBSI 1 configuration {config} requires one of: "
            + ", ".join(allowed_aggregations)
        )
    nondirectional_aggregation = (
        "2.5d"
        if directional_aggregation.startswith("2.5d_")
        else "2d"
        if directional_aggregation.startswith("2d_")
        else "3d"
    )
    selected: list[ReferenceRecord] = []
    for record in records:
        if record.specification != "IBSI 1" or record.configuration != config:
            continue
        include = False
        if record.family in {"glcm", "glrlm"}:
            include = record.aggregation == directional_aggregation
        elif record.family in {"glszm", "gldzm", "ngtdm", "ngldm"}:
            include = record.aggregation == nondirectional_aggregation
        elif record.family != "diagnostics":
            include = record.aggregation == "not_applicable"
        if include:
            selected.append(
                replace(
                    record,
                    profile=f"configuration_{config}_{directional_aggregation}",
                    in_profile=True,
                )
            )
    expected_rows, expected_standardized = IBSI1_CONFIGURATION_PROFILE_COUNTS[config]
    standardized = [record for record in selected if record.standardized]
    if (len(selected), len(standardized)) != (expected_rows, expected_standardized):
        raise ReferenceValidationError(
            f"IBSI 1 configuration {config} profile has {len(selected)}/{len(standardized)}, "
            f"expected {expected_rows}/{expected_standardized}"
        )
    if len({record.semantic_key for record in selected}) != expected_rows:
        raise ReferenceValidationError(
            f"IBSI 1 configuration {config} does not contain {expected_rows} "
            "unique workbook instances"
        )
    return selected


def import_ibsi1_workbook(
    workbook_path: Path,
    output_dir: Path,
    *,
    require_known_hash: bool = True,
) -> dict[str, Any]:
    """Import every row of the official six-sheet workbook without collapsing aggregations."""

    path = Path(workbook_path).expanduser().resolve()
    digest = sha256_file(path)
    if require_known_hash and digest != IBSI1_WORKBOOK_SHA256:
        raise ReferenceValidationError(
            "IBSI 1 workbook hash does not match the reviewed official artifact: "
            + digest
        )
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to import the IBSI 1 workbook"
        ) from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if set(workbook.sheetnames) != set(IBSI1_WORKBOOK_SHEETS):
        raise ReferenceValidationError(
            f"Unexpected IBSI 1 workbook sheets: {workbook.sheetnames!r}"
        )

    records: list[ReferenceRecord] = []
    counts: dict[str, dict[str, int]] = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        header = tuple(
            cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))
        )
        expected_header = (
            "dataset",
            "family",
            "feature",
            "consensus",
            "reference value",
            "tolerance",
            "your result",
            "difference",
            "check",
            "tag",
        )
        if header != expected_header:
            raise ReferenceValidationError(
                f"Unexpected columns in sheet {sheet_name!r}: {header!r}"
            )

        sheet_rows = 0
        standardized_rows = 0
        diagnostic_rows = 0
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            if all(value is None for value in values):
                continue
            if len(values) != 10:
                raise ReferenceValidationError(
                    f"Sheet {sheet_name!r} row {row_number} has {len(values)} cells"
                )
            (
                dataset,
                family_label,
                feature_name,
                consensus,
                reference,
                tolerance,
                _,
                _,
                _,
                tag,
            ) = values
            feature_tag = str(tag or "").strip()
            if not feature_tag:
                raise ReferenceValidationError(
                    f"Missing feature tag at {sheet_name}!{row_number}"
                )
            reference_value = _finite_optional(reference)
            tolerance_value = _finite_optional(tolerance)
            standardized = reference_value is not None and tolerance_value is not None
            if (reference_value is None) != (tolerance_value is None):
                raise ReferenceValidationError(
                    f"Partial reference/tolerance pair at {sheet_name}!{row_number}"
                )
            if tolerance_value is not None and tolerance_value < 0.0:
                raise ReferenceValidationError(
                    f"Negative tolerance at {sheet_name}!{row_number}"
                )

            family_label_text = str(family_label or "").strip()
            diagnostic_tags = IBSI1_DIAGNOSTIC_FAMILY_TAGS.get(family_label_text)
            if diagnostic_tags is not None:
                if (
                    sheet_name == "digital phantom"
                    or feature_tag not in diagnostic_tags
                ):
                    raise ReferenceValidationError(
                        f"Invalid IBSI 1 diagnostic tag/family at "
                        f"{sheet_name}!{row_number}: {family_label_text!r}, {feature_tag!r}"
                    )
                code = f"diagnostic:{feature_tag}"
                family = "diagnostics"
                diagnostic_rows += 1
            else:
                if family_label_text.casefold().startswith("diagnostics"):
                    raise ReferenceValidationError(
                        f"Unknown IBSI 1 diagnostic family at "
                        f"{sheet_name}!{row_number}: {family_label_text!r}"
                    )
                code = map_mirp(feature_tag) or ""
                family = CODE_TO_FAMILY.get(code)
                if not code or not family:
                    raise ReferenceValidationError(
                        f"Unmapped IBSI 1 radiomic feature tag at "
                        f"{sheet_name}!{row_number}: {feature_tag}"
                    )
            configuration = _canonical_configuration(sheet_name, str(dataset or ""))
            aggregation = _aggregation(family_label_text)
            in_profile = _in_ibsi1_profile(configuration, family, aggregation)
            records.append(
                ReferenceRecord(
                    specification="IBSI 1",
                    phase="phase1" if configuration == "digital_phantom" else "phase2",
                    dataset=str(dataset or "").strip(),
                    configuration=configuration,
                    profile=IBSI1_PROFILE if in_profile else "",
                    in_profile=in_profile,
                    aggregation=aggregation,
                    family=family,
                    feature_name=str(feature_name or "").strip(),
                    feature_tag=feature_tag,
                    semantic_key=semantic_key_from_tag(feature_tag, family),
                    ibsi_code=code,
                    consensus=str(consensus or "").strip(),
                    reference_value=reference_value,
                    tolerance=tolerance_value,
                    standardized=standardized,
                    source_sheet=sheet_name,
                    source_row=row_number,
                )
            )
            sheet_rows += 1
            standardized_rows += int(standardized)

        expected_rows, expected_standardized = IBSI1_WORKBOOK_SHEETS[sheet_name]
        if (sheet_rows, standardized_rows) != (expected_rows, expected_standardized):
            raise ReferenceValidationError(
                f"Sheet {sheet_name!r} has {(sheet_rows, standardized_rows)}, expected "
                f"{(expected_rows, expected_standardized)} rows/standardized rows"
            )
        if digest == IBSI1_WORKBOOK_SHA256:
            expected_diagnostics = IBSI1_DIAGNOSTIC_ROWS_BY_SHEET[sheet_name]
            if diagnostic_rows != expected_diagnostics:
                raise ReferenceValidationError(
                    f"Sheet {sheet_name!r} has {diagnostic_rows} diagnostic rows, "
                    f"expected {expected_diagnostics}"
                )
        counts[sheet_name] = {
            "rows": sheet_rows,
            "standardized": standardized_rows,
            "diagnostics": diagnostic_rows,
            "radiomic_features": sheet_rows - diagnostic_rows,
        }
    workbook.close()

    if len({record.key for record in records}) != len(records):
        raise ReferenceValidationError(
            "IBSI 1 workbook contains duplicate configuration-aware tags"
        )
    profile = [record for record in records if record.in_profile]
    profile_standardized = [record for record in profile if record.standardized]
    if (
        len(profile) != IBSI1_PROFILE_ROWS
        or len(profile_standardized) != IBSI1_PROFILE_STANDARDIZED
    ):
        raise ReferenceValidationError(
            f"IBSI 1 canonical profile has {len(profile)}/{len(profile_standardized)} rows; "
            f"expected {IBSI1_PROFILE_ROWS}/{IBSI1_PROFILE_STANDARDIZED}"
        )
    if len({record.semantic_key for record in profile}) != IBSI1_PROFILE_ROWS:
        raise ReferenceValidationError(
            "IBSI 1 canonical profile does not contain 174 unique workbook instances"
        )

    output = Path(output_dir).expanduser().resolve()
    table_path = output / "ibsi1_references.csv"
    table_sha = atomic_write_text(table_path, _record_csv(records))
    manifest = {
        "schema_version": REFERENCE_SCHEMA_VERSION,
        "specification": "IBSI 1",
        "source": {
            "url": IBSI1_WORKBOOK_URL,
            # Do not bind a user's account name or clone location into a
            # normative reference manifest. The source URL and reviewed byte
            # hash are the identity; the basename is sufficient diagnostics.
            "path_at_import": path.name,
            "sha256": digest,
            "known_reviewed_sha256": IBSI1_WORKBOOK_SHA256,
            "hash_verified": digest == IBSI1_WORKBOOK_SHA256,
        },
        "reference_table": {"path": table_path.name, "sha256": table_sha},
        "counts_by_sheet": counts,
        "total_rows": len(records),
        "diagnostic_rows": sum(record.family == "diagnostics" for record in records),
        "radiomic_feature_rows": sum(
            record.family != "diagnostics" for record in records
        ),
        "unique_configuration_aware_tags": len(
            {(r.configuration, r.feature_tag) for r in records}
        ),
        "canonical_profile": {
            "name": IBSI1_PROFILE,
            "rows": len(profile),
            "standardized_rows": len(profile_standardized),
            "not_standardized_rows": len(profile) - len(profile_standardized),
            "aggregation_policy": (
                "scalar features plus GLCM/GLRLM 3D merged and single-matrix "
                "GLSZM/GLDZM/NGTDM/NGLDM 3D"
            ),
        },
        "tolerance_policy": "official workbook difference ROUNDDOWN rule",
    }
    manifest_path = output / "ibsi1_reference_manifest.json"
    atomic_write_json(manifest_path, manifest)
    return manifest


def import_ibsi2_phase2_csv(
    source_path: Path,
    output_dir: Path,
    *,
    require_reviewed_derived_hash: bool = True,
) -> dict[str, Any]:
    """Import the official Phase 2 table and retain the full protocol surface.

    IBSI publishes reference rows for configurations 1.A--9.B (324 rows).  The
    protocol also defines configurations 10.A--11.B, whose 72 statistic cells
    have no published consensus values.  The canonical table records all 396
    defined cells so coverage and accuracy denominators cannot be conflated.
    """

    path = Path(source_path).expanduser().resolve()
    digest = sha256_file(path)
    if require_reviewed_derived_hash and digest != IBSI2_PHASE2_REFERENCE_SHA256:
        raise ReferenceValidationError(
            "IBSI 2 Phase 2 CSV hash is not the reviewed official artifact: " + digest
        )
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        reader = csv.DictReader(stream, delimiter=";")
        expected = [
            "filter_id",
            "feature",
            "feature_tag",
            "consensus_value",
            "tolerance",
        ]
        if reader.fieldnames != expected:
            raise ReferenceValidationError(
                f"Unexpected IBSI 2 Phase 2 columns: {reader.fieldnames}"
            )
        raw_rows = list(reader)
    if len(raw_rows) != IBSI2_PHASE2_PUBLISHED_REFERENCE_ROWS:
        raise ReferenceValidationError(
            "IBSI 2 Phase 2 official CSV must contain "
            f"{IBSI2_PHASE2_PUBLISHED_REFERENCE_ROWS} rows, got {len(raw_rows)}"
        )

    pair_counts: Counter[tuple[str, str]] = Counter()
    records: list[ReferenceRecord] = []
    for row_number, row in enumerate(raw_rows, start=2):
        filter_id = str(row["filter_id"] or "").strip().upper()
        feature_tag = str(row["feature_tag"] or "").strip()
        if filter_id not in IBSI2_PHASE2_PUBLISHED_FILTER_IDS:
            raise ReferenceValidationError(
                f"Unknown IBSI 2 filter ID at row {row_number}: {filter_id}"
            )
        if feature_tag not in IBSI2_PHASE2_TAGS:
            raise ReferenceValidationError(
                f"Unknown IBSI 2 feature tag at row {row_number}: {feature_tag}"
            )
        pair_counts[(filter_id, feature_tag)] += 1
        reference = _finite_optional(row["consensus_value"])
        tolerance = _finite_optional(row["tolerance"])
        if (reference is None) != (tolerance is None):
            raise ReferenceValidationError(
                f"Partial value/tolerance pair at row {row_number}"
            )
        if tolerance is not None and tolerance < 0.0:
            raise ReferenceValidationError(f"Negative tolerance at row {row_number}")
        code = map_mirp(feature_tag) or ""
        if not code:
            raise ReferenceValidationError(
                f"Unmapped IBSI 2 feature tag: {feature_tag}"
            )
        standardized = reference is not None and tolerance is not None
        records.append(
            ReferenceRecord(
                specification="IBSI 2",
                phase="phase2",
                dataset="IBSI 2 CT radiomics phantom",
                configuration=filter_id,
                profile=IBSI2_PHASE2_PROFILE,
                in_profile=True,
                # A/B identify filter/preprocessing dimensionality, not feature
                # aggregation.  These first-order statistics use the complete
                # 3D ROI, for which matrix aggregation is not applicable.
                aggregation="not_applicable",
                family="intensity",
                feature_name=str(row["feature"] or "").strip(),
                feature_tag=feature_tag,
                semantic_key=semantic_key_from_tag(feature_tag, "intensity"),
                ibsi_code=code,
                consensus="standardized" if standardized else "not standardized",
                reference_value=reference,
                tolerance=tolerance,
                standardized=standardized,
                source_sheet=IBSI2_PHASE2_REFERENCE_PATH,
                source_row=row_number,
            )
        )

    expected_pairs = {
        (filter_id, tag)
        for filter_id in IBSI2_PHASE2_PUBLISHED_FILTER_IDS
        for tag in IBSI2_PHASE2_TAGS
    }
    if set(pair_counts) != expected_pairs or any(
        count != 1 for count in pair_counts.values()
    ):
        raise ReferenceValidationError(
            "IBSI 2 Phase 2 filter/feature grid is incomplete or duplicated"
        )
    published_standardized_count = sum(record.standardized for record in records)
    if published_standardized_count != IBSI2_PHASE2_STANDARDIZED_CHECKS:
        raise ReferenceValidationError(
            "IBSI 2 Phase 2 official CSV must contain "
            f"{IBSI2_PHASE2_STANDARDIZED_CHECKS} standardized references, got "
            f"{published_standardized_count}"
        )
    published_nonstandardized = {
        (record.configuration, record.feature_tag)
        for record in records
        if not record.standardized
    }
    if published_nonstandardized != {IBSI2_PHASE2_NONSTANDARDIZED_PUBLISHED_PAIR}:
        raise ReferenceValidationError(
            "The official Phase 2 table must have exactly one blank reference: "
            "8.B/stat_qcod"
        )
    if (
        len({(record.configuration, record.semantic_key) for record in records})
        != IBSI2_PHASE2_PUBLISHED_REFERENCE_ROWS
    ):
        raise ReferenceValidationError(
            "IBSI 2 Phase 2 semantic feature grid is not one-to-one"
        )

    for filter_id in IBSI2_PHASE2_NONSTANDARDIZED_FILTER_IDS:
        for feature_tag in IBSI2_PHASE2_TAGS:
            code = map_mirp(feature_tag) or ""
            if not code:
                raise ReferenceValidationError(
                    f"Unmapped IBSI 2 feature tag: {feature_tag}"
                )
            records.append(
                ReferenceRecord(
                    specification="IBSI 2",
                    phase="phase2",
                    dataset="IBSI 2 CT radiomics phantom",
                    configuration=filter_id,
                    profile=IBSI2_PHASE2_PROFILE,
                    in_profile=True,
                    aggregation="not_applicable",
                    family="intensity",
                    feature_name=IBSI2_PHASE2_FEATURE_NAMES[feature_tag],
                    feature_tag=feature_tag,
                    semantic_key=semantic_key_from_tag(feature_tag, "intensity"),
                    ibsi_code=code,
                    consensus="not standardized",
                    reference_value=None,
                    tolerance=None,
                    standardized=False,
                    source_sheet="IBSI 2 protocol-defined; no published reference row",
                    source_row=0,
                )
            )

    _validate_reviewed_table_structure(
        records,
        specification="IBSI 2",
        phase="phase2",
    )
    standardized_count = sum(record.standardized for record in records)

    output = Path(output_dir).expanduser().resolve()
    table_path = output / "ibsi2_phase2_references.csv"
    table_sha = atomic_write_text(table_path, _record_csv(records))
    manifest = {
        "schema_version": REFERENCE_SCHEMA_VERSION,
        "specification": "IBSI 2",
        "phase": "phase2",
        "source": {
            "reference_repository": IBSI2_REFERENCE_REPOSITORY,
            "reference_commit": IBSI2_REFERENCE_COMMIT,
            "reference_path": IBSI2_PHASE2_REFERENCE_PATH,
            "license": "CC0-1.0",
            "sha256": digest,
            "known_reviewed_sha256": IBSI2_PHASE2_REFERENCE_SHA256,
            "hash_verified": digest == IBSI2_PHASE2_REFERENCE_SHA256,
            "analysis_repository": IBSI2_ANALYSIS_REPOSITORY,
            "analysis_commit": IBSI2_ANALYSIS_COMMIT,
            "generator": IBSI2_PHASE2_ANALYSIS_SOURCE,
            "generator_sha256": IBSI2_PHASE2_ANALYSIS_SOURCE_SHA256,
        },
        "reference_table": {"path": table_path.name, "sha256": table_sha},
        "defined_checks": len(records),
        "published_reference_rows": len(raw_rows),
        "rows": len(records),
        "standardized_rows": standardized_count,
        "not_standardized_rows": len(records) - standardized_count,
        "defined_filter_configurations": len(IBSI2_PHASE2_DEFINED_FILTER_IDS),
        "published_filter_configurations": len(IBSI2_PHASE2_PUBLISHED_FILTER_IDS),
        "filter_configurations": len(IBSI2_PHASE2_DEFINED_FILTER_IDS),
        "features_per_configuration": len(IBSI2_PHASE2_TAGS),
        "nonstandardized_cells": {
            "unpublished_configurations": list(IBSI2_PHASE2_NONSTANDARDIZED_FILTER_IDS),
            "published_blank": list(IBSI2_PHASE2_NONSTANDARDIZED_PUBLISHED_PAIR),
        },
        "feature_aggregation": "not_applicable; first-order statistics over the complete 3D ROI",
        "tolerance_policy": "raw absolute error within published tolerance",
    }
    atomic_write_json(output / "ibsi2_phase2_reference_manifest.json", manifest)
    return manifest


def _phase1_id_from_name(path: Path) -> Optional[str]:
    name = path.name.casefold()
    if name.endswith(".nii.gz"):
        name = name[:-7]
    elif name.endswith(".nii"):
        name = name[:-4]
    name = re.sub(r"-validcrm$", "", name)
    name = name.replace("_", ".").replace("-", ".")
    name = re.sub(r"[^0-9a-z.]+", ".", name)
    name = re.sub(r"\.+", ".", name).strip(".")
    for test_id in sorted(IBSI2_PHASE1_TEST_IDS, key=len, reverse=True):
        if name == test_id or name.startswith(test_id + "."):
            return test_id
    return None


def validate_ibsi2_phase1_bundle(
    reference_dir: Path,
    *,
    manifest_path: Optional[Path] = None,
) -> dict[str, Any]:
    """Validate a consensus response-map bundle against the 33 referencable tests."""

    root = Path(reference_dir).expanduser().resolve()
    files = sorted(
        path
        for path in root.rglob("*")
        if path.is_file() and path.name.casefold().endswith((".nii", ".nii.gz"))
    )
    by_id: dict[str, list[Path]] = defaultdict(list)
    unrecognized: list[str] = []
    for path in files:
        test_id = _phase1_id_from_name(path)
        if test_id is None:
            unrecognized.append(str(path.relative_to(root)))
        else:
            by_id[test_id].append(path)

    found = set(by_id)
    missing = sorted(set(IBSI2_PHASE1_STANDARDIZED_IDS) - found)
    forbidden = sorted(set(IBSI2_PHASE1_NONSTANDARDIZED_IDS).intersection(found))
    unexpected = sorted(found - set(IBSI2_PHASE1_TEST_IDS))
    duplicated = {
        test_id: [str(p.relative_to(root)) for p in paths]
        for test_id, paths in by_id.items()
        if len(paths) != 1
    }
    diagnostics = {
        "missing_standardized_ids": missing,
        "forbidden_nonstandardized_ids": forbidden,
        "unexpected_ids": unexpected,
        "duplicate_ids": duplicated,
        "unrecognized_nifti_files": unrecognized,
    }
    if any(diagnostics.values()) or len(files) != 33:
        raise ReferenceValidationError(
            "Invalid IBSI 2 Phase 1 reference bundle: "
            + json.dumps(diagnostics, sort_keys=True)
        )

    try:
        import nibabel as nib
        import numpy as np
    except ImportError as exc:
        raise RuntimeError(
            "nibabel and numpy are required to validate IBSI 2 response maps"
        ) from exc

    map_records: list[dict[str, Any]] = []
    for test_id in IBSI2_PHASE1_STANDARDIZED_IDS:
        path = by_id[test_id][0]
        digest = sha256_file(path)
        expected_digest = IBSI2_PHASE1_REFERENCE_SHA256[test_id]
        if digest != expected_digest:
            raise ReferenceValidationError(
                f"Response map {test_id} does not match the pinned official hash: "
                f"{digest}"
            )
        image = nib.load(str(path))
        data = np.asanyarray(image.dataobj)
        if data.ndim != 3:
            raise ReferenceValidationError(
                f"Response map {path} is not 3D: {data.shape}"
            )
        if not np.isfinite(data).all():
            raise ReferenceValidationError(
                f"Response map {path} contains non-finite voxels"
            )
        map_records.append(
            {
                "test_id": test_id,
                "path": str(path.relative_to(root)),
                "sha256": digest,
                "shape": [int(value) for value in data.shape],
                "affine": [
                    [float(value) for value in row] for row in image.affine.tolist()
                ],
                "dtype": str(data.dtype),
                "minimum": float(np.min(data)),
                "maximum": float(np.max(data)),
                "reference_range": float(np.max(data) - np.min(data)),
            }
        )

    manifest = {
        "schema_version": REFERENCE_SCHEMA_VERSION,
        "specification": "IBSI 2",
        "phase": "phase1",
        "source": {
            "reference_repository": IBSI2_REFERENCE_REPOSITORY,
            "reference_commit": IBSI2_REFERENCE_COMMIT,
            "reference_readme_sha256": IBSI2_REFERENCE_README_SHA256,
            "license": "CC0-1.0",
            "analysis_repository": IBSI2_ANALYSIS_REPOSITORY,
            "analysis_commit": IBSI2_ANALYSIS_COMMIT,
            "comparison_source": IBSI2_PHASE1_COMPARISON_SOURCE,
            "comparison_source_sha256": IBSI2_PHASE1_COMPARISON_SOURCE_SHA256,
            "hash_verified": True,
        },
        "defined_tests": len(IBSI2_PHASE1_TEST_IDS),
        "standardized_reference_tests": len(IBSI2_PHASE1_STANDARDIZED_IDS),
        "nonstandardized_tests": list(IBSI2_PHASE1_NONSTANDARDIZED_IDS),
        "comparison_rule": IBSI2_PHASE1_COMPARISON_RULE,
        "maps": map_records,
    }
    if manifest_path is not None:
        atomic_write_json(Path(manifest_path).expanduser().resolve(), manifest)
    return manifest
