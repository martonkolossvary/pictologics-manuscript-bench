"""Accessible tables and figures for IBSI feature/filter compliance results."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
from collections import Counter, defaultdict
from pathlib import Path, PurePosixPath, PureWindowsPath
from typing import Any, Iterable, Mapping, Optional, Sequence

from bench.benchmark_ledger import atomic_write_json, atomic_write_text
from bench.compliance.ibsi2_protocol import (
    PHASE1_FILTER_SPECS_BY_ID,
    PHASE2_FILTER_SPECS_BY_ID,
)
from bench.compliance.models import ComparisonRecord
from bench.compliance.references import (
    IBSI1_BENCHMARK_INSTANCES_BY_FAMILY,
    IBSI1_PROFILE_ROWS,
    IBSI1_PROFILE_STANDARDIZED,
    IBSI1_STANDARDIZED_INSTANCES_BY_FAMILY,
    IBSI1_TABLE2_DEFINITION_COUNT,
    IBSI1_TABLE2_FAMILY_DEFINITIONS,
    IBSI2_PHASE1_COMPARISON_RULE,
    IBSI2_PHASE1_NONSTANDARDIZED_IDS,
    IBSI2_PHASE1_STANDARDIZED_IDS,
    IBSI2_PHASE1_TEST_IDS,
    IBSI2_PHASE2_DEFINED_CHECKS,
    IBSI2_PHASE2_DEFINED_FILTER_IDS,
    IBSI2_PHASE2_NONSTANDARDIZED_CHECKS,
    IBSI2_PHASE2_NONSTANDARDIZED_PUBLISHED_PAIR,
    IBSI2_PHASE2_PUBLISHED_FILTER_IDS,
    IBSI2_PHASE2_PUBLISHED_REFERENCE_ROWS,
    IBSI2_PHASE2_STANDARDIZED_CHECKS,
    IBSI2_PHASE2_TAGS,
    ibsi1_table2_definition_key,
)
from bench.ibsi_families import FAMILY_LABELS, FAMILY_ORDER


ADAPTER_STYLES = {
    "pictologics": ("#1764ab", "o", "/"),
    "pyradiomics": ("#d55e00", "s", "\\"),
    "mirp": ("#6f4e9c", "^", "xx"),
    "medimage": ("#26734d", "D", ".."),
    "zrad": ("#a51c30", "P", "++"),
}

_COMPARISON_FIELDS = tuple(ComparisonRecord.__dataclass_fields__)

_PHASE2_COMPONENT_SCOPE = (
    "Component-level filter/statistic evaluation after protocol-controlled common "
    "preprocessing; adapter-native image import and resampling are outside this "
    "comparison."
)

_PHASE1_STRICT_3D_TEST_IDS = tuple(
    test_id
    for test_id in IBSI2_PHASE1_TEST_IDS
    if int(PHASE1_FILTER_SPECS_BY_ID[test_id].parameters["dimensionality"]) == 3
)
_PHASE1_STRICT_3D_STANDARDIZED_IDS = tuple(
    test_id
    for test_id in _PHASE1_STRICT_3D_TEST_IDS
    if test_id in IBSI2_PHASE1_STANDARDIZED_IDS
)
_PHASE1_STRICT_3D_FILTER_ORDER = (
    "mean",
    "log",
    "laws",
    "wavelet",
    "simoncelli",
    "riesz_log",
    "riesz_simoncelli",
)
_FILTER_LABELS = {
    "none": "Identity",
    "mean": "Mean",
    "log": "LoG",
    "laws": "Laws",
    "gabor": "Gabor",
    "wavelet": "Wavelet",
    "simoncelli": "Simoncelli",
    "riesz_log": "Riesz-LoG",
    "riesz_simoncelli": "Riesz-Simoncelli",
}
_PHASE2_B_DEFINED_FILTER_IDS = tuple(
    filter_id
    for filter_id in IBSI2_PHASE2_DEFINED_FILTER_IDS
    if filter_id.endswith(".B")
)
_PHASE2_B_PUBLISHED_FILTER_IDS = tuple(
    filter_id
    for filter_id in IBSI2_PHASE2_PUBLISHED_FILTER_IDS
    if filter_id.endswith(".B")
)
_PHASE2_B_PUBLISHED_REFERENCE_ROWS = len(_PHASE2_B_PUBLISHED_FILTER_IDS) * len(
    IBSI2_PHASE2_TAGS
)
_PHASE2_B_STANDARDIZED_CHECKS = _PHASE2_B_PUBLISHED_REFERENCE_ROWS - 1

if (
    len(_PHASE1_STRICT_3D_TEST_IDS) != 27
    or len(_PHASE1_STRICT_3D_STANDARDIZED_IDS) != 24
    or len(_PHASE2_B_DEFINED_FILTER_IDS) != 11
    or len(_PHASE2_B_PUBLISHED_FILTER_IDS) != 9
    or _PHASE2_B_PUBLISHED_REFERENCE_ROWS != 162
    or _PHASE2_B_STANDARDIZED_CHECKS != 161
):
    raise RuntimeError(
        "IBSI 2 scoped report denominators differ from the reviewed design"
    )

_FIELD_DEFINITIONS = {
    "specification": "IBSI specification represented by the row.",
    "phase": "IBSI benchmark phase represented by the row.",
    "profile": "Named analysis profile used to group otherwise comparable rows.",
    "adapter": "Radiomics adapter or software implementation evaluated.",
    "software_version": "Installed adapter software version used for calculation.",
    "generator_distribution": "Python distribution that generated the candidate response map.",
    "generator_version": "Installed generator version that produced the candidate response map.",
    "configuration": "IBSI filter configuration identifier.",
    "protocol_variant": (
        "IBSI Phase 2 pathway: A is slice-wise 2D and B is volumetric 3D. "
        "Configuration 5.B is the specified exception, using 2D Gabor responses "
        "averaged over three orthogonal planes."
    ),
    "aggregation": "Feature aggregation mode recorded for the calculation.",
    "family": "IBSI feature family or statistic family.",
    "feature_name": "Human-readable feature or statistic name.",
    "feature_tag": "IBSI workbook feature/statistic tag.",
    "semantic_key": "Stable semantic identifier used for adapter-independent mapping.",
    "ibsi_code": "IBSI identifier or code for the feature/statistic.",
    "standardized": "Whether an official consensus value and tolerance are standardized.",
    "observed_supported": (
        "Whether the row belongs to an explicitly reviewed native-supported "
        "configuration; never inferred from a finite result."
    ),
    "native_supported": (
        "Yes only for an explicitly reviewed native-supported filter configuration or test."
    ),
    "native_support_status": (
        "Reviewed native-support classification: supported, unsupported, or incomplete."
    ),
    "mapped": "Whether the adapter output was mapped unambiguously to this IBSI row.",
    "attempted": "Whether calculation was attempted for this row.",
    "finite": "Whether calculation produced a finite numerical result.",
    "referencable": "Whether an official reference value and tolerance can be applied.",
    "evaluated": "Whether a finite mapped result was compared with an official reference.",
    "passed": "Whether the evaluated result was within the official tolerance boundary.",
    "status": "Explicit row outcome; unsupported and unavailable rows are not counted as passes.",
    "native_feature_names": "Native adapter output name or names supporting the mapped row.",
    "value": "Calculated candidate value.",
    "reference_value": "Official IBSI consensus reference value.",
    "tolerance": "Official IBSI absolute tolerance for this comparison.",
    "raw_abs_error": "Absolute candidate-minus-reference error before policy transformations.",
    "comparison_error": "Error quantity used by the declared comparison policy.",
    "error_tolerance_ratio": "Comparison error divided by tolerance when that ratio is defined.",
    "comparison_policy": "Rule used to determine the tolerance outcome.",
    "detail": "Audit explanation, including reviewed support-declaration evidence where applicable.",
    "defined_features": "Specification-specific denominator; consult the declared denominator policy.",
    "supported_features": "Specification-specific numerator; consult the declared denominator policy.",
    "native_supported_definitions": (
        "IBSI 1 Table 2 definitions for which every required workbook instance "
        "maps to a reviewed native output or audited exact identity."
    ),
    "finite_definitions": (
        "IBSI 1 Table 2 definitions for which every required workbook instance "
        "has a finite calculated value."
    ),
    "distinct_source_outputs": "Number of distinct named native outputs contributing to mapped rows.",
    "documented_exact_identity_definitions": "Number of semantic definitions assigned by an audited exact identity.",
    "feature_support_rate": "supported_features divided by defined_features.",
    "defined_benchmark_instances": "Number of designed benchmark rows in the applicable profile.",
    "supported_benchmark_instances": "Designed rows belonging to reviewed native-supported configurations.",
    "finite_benchmark_instances": "Designed rows with finite calculated values.",
    "benchmark_instance_coverage_rate": "Finite benchmark rows divided by designed benchmark rows.",
    "defined_standardized_checks": "Fixed number of standardized tolerance checks in scope.",
    "supported_checks": "Standardized rows belonging to reviewed native-supported configurations.",
    "conditional_accuracy": "Passed checks divided by evaluated checks; unsupported or unevaluated rows are outside this denominator.",
    "overall_referenced_success": "Passed checks divided by all standardized reference checks in scope.",
    "overall_standardized_success": "Passed checks divided by the fixed standardized-check denominator.",
    "native_supported_filters": "Number of explicitly reviewed native-supported Phase 2 filter configurations.",
    "defined_filters": "Number of Phase 2 filter configurations in the full design (22).",
    "native_filter_support_rate": "Reviewed native-supported configurations divided by 22 defined configurations.",
    "finite_designed_checks": "Finite configuration-statistic calculations among 396 designed Phase 2 checks.",
    "finite_calculation_denominator": "Fixed designed-check denominator for finite calculation coverage.",
    "published_reference_rows": "Official Phase 2 comparison rows published by IBSI (324).",
    "finite_published_reference_rows": "Published reference rows with a finite calculated value.",
    "standardized_reference_checks": "Published rows with standardized consensus values and tolerances.",
    "evaluated_standardized_checks": "Standardized rows actually compared with a reference.",
    "passed_standardized_checks": "Evaluated standardized rows within tolerance.",
    "failed_standardized_checks": "Evaluated standardized rows outside tolerance.",
    "overall_standardized_success_denominator": "Fixed standardized-check denominator, including unsupported and unevaluated rows.",
    "nonstandardized_defined_checks": "Designed checks without a standardized tolerance.",
    "defined_checks": "Configuration-statistic checks defined for this filter configuration (18).",
    "finite_checks": "Finite statistics calculated for this filter configuration.",
    "finite_calculation_rate": "finite_checks divided by defined_checks.",
    "support_declaration": "Reviewed native-support evidence retained from the candidate declaration.",
    "controlled_preprocessing_scope": "Scope limitation for the controlled common-preprocessing component evaluation.",
    "defined_filter_tests": "IBSI 2 Phase 1 filter tests in the full design (36).",
    "native_supported_filter_tests": "Explicitly reviewed native-supported Phase 1 filter tests.",
    "candidate_maps_supplied": "Provenance-attested candidate response maps supplied for evaluation.",
    "standardized_reference_tests": "Phase 1 response-map tests with standardized references (33).",
    "candidate_supplied": "Whether a provenance-attested candidate response map was supplied.",
    "test_id": "IBSI 2 Phase 1 filter-test identifier.",
    "filter_family": "Reviewed IBSI filter family derived from the exact protocol configuration.",
    "strict_3d_scope": "Whether the Phase 1 test uses an IBSI-defined three-dimensional filter.",
    "kernel_dimensionality": "Dimensionality of the native filter kernel in the reviewed IBSI configuration.",
    "workflow_dimensionality": "Dimensionality of the IBSI pathway in which the filter is evaluated.",
    "published_configuration": "Whether official Phase 2 statistic references are published for the configuration.",
    "three_plane_gabor_exception": (
        "True only for Phase 2 configuration 5.B, which averages 2D Gabor "
        "responses over three orthogonal planes within the volumetric B pathway."
    ),
    "native_supported_b_designed_filters": "Reviewed native-supported configurations among the 11 designed Phase 2 B configurations.",
    "native_supported_b_published_filters": "Reviewed native-supported configurations among the 9 published Phase 2 B configurations.",
    "defined_b_filters": "All 11 designed Phase 2 Configuration-B filters.",
    "published_b_filters": "The 9 Configuration-B filters with published Phase 2 statistic rows.",
    "finite_published_b_checks": "Finite calculated values among the 162 published Configuration-B rows.",
    "published_b_reference_rows": "The 162 published Configuration-B statistic rows.",
    "standardized_b_checks": "The 161 published Configuration-B rows with consensus values and tolerances.",
    "native_supported_nonstandardized_tests": "Native-supported tests among the three strict-3D Phase 1 tests without standardized response maps.",
    "voxel_tolerance": "Allowed all-voxel absolute error: one percent of the reference-map range.",
    "max_abs_error": "Largest absolute voxel error observed in the response map.",
}


def _bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes"}


def _optional_bool(value: Any) -> Optional[bool]:
    if value is None or str(value).strip() == "":
        return None
    return _bool(value)


def _optional_float(value: Any) -> Optional[float]:
    if value is None or str(value).strip() == "":
        return None
    result = float(value)
    return result if math.isfinite(result) else None


def _public_candidate_path(value: str) -> str:
    """Strip local roots from candidate paths written to public artifacts."""

    native = Path(value)
    posix = PurePosixPath(value)
    windows = PureWindowsPath(value)
    unsafe = (
        native.is_absolute()
        or posix.is_absolute()
        or windows.is_absolute()
        or bool(windows.root)
        or ".." in native.parts
        or ".." in posix.parts
        or ".." in windows.parts
    )
    if not unsafe:
        return value
    return PurePosixPath(value.replace("\\", "/")).name


def comparison_to_csv(records: Iterable[ComparisonRecord]) -> str:
    stream = io.StringIO(newline="")
    writer = csv.DictWriter(stream, fieldnames=_COMPARISON_FIELDS, lineterminator="\n")
    writer.writeheader()
    for record in records:
        row = record.to_dict()
        for key in (
            "value",
            "reference_value",
            "tolerance",
            "raw_abs_error",
            "comparison_error",
            "error_tolerance_ratio",
        ):
            if row[key] is None:
                row[key] = ""
        if row["passed"] is None:
            row["passed"] = ""
        writer.writerow(row)
    return stream.getvalue()


def load_comparison_csv(path: Path) -> list[ComparisonRecord]:
    with Path(path).open("r", encoding="utf-8", newline="") as stream:
        rows = list(csv.DictReader(stream))
    return [
        ComparisonRecord(
            specification=row["specification"],
            phase=row["phase"],
            adapter=row["adapter"],
            software_version=row["software_version"],
            configuration=row["configuration"],
            profile=row["profile"],
            aggregation=row["aggregation"],
            family=row["family"],
            feature_name=row["feature_name"],
            feature_tag=row["feature_tag"],
            semantic_key=row["semantic_key"],
            ibsi_code=row["ibsi_code"],
            standardized=_bool(row["standardized"]),
            observed_supported=_bool(row["observed_supported"]),
            mapped=_bool(row["mapped"]),
            attempted=_bool(row["attempted"]),
            finite=_bool(row["finite"]),
            referencable=_bool(row["referencable"]),
            evaluated=_bool(row["evaluated"]),
            passed=_optional_bool(row["passed"]),
            status=row["status"],
            native_feature_names=row.get("native_feature_names", ""),
            value=_optional_float(row.get("value")),
            reference_value=_optional_float(row.get("reference_value")),
            tolerance=_optional_float(row.get("tolerance")),
            raw_abs_error=_optional_float(row.get("raw_abs_error")),
            comparison_error=_optional_float(row.get("comparison_error")),
            error_tolerance_ratio=_optional_float(row.get("error_tolerance_ratio")),
            comparison_policy=row.get("comparison_policy", ""),
            detail=row.get("detail", ""),
        )
        for row in rows
    ]


def _csv_text(rows: Sequence[Mapping[str, Any]], fields: Sequence[str]) -> str:
    stream = io.StringIO(newline="")
    writer = csv.DictWriter(
        stream, fieldnames=list(fields), extrasaction="ignore", lineterminator="\n"
    )
    writer.writeheader()
    writer.writerows(rows)
    return stream.getvalue()


def _data_dictionary_rows(
    tables: Sequence[tuple[str, Sequence[Mapping[str, Any]]]],
) -> list[dict[str, str]]:
    """Build a machine-readable dictionary for every published table column."""

    rows: list[dict[str, str]] = []
    for table_name, table_rows in tables:
        fields = list(table_rows[0]) if table_rows else []
        for field in fields:
            rows.append(
                {
                    "table": table_name,
                    "field": field,
                    "definition": _FIELD_DEFINITIONS.get(
                        field,
                        f"Report field '{field.replace('_', ' ')}'; interpret with "
                        "the report denominator policy and row status.",
                    ),
                }
            )
    return rows


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _write_release_manifest(
    output: Path,
    artifacts: Sequence[tuple[str, str]],
    *,
    filename: str = "release_manifest.json",
) -> dict[str, Any]:
    """Write hashes for publication artifacts using output-relative POSIX paths."""

    entries = []
    seen: set[str] = set()
    for relative_name, role in artifacts:
        relative = Path(relative_name)
        if relative.is_absolute() or ".." in relative.parts:
            raise ValueError(f"Release artifact path must be relative: {relative_name}")
        normalized = relative.as_posix()
        if normalized in seen:
            continue
        seen.add(normalized)
        path = output / relative
        if not path.is_file():
            raise FileNotFoundError(f"Declared release artifact does not exist: {path}")
        entries.append(
            {
                "path": normalized,
                "role": role,
                "sha256": _sha256_file(path),
                "bytes": path.stat().st_size,
            }
        )
    manifest = {
        "schema_version": 1,
        "hash_algorithm": "sha256",
        "path_policy": "POSIX paths relative to this release-manifest directory",
        "artifacts": sorted(entries, key=lambda entry: entry["path"]),
    }
    atomic_write_json(output / filename, manifest)
    return manifest


def _is_ibsi1_digital_phantom(records: Sequence[ComparisonRecord]) -> bool:
    return bool(records) and all(
        record.specification == "IBSI 1"
        and record.phase == "phase1"
        and record.configuration == "digital_phantom"
        for record in records
    )


def _is_ibsi2_phase2(records: Sequence[ComparisonRecord]) -> bool:
    return bool(records) and all(
        record.specification == "IBSI 2" and record.phase == "phase2"
        for record in records
    )


def _ibsi2_phase2_row_complete(records: Sequence[ComparisonRecord]) -> bool:
    """Return whether every adapter has an exhaustive, processed support design."""

    if not _is_ibsi2_phase2(records):
        return False
    grouped: dict[tuple[str, str, str], list[ComparisonRecord]] = defaultdict(list)
    for record in records:
        grouped[(record.adapter, record.software_version, record.profile)].append(
            record
        )
    incomplete_statuses = {
        "candidate_not_supplied",
        "configuration_unsupported_or_not_supplied",
        "filter_unsupported_or_not_supplied",
        "error",
    }
    for values in grouped.values():
        if len(values) != IBSI2_PHASE2_DEFINED_CHECKS or any(
            record.status in incomplete_statuses for record in values
        ):
            return False
        by_filter: dict[str, list[ComparisonRecord]] = defaultdict(list)
        for record in values:
            by_filter[record.configuration].append(record)
        if set(by_filter) != set(IBSI2_PHASE2_DEFINED_FILTER_IDS):
            return False
        for filter_records in by_filter.values():
            supported_declaration = all(
                record.detail.startswith(
                    "reviewed native-supported filter declaration:"
                )
                for record in filter_records
            )
            unsupported_declaration = all(
                record.status == "native_unsupported"
                and record.detail.startswith("reviewed native-support declaration:")
                for record in filter_records
            )
            if not (supported_declaration or unsupported_declaration):
                return False
    return bool(grouped)


def _ibsi2_phase2_native_supported_filters(
    records: Sequence[ComparisonRecord],
) -> set[str]:
    """Recover explicitly reviewed native support from row classifications."""

    grouped: dict[str, list[ComparisonRecord]] = defaultdict(list)
    for record in records:
        grouped[record.configuration].append(record)
    return {
        configuration
        for configuration, values in grouped.items()
        if values
        and all(
            record.detail.startswith("reviewed native-supported filter declaration:")
            for record in values
        )
    }


def _ibsi2_phase2_native_support_status(
    records: Sequence[ComparisonRecord],
) -> str:
    if records and all(
        record.detail.startswith("reviewed native-supported filter declaration:")
        for record in records
    ):
        return "supported"
    if records and all(
        record.status == "native_unsupported"
        and record.detail.startswith("reviewed native-support declaration:")
        for record in records
    ):
        return "unsupported"
    return "incomplete"


def _ibsi2_phase2_configuration_summary(
    records: Sequence[ComparisonRecord],
) -> list[dict[str, Any]]:
    """Summarize each adapter/configuration without conflating four outcomes."""

    if not _is_ibsi2_phase2(records):
        return []
    groups: dict[tuple[str, str, str, str], list[ComparisonRecord]] = defaultdict(list)
    for record in records:
        groups[
            (
                record.adapter,
                record.software_version,
                record.profile,
                record.configuration,
            )
        ].append(record)
    expected_checks = len(IBSI2_PHASE2_TAGS)
    published = set(IBSI2_PHASE2_PUBLISHED_FILTER_IDS)
    rows: list[dict[str, Any]] = []
    for (adapter, software_version, profile, configuration), values in sorted(
        groups.items()
    ):
        native_status = _ibsi2_phase2_native_support_status(values)
        standardized = [record for record in values if record.standardized]
        evaluated = sum(record.evaluated for record in standardized)
        passed = sum(record.passed is True for record in standardized)
        failed = sum(record.passed is False for record in standardized)
        finite = sum(record.finite for record in values)
        published_rows = expected_checks if configuration in published else 0
        declaration = " | ".join(
            sorted(
                {
                    record.detail
                    for record in values
                    if record.detail.startswith("reviewed native-")
                }
            )
        )
        rows.append(
            {
                "specification": "IBSI 2",
                "phase": "phase2",
                "profile": profile,
                "adapter": adapter,
                "software_version": software_version,
                "configuration": configuration,
                "protocol_variant": configuration.rsplit(".", 1)[-1],
                "native_support_status": native_status,
                "native_supported": native_status == "supported",
                "defined_checks": expected_checks,
                "finite_checks": finite,
                "finite_calculation_rate": finite / expected_checks,
                "published_reference_rows": published_rows,
                "finite_published_reference_rows": (
                    finite if configuration in published else 0
                ),
                "standardized_reference_checks": len(standardized),
                "evaluated_standardized_checks": evaluated,
                "passed_standardized_checks": passed,
                "failed_standardized_checks": failed,
                "conditional_accuracy": passed / evaluated if evaluated else None,
                "overall_standardized_success": (
                    passed / len(standardized) if standardized else None
                ),
                "overall_standardized_success_denominator": len(standardized),
                "nonstandardized_defined_checks": expected_checks - len(standardized),
                "controlled_preprocessing_scope": _PHASE2_COMPONENT_SCOPE,
                "support_declaration": declaration,
            }
        )
    return rows


def _ibsi2_phase2_b_configuration_summary(
    records: Sequence[ComparisonRecord],
) -> list[dict[str, Any]]:
    """Return the 11 designed B configurations with explicit kernel scope."""

    rows = []
    for row in _ibsi2_phase2_configuration_summary(records):
        configuration = str(row["configuration"])
        if configuration not in _PHASE2_B_DEFINED_FILTER_IDS:
            continue
        parameters = PHASE2_FILTER_SPECS_BY_ID[configuration].parameters
        rows.append(
            {
                **row,
                "filter_family": str(parameters["filter"]),
                "workflow_dimensionality": 3,
                "kernel_dimensionality": int(parameters["dimensionality"]),
                "published_configuration": (
                    configuration in _PHASE2_B_PUBLISHED_FILTER_IDS
                ),
                "three_plane_gabor_exception": configuration == "5.B",
            }
        )
    return rows


def _ibsi2_phase2_b_summary(
    records: Sequence[ComparisonRecord],
) -> list[dict[str, Any]]:
    """Summarize the volumetric B pathway without mixing in A results."""

    grouped: dict[tuple[str, str, str], list[ComparisonRecord]] = defaultdict(list)
    for record in records:
        if record.configuration in _PHASE2_B_DEFINED_FILTER_IDS:
            grouped[(record.adapter, record.software_version, record.profile)].append(
                record
            )

    rows: list[dict[str, Any]] = []
    published_ids = set(_PHASE2_B_PUBLISHED_FILTER_IDS)
    for (adapter, software_version, profile), values in sorted(grouped.items()):
        published = [
            record for record in values if record.configuration in published_ids
        ]
        standardized = [record for record in published if record.standardized]
        evaluated = [record for record in standardized if record.evaluated]
        passed = sum(record.passed is True for record in standardized)
        supported = _ibsi2_phase2_native_supported_filters(values)
        rows.append(
            {
                "specification": "IBSI 2",
                "phase": "phase2",
                "profile": profile,
                "adapter": adapter,
                "software_version": software_version,
                "defined_b_filters": len(_PHASE2_B_DEFINED_FILTER_IDS),
                "native_supported_b_designed_filters": len(
                    supported.intersection(_PHASE2_B_DEFINED_FILTER_IDS)
                ),
                "published_b_filters": len(_PHASE2_B_PUBLISHED_FILTER_IDS),
                "native_supported_b_published_filters": len(
                    supported.intersection(_PHASE2_B_PUBLISHED_FILTER_IDS)
                ),
                "published_b_reference_rows": _PHASE2_B_PUBLISHED_REFERENCE_ROWS,
                "finite_published_b_checks": sum(record.finite for record in published),
                "standardized_b_checks": _PHASE2_B_STANDARDIZED_CHECKS,
                "evaluated_standardized_checks": len(evaluated),
                "passed_standardized_checks": passed,
                "failed_standardized_checks": sum(
                    record.passed is False for record in standardized
                ),
                "conditional_accuracy": (
                    passed / len(evaluated) if evaluated else None
                ),
                "overall_standardized_success": (
                    passed / _PHASE2_B_STANDARDIZED_CHECKS
                ),
                "overall_standardized_success_denominator": (
                    _PHASE2_B_STANDARDIZED_CHECKS
                ),
                "controlled_preprocessing_scope": _PHASE2_COMPONENT_SCOPE,
            }
        )
    return rows


def _ibsi1_definition_groups(
    records: Sequence[ComparisonRecord],
) -> dict[tuple[str, str], list[ComparisonRecord]]:
    groups: dict[tuple[str, str], list[ComparisonRecord]] = defaultdict(list)
    for record in records:
        groups[
            (record.family, ibsi1_table2_definition_key(record.semantic_key))
        ].append(record)
    return groups


def _ibsi1_definition_count(
    records: Sequence[ComparisonRecord],
    *,
    attribute: str,
) -> int:
    """Count Table 2 definitions satisfying an instance-level boolean field.

    The two parameterised IVH definitions each require both their 10% and 90%
    workbook instances to satisfy the field before the definition is counted.
    """

    count = 0
    for (_, definition), values in _ibsi1_definition_groups(records).items():
        expected_instances = (
            2
            if definition
            in {
                "ivh.volume_at_intensity_fraction",
                "ivh.intensity_at_volume_fraction",
            }
            else 1
        )
        if len(values) == expected_instances and all(
            bool(getattr(record, attribute)) for record in values
        ):
            count += 1
    return count


def _ibsi1_finite_definition_count(records: Sequence[ComparisonRecord]) -> int:
    """Count definitions whose complete workbook instance set is finite."""

    return _ibsi1_definition_count(records, attribute="finite")


def _ibsi1_native_supported_definition_count(
    records: Sequence[ComparisonRecord],
) -> int:
    """Count definitions backed by reviewed native outputs or exact identities."""

    return _ibsi1_definition_count(records, attribute="observed_supported")


def _source_output_and_alias_counts(
    records: Sequence[ComparisonRecord],
) -> tuple[int, int]:
    """Return distinct source outputs and exact-alias semantic definitions."""

    source_outputs: set[str] = set()
    alias_definitions: set[str] = set()
    alias_suffix = " [documented exact alias]"
    for record in records:
        if not record.observed_supported:
            continue
        for name in record.native_feature_names.split(", "):
            if not name:
                continue
            if name.endswith(alias_suffix):
                source_outputs.add(name[: -len(alias_suffix)])
                if record.semantic_key:
                    alias_definitions.add(record.semantic_key)
            else:
                source_outputs.add(name)
    return len(source_outputs), len(alias_definitions)


def _overall_summary(records: Sequence[ComparisonRecord]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str, str, str, str], list[ComparisonRecord]] = defaultdict(
        list
    )
    for record in records:
        groups[
            (
                record.specification,
                record.phase,
                record.profile,
                record.adapter,
                record.software_version,
            )
        ].append(record)
    rows: list[dict[str, Any]] = []
    for (
        specification,
        phase,
        profile,
        adapter,
        software_version,
    ), values in sorted(groups.items()):
        standardized = [record for record in values if record.standardized]
        is_ibsi1 = _is_ibsi1_digital_phantom(values)
        is_ibsi2_phase2 = _is_ibsi2_phase2(values)
        defined_codes = {
            record.semantic_key for record in values if record.semantic_key
        }
        supported_codes = {
            record.semantic_key
            for record in values
            if record.semantic_key and record.observed_supported
        }
        standardized_codes = {
            record.semantic_key for record in standardized if record.semantic_key
        }
        standardized_supported_codes = {
            record.semantic_key
            for record in standardized
            if record.semantic_key and record.observed_supported
        }
        evaluated = sum(record.evaluated for record in standardized)
        passed = sum(record.passed is True for record in standardized)
        referencable = sum(record.referencable for record in standardized)
        finite = sum(record.finite for record in standardized)
        supported_checks = sum(record.observed_supported for record in standardized)
        unsupported_statuses = {
            "unsupported",
            "native_unsupported",
            "configuration_unsupported_or_not_supplied",
            "filter_unsupported_or_not_supplied",
            "candidate_not_supplied",
        }
        unsupported = sum(
            record.status in unsupported_statuses for record in standardized
        )
        supported_benchmark_instances = sum(
            record.observed_supported for record in values
        )
        finite_benchmark_instances = sum(record.finite for record in values)
        finite_published_reference_rows = sum(
            record.finite and record.configuration in IBSI2_PHASE2_PUBLISHED_FILTER_IDS
            for record in values
        )
        source_output_count, exact_alias_count = _source_output_and_alias_counts(values)
        if is_ibsi1:
            defined_feature_count = IBSI1_TABLE2_DEFINITION_COUNT
            supported_feature_count = _ibsi1_finite_definition_count(values)
            defined_benchmark_instances = IBSI1_PROFILE_ROWS
        elif is_ibsi2_phase2:
            # Phase 2 is a configuration-by-statistic design, not an 18-feature
            # inventory.  Only 324 of the 396 designed cells have published
            # reference rows, so finite coverage must retain the full design
            # denominator rather than collapsing across configurations.
            defined_feature_count = IBSI2_PHASE2_DEFINED_CHECKS
            supported_feature_count = finite_benchmark_instances
            defined_benchmark_instances = IBSI2_PHASE2_DEFINED_CHECKS
        else:
            defined_feature_count = len(defined_codes)
            supported_feature_count = len(supported_codes)
            defined_benchmark_instances = len(values)
        row = {
            "specification": specification,
            "phase": phase,
            "profile": profile,
            "adapter": adapter,
            "software_version": software_version,
            "defined_features": defined_feature_count,
            "supported_features": supported_feature_count,
            "distinct_source_outputs": source_output_count,
            "documented_exact_identity_definitions": exact_alias_count,
            "feature_support_rate": (
                supported_feature_count / defined_feature_count
                if defined_feature_count
                else None
            ),
            "defined_benchmark_instances": defined_benchmark_instances,
            "supported_benchmark_instances": supported_benchmark_instances,
            "finite_benchmark_instances": finite_benchmark_instances,
            "benchmark_instance_coverage_rate": (
                finite_benchmark_instances / defined_benchmark_instances
                if defined_benchmark_instances
                else None
            ),
            "standardized_features": (
                IBSI2_PHASE2_STANDARDIZED_CHECKS
                if is_ibsi2_phase2
                else len(standardized_codes)
            ),
            "standardized_supported_features": (
                finite if is_ibsi2_phase2 else len(standardized_supported_codes)
            ),
            "standardized_benchmark_instances": (
                IBSI2_PHASE2_STANDARDIZED_CHECKS
                if is_ibsi2_phase2
                else len(standardized)
            ),
            "supported_standardized_benchmark_instances": supported_checks,
            "defined_standardized_checks": (
                IBSI2_PHASE2_STANDARDIZED_CHECKS
                if is_ibsi2_phase2
                else len(standardized)
            ),
            "supported_checks": supported_checks,
            "attempted": sum(record.attempted for record in standardized),
            "finite": finite,
            "referencable": referencable,
            "evaluated": evaluated,
            "passed": passed,
            "failed": sum(record.passed is False for record in standardized),
            "unsupported": unsupported,
            "status_unsupported": sum(
                record.status == "unsupported" for record in standardized
            ),
            "missing": sum(record.status == "missing" for record in standardized),
            "nonfinite": sum(record.status == "nonfinite" for record in standardized),
            "ambiguous": sum(record.status == "ambiguous" for record in standardized),
            "execution_coverage": finite / supported_checks
            if supported_checks
            else None,
            "conditional_accuracy": passed / evaluated if evaluated else None,
            "overall_referenced_success": passed / referencable
            if referencable
            else None,
        }
        if is_ibsi1:
            row.update(
                {
                    "native_supported_definitions": (
                        _ibsi1_native_supported_definition_count(values)
                    ),
                    "finite_definitions": _ibsi1_finite_definition_count(values),
                }
            )
        if is_ibsi2_phase2:
            native_supported_filters = _ibsi2_phase2_native_supported_filters(values)
            row.update(
                {
                    "native_supported_filters": len(native_supported_filters),
                    "defined_filters": len(IBSI2_PHASE2_DEFINED_FILTER_IDS),
                    "native_filter_support_rate": (
                        len(native_supported_filters)
                        / len(IBSI2_PHASE2_DEFINED_FILTER_IDS)
                    ),
                    "finite_designed_checks": finite_benchmark_instances,
                    "finite_calculation_denominator": IBSI2_PHASE2_DEFINED_CHECKS,
                    "published_reference_rows": IBSI2_PHASE2_PUBLISHED_REFERENCE_ROWS,
                    "finite_published_reference_rows": finite_published_reference_rows,
                    "nonstandardized_defined_checks": IBSI2_PHASE2_NONSTANDARDIZED_CHECKS,
                    "conditional_accuracy_numerator": passed,
                    "conditional_accuracy_denominator": evaluated,
                    "overall_standardized_success_numerator": passed,
                    "overall_standardized_success_denominator": IBSI2_PHASE2_STANDARDIZED_CHECKS,
                    "overall_standardized_success": (
                        passed / IBSI2_PHASE2_STANDARDIZED_CHECKS
                    ),
                    "controlled_preprocessing_scope": _PHASE2_COMPONENT_SCOPE,
                    "filter_support_inferred_from_candidate_maps": False,
                }
            )
        rows.append(row)
    return rows


def _family_summary(records: Sequence[ComparisonRecord]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str, str, str, str, str], list[ComparisonRecord]] = (
        defaultdict(list)
    )
    for record in records:
        groups[
            (
                record.specification,
                record.phase,
                record.profile,
                record.adapter,
                record.software_version,
                record.family,
            )
        ].append(record)
    rows = []
    for (
        specification,
        phase,
        profile,
        adapter,
        software_version,
        family,
    ), values in sorted(groups.items()):
        defined = {record.semantic_key for record in values if record.semantic_key}
        supported = {
            record.semantic_key
            for record in values
            if record.semantic_key and record.observed_supported
        }
        is_ibsi1 = _is_ibsi1_digital_phantom(values)
        is_ibsi2_phase2 = _is_ibsi2_phase2(values)
        standardized_values = [record for record in values if record.standardized]
        evaluated = sum(record.evaluated for record in standardized_values)
        passed = sum(record.passed is True for record in standardized_values)
        source_output_count, exact_alias_count = _source_output_and_alias_counts(values)
        row = {
            "specification": specification,
            "phase": phase,
            "profile": profile,
            "adapter": adapter,
            "software_version": software_version,
            "family": family,
            "defined_features": (
                IBSI1_TABLE2_FAMILY_DEFINITIONS[family]
                if is_ibsi1
                else (IBSI2_PHASE2_DEFINED_CHECKS if is_ibsi2_phase2 else len(defined))
            ),
            "supported_features": (
                _ibsi1_finite_definition_count(values)
                if is_ibsi1
                else (
                    sum(record.finite for record in values)
                    if is_ibsi2_phase2
                    else len(supported)
                )
            ),
            "distinct_source_outputs": source_output_count,
            "documented_exact_identity_definitions": exact_alias_count,
            "benchmark_instances": (
                IBSI1_BENCHMARK_INSTANCES_BY_FAMILY[family]
                if is_ibsi1
                else (IBSI2_PHASE2_DEFINED_CHECKS if is_ibsi2_phase2 else len(values))
            ),
            "supported_benchmark_instances": sum(
                record.observed_supported for record in values
            ),
            "finite_benchmark_instances": sum(record.finite for record in values),
            "benchmark_instance_coverage_rate": (
                sum(record.finite for record in values)
                / IBSI1_BENCHMARK_INSTANCES_BY_FAMILY[family]
                if is_ibsi1
                else sum(record.finite for record in values)
                / IBSI2_PHASE2_DEFINED_CHECKS
                if is_ibsi2_phase2
                else sum(record.finite for record in values) / len(values)
                if values
                else None
            ),
            "standardized_checks": len(standardized_values),
            "evaluated_checks": evaluated,
            "passed_checks": passed,
            "failed_checks": sum(
                record.passed is False for record in standardized_values
            ),
            "conditional_accuracy": passed / evaluated if evaluated else None,
        }
        if is_ibsi1:
            row.update(
                {
                    "native_supported_definitions": (
                        _ibsi1_native_supported_definition_count(values)
                    ),
                    "finite_definitions": _ibsi1_finite_definition_count(values),
                }
            )
        if is_ibsi2_phase2:
            row.update(
                {
                    "finite_designed_checks": sum(record.finite for record in values),
                    "published_reference_rows": IBSI2_PHASE2_PUBLISHED_REFERENCE_ROWS,
                    "standardized_reference_checks": IBSI2_PHASE2_STANDARDIZED_CHECKS,
                    "nonstandardized_defined_checks": IBSI2_PHASE2_NONSTANDARDIZED_CHECKS,
                }
            )
        rows.append(row)
    return rows


def _save_figure(fig, stem: Path) -> list[str]:
    outputs = []
    for suffix in ("pdf", "svg", "png"):
        path = stem.with_suffix("." + suffix)
        fig.savefig(path, dpi=300, bbox_inches="tight", facecolor="white")
        outputs.append(path.name)
    return outputs


def _style_axes(ax) -> None:
    ax.set_facecolor("white")
    for spine in ax.spines.values():
        spine.set_color("#333333")
        spine.set_linewidth(1.0)
    ax.tick_params(labelsize=12, colors="#222222")
    ax.title.set_fontsize(14)


def _pyplot():
    """Use a deterministic non-interactive backend for CLI/report execution."""

    import matplotlib

    matplotlib.use("Agg", force=True)
    import matplotlib.pyplot as plt

    return plt


def _summary_labels(summary: Sequence[Mapping[str, Any]]) -> list[str]:
    adapter_version_counts = Counter(
        (str(row["adapter"]), str(row.get("software_version", ""))) for row in summary
    )
    labels = []
    for row in summary:
        adapter = str(row["adapter"]).replace("_", " ").title()
        version = str(row.get("software_version", "")).strip()
        label = f"{adapter} {version}" if version else adapter
        key = (str(row["adapter"]), str(row.get("software_version", "")))
        if adapter_version_counts[key] > 1:
            profile = str(row.get("profile", "")).replace("_", " ")
            label = f"{label} — {profile}"
        labels.append(label)
    return labels


def _plot_support(summary: Sequence[Mapping[str, Any]], output: Path) -> dict[str, Any]:
    plt = _pyplot()

    is_ibsi1 = all(
        row.get("specification") == "IBSI 1"
        and row.get("profile") == "digital_phantom_3d_merged"
        for row in summary
    )
    is_ibsi2_phase2 = all(
        row.get("specification") == "IBSI 2" and row.get("phase") == "phase2"
        for row in summary
    )
    adapters = [str(row["adapter"]) for row in summary]
    labels = _summary_labels(summary)
    supported = [
        int(
            row.get("finite_designed_checks", row["supported_features"])
            if is_ibsi2_phase2
            else row["supported_features"]
        )
        for row in summary
    ]
    defined = [int(row["defined_features"]) for row in summary]
    fig, ax = plt.subplots(figsize=(9.5, max(4.5, 0.72 * len(adapters) + 2.0)))
    positions = list(range(len(adapters)))
    for position, adapter, value, denominator, row in zip(
        positions, adapters, supported, defined, summary
    ):
        color, _, hatch = ADAPTER_STYLES.get(adapter, ("#555555", "o", "//"))
        ax.barh(
            position,
            value,
            color=color,
            edgecolor="#111111",
            linewidth=1.2,
            hatch=hatch,
            height=0.62,
        )
        if is_ibsi1:
            direct_label = (
                f"defs {value}/{denominator}; inst "
                f"{int(row['finite_benchmark_instances'])}/"
                f"{int(row['defined_benchmark_instances'])}"
            )
        elif is_ibsi2_phase2:
            direct_label = (
                f"native {int(row['native_supported_filters'])}/"
                f"{int(row['defined_filters'])}; finite {value}/{denominator}"
            )
        else:
            direct_label = f"{value}/{denominator}"
        ax.text(
            value + max(0.5, max(defined, default=1) * 0.01),
            position,
            direct_label,
            va="center",
            fontsize=12,
        )
    ax.set_yticks(positions, labels)
    ax.invert_yaxis()
    ax.set_xlim(0, max(defined, default=1) * 1.28)
    if is_ibsi1:
        x_label = (
            "Finite semantic Table 2 definitions (labels also give workbook instances)"
        )
        title = "IBSI feature support by installed adapter"
    elif is_ibsi2_phase2:
        x_label = "Finite Phase 2 configuration-statistic checks (of 396 designed)"
        title = "IBSI 2 Phase 2 controlled-input component coverage"
    else:
        x_label = "Finite IBSI-defined semantic features"
        title = "IBSI feature support by installed adapter"
    ax.set_xlabel(x_label, fontsize=12)
    ax.set_title(title)
    ax.grid(axis="x", color="#cccccc", linestyle=":", linewidth=0.8)
    _style_axes(ax)
    fig.tight_layout()
    files = _save_figure(fig, output / "figure_feature_support")
    plt.close(fig)
    identity_notes = [
        (
            f"{str(row['adapter']).replace('_', ' ').title()} covers "
            f"{int(row['supported_features'])} definitions using "
            f"{int(row['distinct_source_outputs'])} distinct source outputs plus "
            f"{int(row['documented_exact_identity_definitions'])} documented exact identities."
        )
        for row in summary
        if not is_ibsi2_phase2
        and int(row.get("documented_exact_identity_definitions", 0)) > 0
    ]
    identity_disclosure = " " + " ".join(identity_notes) if identity_notes else ""
    if is_ibsi2_phase2:
        result_text = "; ".join(
            f"{label}: {value}/{denominator} finite"
            for label, value, denominator in zip(labels, supported, defined)
        )
        return {
            "id": "figure_feature_support",
            "files": files,
            "caption": (
                "Finite IBSI 2 Phase 2 configuration-statistic checks out of the 396-check "
                "design after protocol-controlled common preprocessing. This is a "
                "component-level native-filter/statistic comparison and does not test each "
                "adapter's native image-import or resampling path. The official analysis "
                "publishes 324 reference rows, of which 323 "
                "have standardized tolerances; 73 designed checks therefore have no "
                "standardized tolerance (72 unpublished design cells and one published "
                "nonstandardized row). Direct labels separately show the explicitly reviewed "
                "native filter coverage out of 22 filter configurations and finite calculation "
                "coverage out of 396 designed checks. "
                "Hatch patterns redundantly identify adapters. Results: "
                + result_text
                + "."
            ),
            "alt_text": (
                "Horizontal hatched bars show finite component-level IBSI 2 Phase 2 checks "
                "after controlled preprocessing; direct labels also show reviewed native "
                "filters out of 22. "
                + result_text
                + ". The published reference table contains 324 rows, 323 standardized; "
                "73 of the 396 designed checks lack a standardized tolerance."
            ),
        }
    return {
        "id": "figure_feature_support",
        "files": files,
        "caption": (
            "IBSI-defined semantic feature coverage from each installed adapter's calculated output surface. "
            + (
                "For IBSI 1, coverage requires finite digital-phantom values and uses the 172 Table 2 definitions. "
                "Direct labels separately show finite definitions out of 172 and finite workbook instances "
                "out of 174; numerical accuracy is assessed separately against 169 standardized instances. "
                if is_ibsi1
                else "Direct labels show finite semantic features over the defined feature denominator. "
            )
            + "Audited exact identities retain their shared native-output provenance and are not independent calculations. "
            "Hatch patterns redundantly identify adapters." + identity_disclosure
        ),
        "alt_text": (
            (
                "Horizontal hatched bars compare adapter coverage of 172 IBSI definitions. "
                "Direct labels also give finite workbook instances out of 174; 169 instances have standardized tolerances."
                if is_ibsi1
                else "Horizontal hatched bars compare finite IBSI semantic-feature coverage by adapter."
            )
            + identity_disclosure
        ),
    }


def _plot_outcomes(
    summary: Sequence[Mapping[str, Any]], output: Path
) -> dict[str, Any]:
    plt = _pyplot()

    is_ibsi1 = all(
        row.get("specification") == "IBSI 1"
        and row.get("profile") == "digital_phantom_3d_merged"
        for row in summary
    )
    is_ibsi2_phase2 = all(
        row.get("specification") == "IBSI 2" and row.get("phase") == "phase2"
        for row in summary
    )
    adapters = [str(row["adapter"]) for row in summary]
    labels = _summary_labels(summary)
    pass_values = [int(row["passed"]) for row in summary]
    fail_values = [int(row["failed"]) for row in summary]
    unevaluated = [
        int(row["defined_standardized_checks"])
        - int(row["passed"])
        - int(row["failed"])
        - int(row["unsupported"])
        for row in summary
    ]
    unsupported = [int(row["unsupported"]) for row in summary]
    series = (
        ("Pass", pass_values, "#287a3d", ""),
        ("Fail", fail_values, "#b22222", "xx"),
        ("Unevaluated", unevaluated, "#b8b8b8", ".."),
        (
            "Checks in native-unsupported configurations"
            if is_ibsi2_phase2
            else "Unsupported",
            unsupported,
            "#f0f0f0",
            "//",
        ),
    )
    fig, ax = plt.subplots(figsize=(9.5, 5.8))
    left = [0] * len(adapters)
    for label, values, color, hatch in series:
        bars = ax.barh(
            range(len(adapters)),
            values,
            left=left,
            label=label,
            color=color,
            edgecolor="#222222",
            linewidth=0.9,
            hatch=hatch,
        )
        for index, (bar, value) in enumerate(zip(bars, values)):
            if value:
                ax.text(
                    left[index] + value / 2.0,
                    bar.get_y() + bar.get_height() / 2.0,
                    str(value),
                    ha="center",
                    va="center",
                    fontsize=12,
                    color="#111111"
                    if label
                    in {
                        "Unevaluated",
                        "Unsupported",
                        "Checks in native-unsupported configurations",
                    }
                    else "white",
                )
        left = [a + b for a, b in zip(left, values)]
    ax.set_yticks(range(len(adapters)), labels)
    ax.invert_yaxis()
    ax.set_xlabel("Standardized reference checks", fontsize=12)
    ax.set_title("IBSI tolerance outcomes and unevaluated checks")
    ax.legend(fontsize=12, frameon=True, edgecolor="#333333", ncol=2)
    ax.grid(axis="x", color="#cccccc", linestyle=":", linewidth=0.8)
    _style_axes(ax)
    fig.tight_layout()
    files = _save_figure(fig, output / "figure_tolerance_outcomes")
    plt.close(fig)
    denominator_context = (
        " For the digital phantom, 169 of 174 workbook instances have standardized "
        "tolerances; those instances represent 172 Table 2 definitions."
        if is_ibsi1
        else (
            " Phase 2 defines 396 configuration-statistic checks; the published "
            "reference table contains 324 rows, 323 standardized and one "
            "nonstandardized. Together with 72 unpublished design cells, 73 "
            "designed checks have no standardized tolerance."
            if is_ibsi2_phase2
            else ""
        )
    )
    phase2_results = "; ".join(
        f"{label}: {passed} passed, {failed} failed, "
        f"{not_evaluated} unevaluated, {not_supported} unavailable"
        for label, passed, failed, not_evaluated, not_supported in zip(
            labels, pass_values, fail_values, unevaluated, unsupported
        )
    )
    return {
        "id": "figure_tolerance_outcomes",
        "files": files,
        "caption": (
            (
                "Counts of standardized checks that passed, failed, belonged to an explicitly native-unsupported filter, or could not be evaluated. "
                if is_ibsi2_phase2
                else "Counts of standardized checks that passed, failed, were unsupported, or could not be evaluated. "
            )
            + "Counts are printed within outlined, patterned segments."
            + denominator_context
            + (
                " These outcomes assess native filters and statistics after "
                "protocol-controlled common preprocessing."
                if is_ibsi2_phase2
                else ""
            )
        ),
        "alt_text": (
            "Stacked patterned bars compare adapter outcomes among 169 standardized digital-phantom instances, "
            "drawn from 174 workbook instances representing 172 IBSI definitions; segment counts are printed."
            if is_ibsi1
            else (
                "Stacked patterned bars show outcomes among 323 standardized IBSI 2 Phase 2 checks. "
                + phase2_results
                + ". The reference table publishes 324 of 396 designed checks; 73 designed checks lack a standardized tolerance."
                if is_ibsi2_phase2
                else "Stacked patterned bars show pass, fail, unevaluated, and unsupported standardized IBSI checks for every adapter, with counts printed in segments."
            )
        ),
    }


def _matrix_dimension(records: Sequence[ComparisonRecord]) -> tuple[str, list[str]]:
    families = sorted({record.family for record in records if record.standardized})
    configurations = sorted(
        {record.configuration for record in records if record.standardized}
    )
    if len(families) > 1 and len(families) <= 12:
        return "family", families
    return "configuration", configurations


def _plot_matrix(records: Sequence[ComparisonRecord], output: Path) -> dict[str, Any]:
    import numpy as np

    plt = _pyplot()

    dimension, columns = _matrix_dimension(records)
    row_keys = sorted(
        {
            (record.adapter, record.software_version, record.profile)
            for record in records
        }
    )
    adapter_version_counts = Counter(
        (adapter, software_version) for adapter, software_version, _ in row_keys
    )
    row_labels = [
        (
            f"{adapter.title()} {software_version}"
            if adapter_version_counts[(adapter, software_version)] == 1
            else (f"{adapter.title()} {software_version} — {profile.replace('_', ' ')}")
        ).strip()
        for adapter, software_version, profile in row_keys
    ]
    matrix = np.full((len(row_keys), len(columns)), np.nan)
    labels = np.empty(matrix.shape, dtype=object)
    for row_index, (adapter, software_version, profile) in enumerate(row_keys):
        for column_index, column in enumerate(columns):
            selected = [
                record
                for record in records
                if record.adapter == adapter
                and record.software_version == software_version
                and record.profile == profile
                and record.standardized
                and getattr(record, dimension) == column
            ]
            evaluated = sum(record.evaluated for record in selected)
            passed = sum(record.passed is True for record in selected)
            supported = sum(record.observed_supported for record in selected)
            matrix[row_index, column_index] = (
                passed / evaluated if evaluated else np.nan
            )
            labels[row_index, column_index] = (
                f"P {passed}/{evaluated}\nS {supported}/{len(selected)}"
            )

    fig_width = max(9.0, 1.15 * len(columns) + 3.0)
    fig, ax = plt.subplots(figsize=(fig_width, max(4.8, 0.8 * len(row_keys) + 2.5)))
    masked = np.ma.masked_invalid(matrix)
    cmap = plt.get_cmap("cividis").with_extremes(bad="#e6e6e6")
    image = ax.imshow(masked, cmap=cmap, vmin=0.0, vmax=1.0, aspect="auto")
    ax.set_facecolor("#eeeeee")
    ax.set_xticks(
        range(len(columns)),
        [value.replace("_", " ").upper() for value in columns],
        rotation=38,
        ha="right",
    )
    ax.set_yticks(range(len(row_keys)), row_labels)
    ax.set_xticks(np.arange(len(columns) + 1) - 0.5, minor=True)
    ax.set_yticks(np.arange(len(row_keys) + 1) - 0.5, minor=True)
    ax.grid(which="minor", color="#222222", linewidth=1.0)
    ax.tick_params(which="minor", bottom=False, left=False)
    for row_index in range(len(row_keys)):
        for column_index in range(len(columns)):
            value = matrix[row_index, column_index]
            color = "white" if math.isfinite(value) and value < 0.45 else "#111111"
            ax.text(
                column_index,
                row_index,
                labels[row_index, column_index],
                ha="center",
                va="center",
                fontsize=12,
                color=color,
                fontweight="bold",
            )
    colorbar = fig.colorbar(image, ax=ax, pad=0.02)
    colorbar.set_label("Passed / evaluated", fontsize=12)
    colorbar.ax.tick_params(labelsize=12)
    ax.set_xlabel(dimension.title(), fontsize=12)
    ax.set_ylabel("Adapter", fontsize=12)
    ax.set_title("IBSI support and conditional tolerance accuracy")
    _style_axes(ax)
    fig.tight_layout()
    files = _save_figure(fig, output / "figure_support_accuracy_matrix")
    plt.close(fig)
    return {
        "id": "figure_support_accuracy_matrix",
        "files": files,
        "caption": (
            f"Per-{dimension} support and conditional accuracy. Each cell states P passed/evaluated and "
            "S supported/standardized; grey cells have no evaluable result."
        ),
        "alt_text": f"Outlined matrix by adapter and {dimension}; each cell directly labels passed-over-evaluated and supported-over-standardized counts.",
    }


def _plot_ibsi2_phase2_matrix(
    records: Sequence[ComparisonRecord], output: Path
) -> dict[str, Any]:
    """Render Phase 2 as separate A/B design matrices without hiding gaps."""

    import numpy as np

    plt = _pyplot()
    row_keys = sorted(
        {
            (record.adapter, record.software_version, record.profile)
            for record in records
        }
    )
    duplicate_versions = Counter(
        (adapter, software_version) for adapter, software_version, _ in row_keys
    )
    row_labels = []
    for adapter, software_version, profile in row_keys:
        label = f"{adapter.title()} {software_version}".strip()
        if duplicate_versions[(adapter, software_version)] > 1:
            label += f" — {profile.replace('_', ' ')}"
        row_labels.append(label)

    fig, axes = plt.subplots(
        2,
        1,
        figsize=(21.5, max(9.5, 1.55 * len(row_keys) + 6.0)),
        sharey=True,
    )
    cmap = plt.get_cmap("cividis").with_extremes(bad="#e6e6e6")
    last_image = None
    published = set(IBSI2_PHASE2_PUBLISHED_FILTER_IDS)
    for ax, dimension, dimension_label in zip(
        axes,
        ("A", "B"),
        (
            "IBSI protocol preprocessing variant A",
            "IBSI protocol preprocessing variant B",
        ),
    ):
        configurations = [
            filter_id
            for filter_id in IBSI2_PHASE2_DEFINED_FILTER_IDS
            if filter_id.endswith(f".{dimension}")
        ]
        matrix = np.full((len(row_keys), len(configurations)), np.nan)
        cell_labels = np.empty(matrix.shape, dtype=object)
        for row_index, (adapter, software_version, profile) in enumerate(row_keys):
            for column_index, configuration in enumerate(configurations):
                selected = [
                    record
                    for record in records
                    if record.adapter == adapter
                    and record.software_version == software_version
                    and record.profile == profile
                    and record.configuration == configuration
                ]
                finite = sum(record.finite for record in selected)
                native_status = _ibsi2_phase2_native_support_status(selected)
                native_label = (
                    "yes"
                    if native_status == "supported"
                    else ("no" if native_status == "unsupported" else "?")
                )
                standardized = [record for record in selected if record.standardized]
                evaluated = sum(record.evaluated for record in standardized)
                passed = sum(record.passed is True for record in standardized)
                if configuration not in published:
                    cell_labels[row_index, column_index] = (
                        f"N {native_label}\nF {finite}/{len(selected)}\nP n/a"
                    )
                    continue
                matrix[row_index, column_index] = (
                    passed / evaluated if evaluated else np.nan
                )
                cell_labels[row_index, column_index] = (
                    f"N {native_label}\nF {finite}/{len(selected)}\n"
                    f"P {passed}/{evaluated}"
                )

        last_image = ax.imshow(
            np.ma.masked_invalid(matrix),
            cmap=cmap,
            vmin=0.0,
            vmax=1.0,
            aspect="auto",
        )
        ax.set_facecolor("#e6e6e6")
        ax.set_xticks(
            range(len(configurations)),
            [configuration.replace(".", "") for configuration in configurations],
        )
        ax.set_yticks(range(len(row_keys)), row_labels)
        ax.set_xticks(np.arange(len(configurations) + 1) - 0.5, minor=True)
        ax.set_yticks(np.arange(len(row_keys) + 1) - 0.5, minor=True)
        ax.grid(which="minor", color="#222222", linewidth=1.0)
        ax.tick_params(which="minor", bottom=False, left=False)
        for row_index in range(len(row_keys)):
            for column_index in range(len(configurations)):
                value = matrix[row_index, column_index]
                color = "white" if math.isfinite(value) and value < 0.45 else "#111111"
                ax.text(
                    column_index,
                    row_index,
                    cell_labels[row_index, column_index],
                    ha="center",
                    va="center",
                    fontsize=12,
                    color=color,
                    fontweight="bold",
                )
        ax.set_xlabel("Phase 2 filter configuration", fontsize=12)
        ax.set_ylabel("Adapter", fontsize=12)
        ax.set_title(dimension_label)
        _style_axes(ax)

    fig.suptitle(
        "IBSI 2 Phase 2 controlled-preprocessing component evaluation",
        fontsize=15,
    )
    fig.subplots_adjust(left=0.16, right=0.89, top=0.91, bottom=0.08, hspace=0.45)
    if last_image is not None:
        colorbar_axis = fig.add_axes([0.915, 0.18, 0.012, 0.64])
        colorbar = fig.colorbar(last_image, cax=colorbar_axis)
        colorbar.set_label("Passed / evaluated standardized checks", fontsize=12)
        colorbar.ax.tick_params(labelsize=12)
    files = _save_figure(fig, output / "figure_support_accuracy_matrix")
    plt.close(fig)

    result_parts = []
    for label, (adapter, software_version, profile) in zip(row_labels, row_keys):
        selected = [
            record
            for record in records
            if record.adapter == adapter
            and record.software_version == software_version
            and record.profile == profile
        ]
        standardized = [record for record in selected if record.standardized]
        native_filters = _ibsi2_phase2_native_supported_filters(selected)
        passed = sum(record.passed is True for record in standardized)
        result_parts.append(
            f"{label}: {len(native_filters)}/22 native, "
            f"{sum(record.finite for record in selected)}/396 finite, "
            f"{passed}/{sum(record.evaluated for record in standardized)} "
            f"passed/evaluated, {passed}/323 overall standardized success"
        )
    result_text = "; ".join(result_parts)
    return {
        "id": "figure_support_accuracy_matrix",
        "files": files,
        "caption": (
            "IBSI 2 Phase 2 component-level native-filter/statistic results after "
            "protocol-controlled common preprocessing, split into the slice-wise 2D "
            "A pathway and volumetric 3D B pathway. Configuration 5B is the specified "
            "exception and averages native 2D Gabor responses over three orthogonal "
            "planes. "
            "Each cell states N yes/no for reviewed native support, F finite/18 "
            "designed calculations, and P passed/evaluated standardized checks. "
            "N ? marks an incomplete support declaration. Configurations 10A, 10B, 11A, "
            "and 11B account for 72 defined checks without published reference rows "
            "and state finite calculation coverage with P n/a rather than hiding "
            "the candidate results. The remaining published "
            "nonstandardized row gives 73 of 396 designed checks without a "
            "standardized tolerance. Results: " + result_text + "."
        ),
        "alt_text": (
            "Two outlined matrices show the slice-wise 2D A pathway and volumetric "
            "3D B pathway for each adapter. Cells directly label native support, "
            "finite calculations, and passed-over-evaluated accuracy; 5B is the "
            "specified exception and uses two-dimensional Gabor responses averaged "
            "over orthogonal planes. "
            + result_text
            + ". Overall denominators are 396 defined checks, 324 published rows, "
            "323 standardized checks, and 73 checks without standardized tolerances."
        ),
    }


def _plot_ibsi2_phase2_b_matrix(
    records: Sequence[ComparisonRecord], output: Path
) -> dict[str, Any]:
    """Render the primary published Configuration-B numerical scope."""

    import numpy as np

    plt = _pyplot()
    row_keys = sorted(
        {
            (record.adapter, record.software_version, record.profile)
            for record in records
        }
    )
    duplicate_versions = Counter(
        (adapter, software_version) for adapter, software_version, _ in row_keys
    )
    row_labels = []
    for adapter, software_version, profile in row_keys:
        label = f"{adapter.title()} {software_version}".strip()
        if duplicate_versions[(adapter, software_version)] > 1:
            label += f" — {profile.replace('_', ' ')}"
        row_labels.append(label)

    configurations = list(_PHASE2_B_PUBLISHED_FILTER_IDS)
    matrix = np.zeros((len(row_keys), len(configurations)), dtype=float)
    cell_labels = np.empty(matrix.shape, dtype=object)
    for row_index, (adapter, software_version, profile) in enumerate(row_keys):
        for column_index, configuration in enumerate(configurations):
            selected = [
                record
                for record in records
                if record.adapter == adapter
                and record.software_version == software_version
                and record.profile == profile
                and record.configuration == configuration
            ]
            native_status = _ibsi2_phase2_native_support_status(selected)
            native_label = (
                "yes"
                if native_status == "supported"
                else ("no" if native_status == "unsupported" else "?")
            )
            standardized = [record for record in selected if record.standardized]
            evaluated = sum(record.evaluated for record in standardized)
            passed = sum(record.passed is True for record in standardized)
            denominator = len(standardized)
            matrix[row_index, column_index] = (
                passed / denominator if denominator else 0.0
            )
            cell_labels[row_index, column_index] = (
                f"N {native_label}\n"
                f"F {sum(record.finite for record in selected)}/{len(selected)}\n"
                f"E {evaluated}/{denominator}\n"
                f"P {passed}/{denominator}"
            )

    column_labels = []
    for configuration in configurations:
        parameters = PHASE2_FILTER_SPECS_BY_ID[configuration].parameters
        family = str(parameters["filter"])
        label = _FILTER_LABELS[family]
        if family in {"wavelet", "simoncelli"}:
            label += f" L{int(parameters['level'])}"
        if configuration == "5.B":
            label += "†"
        column_labels.append(f"{configuration}\n{label}")

    fig, ax = plt.subplots(figsize=(18.5, max(5.8, 1.08 * len(row_keys) + 2.8)))
    image = ax.imshow(matrix, cmap="cividis", vmin=0.0, vmax=1.0, aspect="auto")
    ax.set_xticks(range(len(configurations)), column_labels)
    ax.set_yticks(range(len(row_keys)), row_labels)
    ax.set_xticks(np.arange(len(configurations) + 1) - 0.5, minor=True)
    ax.set_yticks(np.arange(len(row_keys) + 1) - 0.5, minor=True)
    ax.grid(which="minor", color="#222222", linewidth=1.0)
    ax.tick_params(which="minor", bottom=False, left=False)
    for row_index in range(len(row_keys)):
        for column_index in range(len(configurations)):
            value = matrix[row_index, column_index]
            ax.text(
                column_index,
                row_index,
                cell_labels[row_index, column_index],
                ha="center",
                va="center",
                fontsize=11,
                fontweight="bold",
                color="white" if value < 0.45 else "#111111",
            )
    colorbar = fig.colorbar(image, ax=ax, pad=0.015)
    colorbar.set_label("Passed / fixed Configuration-B reference denominator")
    colorbar.ax.tick_params(labelsize=12)
    ax.set_xlabel("Published IBSI 2 Configuration-B filter", fontsize=12)
    ax.set_ylabel("Adapter", fontsize=12)
    ax.set_title("IBSI 2 Phase 2 volumetric Configuration-B conformity")
    _style_axes(ax)
    fig.tight_layout()
    files = _save_figure(fig, output / "figure_phase2_b_filter_matrix")
    plt.close(fig)

    result_text = "; ".join(
        (
            f"{row['adapter'].title()} {row['software_version']}: "
            f"{row['native_supported_b_published_filters']}/9 native, "
            f"{row['finite_published_b_checks']}/162 finite, "
            f"{row['passed_standardized_checks']}/"
            f"{row['evaluated_standardized_checks']} passed/evaluated, "
            f"{row['passed_standardized_checks']}/161 overall"
        )
        for row in _ibsi2_phase2_b_summary(records)
    )
    return {
        "id": "figure_phase2_b_filter_matrix",
        "files": files,
        "caption": (
            "Primary IBSI 2 Phase 2 Configuration-B result. Each cell states "
            "N reviewed native support, F finite calculations out of 18 published "
            "statistics, E evaluated standardized statistics, and P passes over "
            "the fixed configuration-specific standardized denominator. Cell colour "
            "uses that fixed denominator, so unsupported and unevaluated values are "
            "not hidden by conditional accuracy. Configuration 5.B (dagger) is the "
            "specified exception: a 2D Gabor kernel is applied over three orthogonal "
            "planes within the volumetric B workflow. Configuration 8.B has 17 rather "
            "than 18 standardized values because `stat_qcod` has no published "
            "consensus tolerance. Results: " + result_text + "."
        ),
        "alt_text": (
            "Adapter-by-filter matrix for the nine published volumetric B "
            "configurations. Cells directly label native support, finite values, "
            "evaluations, and passes over fixed denominators. The Gabor 5.B column "
            "is marked as a two-dimensional three-plane exception. " + result_text + "."
        ),
    }


def _plot_ibsi1_family_heatmap(
    family_rows: Sequence[Mapping[str, Any]],
    output: Path,
    *,
    kind: str,
) -> dict[str, Any]:
    """Render a Table 2 ordered coverage or accuracy heatmap with counts."""

    import numpy as np

    plt = _pyplot()
    row_keys = sorted(
        {
            (
                str(row["adapter"]),
                str(row.get("software_version", "")),
                str(row.get("profile", "")),
            )
            for row in family_rows
        }
    )
    lookup = {
        (
            str(row["adapter"]),
            str(row.get("software_version", "")),
            str(row.get("profile", "")),
            str(row["family"]),
        ): row
        for row in family_rows
    }
    duplicate_versions = Counter((adapter, version) for adapter, version, _ in row_keys)
    row_labels = []
    for adapter, version, profile in row_keys:
        label = f"{adapter.title()} {version}".strip()
        if duplicate_versions[(adapter, version)] > 1:
            label += f" — {profile.replace('_', ' ')}"
        row_labels.append(label)

    matrix = np.full((len(row_keys), len(FAMILY_ORDER)), np.nan)
    labels = np.empty(matrix.shape, dtype=object)
    for row_index, (adapter, version, profile) in enumerate(row_keys):
        for column_index, family in enumerate(FAMILY_ORDER):
            row = lookup.get((adapter, version, profile, family))
            if kind == "coverage":
                numerator = int(row["finite_definitions"]) if row else 0
                native = int(row["native_supported_definitions"]) if row else 0
                denominator = IBSI1_TABLE2_FAMILY_DEFINITIONS[family]
                finite_instances = int(row["finite_benchmark_instances"]) if row else 0
                benchmark_instances = IBSI1_BENCHMARK_INSTANCES_BY_FAMILY[family]
                matrix[row_index, column_index] = numerator / denominator
                labels[row_index, column_index] = (
                    f"native {native}/{denominator}\n"
                    f"finite {numerator}/{denominator}\n"
                    f"inst {finite_instances}/{benchmark_instances}"
                )
            elif kind == "accuracy":
                passed = int(row["passed_checks"]) if row else 0
                evaluated = int(row["evaluated_checks"]) if row else 0
                standardized = IBSI1_STANDARDIZED_INSTANCES_BY_FAMILY[family]
                if evaluated:
                    matrix[row_index, column_index] = passed / evaluated
                labels[row_index, column_index] = (
                    f"{passed}/{evaluated}\nstd inst {standardized}"
                )
            elif kind == "overall":
                passed = int(row["passed_checks"]) if row else 0
                evaluated = int(row["evaluated_checks"]) if row else 0
                standardized = IBSI1_STANDARDIZED_INSTANCES_BY_FAMILY[family]
                matrix[row_index, column_index] = passed / standardized
                labels[row_index, column_index] = (
                    f"pass {passed}/{standardized}\neval {evaluated}/{standardized}"
                )
            else:
                raise ValueError(f"Unknown IBSI 1 heatmap kind: {kind}")

    fig, ax = plt.subplots(figsize=(18.5, max(5.5, 0.92 * len(row_keys) + 2.8)))
    cmap = plt.get_cmap("cividis").with_extremes(bad="#e6e6e6")
    image = ax.imshow(
        np.ma.masked_invalid(matrix),
        cmap=cmap,
        vmin=0.0,
        vmax=1.0,
        aspect="auto",
    )
    ax.set_facecolor("#e6e6e6")
    column_labels = [
        f"{FAMILY_LABELS[family]}\n(n={IBSI1_TABLE2_FAMILY_DEFINITIONS[family]})"
        for family in FAMILY_ORDER
    ]
    ax.set_xticks(range(len(FAMILY_ORDER)), column_labels, rotation=35, ha="right")
    ax.set_yticks(range(len(row_keys)), row_labels)
    ax.set_xticks(np.arange(len(FAMILY_ORDER) + 1) - 0.5, minor=True)
    ax.set_yticks(np.arange(len(row_keys) + 1) - 0.5, minor=True)
    ax.grid(which="minor", color="#222222", linewidth=1.0)
    ax.tick_params(which="minor", bottom=False, left=False)
    for row_index in range(len(row_keys)):
        for column_index in range(len(FAMILY_ORDER)):
            value = matrix[row_index, column_index]
            color = "white" if math.isfinite(value) and value < 0.45 else "#111111"
            ax.text(
                column_index,
                row_index,
                labels[row_index, column_index],
                ha="center",
                va="center",
                fontsize=12,
                color=color,
                fontweight="bold",
            )
    colorbar = fig.colorbar(image, ax=ax, pad=0.015)
    if kind == "coverage":
        colorbar.set_label("Finite semantic definitions / defined", fontsize=12)
        title = "IBSI 1 digital-phantom definition and instance coverage by family"
        stem = "figure_ibsi1_table2_coverage_heatmap"
        caption = (
            "Digital-phantom semantic feature coverage under the required 3D-merged profile. "
            "The 172 Table 2 definitions are represented by 174 workbook instances; 169 instances "
            "have standardized tolerances. "
            "Each cell separately gives native-supported and finite Table 2 definitions over the official "
            "family denominator and finite workbook instances over the family instance denominator. "
            "Native support is not inferred from a finite result. The IVH family therefore "
            "shows five definitions and seven instances; both 10% and 90% instances must be finite before "
            "either parameterised IVH definition is covered. Audited exact identities share an emitted "
            "source scalar and do not represent independent calculations."
        )
        alt = (
            "Adapter-by-family heatmap reports native-supported and finite coverage of 172 IBSI "
            "definitions derived from 174 workbook instances; cells directly label the counts, "
            "and 169 instances have standardized tolerances."
        )
    elif kind == "accuracy":
        colorbar.set_label("Passed / evaluated standardized checks", fontsize=12)
        title = "IBSI 1 digital-phantom tolerance accuracy by Table 2 family"
        stem = "figure_ibsi1_accuracy_heatmap"
        caption = (
            "Numerical accuracy against standardized digital-phantom references under the required "
            "3D-merged profile. Of 174 workbook instances representing 172 Table 2 definitions, "
            "169 have standardized tolerances. Each cell gives passed/evaluated checks and the full "
            "standardized family denominator; grey cells have no evaluable result."
        )
        alt = (
            "Adapter-by-family heatmap reports passed and evaluated checks among 169 standardized "
            "references, derived from 174 workbook instances representing 172 IBSI definitions."
        )
    else:
        colorbar.set_label("Passed / fixed standardized denominator", fontsize=12)
        title = "IBSI 1 overall standardized compatibility by Table 2 family"
        stem = "figure_ibsi1_overall_standardized_success_heatmap"
        caption = (
            "Overall standardized compatibility against the fixed digital-phantom family "
            "denominators. The 169 standardized checks come from 174 workbook instances "
            "representing 172 Table 2 definitions. Cell colour encodes passed checks divided by every standardized "
            "workbook instance in the family, so unsupported and unevaluated instances remain "
            "in the denominator. Cell labels separately state pass/standardized and "
            "evaluated/standardized. This complements, rather than replaces, the definition "
            "coverage and conditional passed/evaluated accuracy heatmaps."
        )
        alt = (
            "Adapter-by-family heatmap reports passed checks over each fixed standardized "
            "family denominator, with unsupported and unevaluated instances retained; labels "
            "also report how many standardized instances were evaluated. There are 169 "
            "standardized checks among 174 workbook instances representing 172 definitions."
        )
    colorbar.ax.tick_params(labelsize=12)
    ax.set_xlabel("IBSI Table 2 feature family", fontsize=12)
    ax.set_ylabel("Adapter", fontsize=12)
    ax.set_title(title)
    _style_axes(ax)
    fig.tight_layout()
    files = _save_figure(fig, output / stem)
    plt.close(fig)
    return {"id": stem, "files": files, "caption": caption, "alt_text": alt}


def _plot_tolerance_ratios(
    records: Sequence[ComparisonRecord], output: Path
) -> Optional[dict[str, Any]]:
    plt = _pyplot()

    evaluated = [
        record
        for record in records
        if record.evaluated
        and record.error_tolerance_ratio is not None
        and math.isfinite(record.error_tolerance_ratio)
    ]
    if not evaluated:
        return None
    adapters = sorted(
        {(record.adapter, record.software_version) for record in evaluated}
    )
    fig, ax = plt.subplots(figsize=(10.5, 6.2))
    for adapter_index, (adapter, software_version) in enumerate(adapters):
        selected = [
            record
            for record in evaluated
            if record.adapter == adapter and record.software_version == software_version
        ]
        selected.sort(
            key=lambda record: (
                record.configuration,
                record.family,
                record.semantic_key,
            )
        )
        ratios = [max(float(record.error_tolerance_ratio), 1e-6) for record in selected]
        x_values = [
            index + (adapter_index - (len(adapters) - 1) / 2.0) * 0.08
            for index in range(len(selected))
        ]
        color, marker, _ = ADAPTER_STYLES.get(adapter, ("#555555", "o", ""))
        label = f"{adapter.title()} {software_version}".strip()
        ax.scatter(
            x_values,
            ratios,
            label=label,
            color=color,
            marker=marker,
            s=34,
            edgecolors="#111111",
            linewidths=0.45,
            alpha=0.82,
        )
    ax.axhline(
        1.0, color="#111111", linestyle="--", linewidth=1.5, label="Tolerance boundary"
    )
    ax.set_yscale("log")
    ax.set_xlabel(
        "Deterministic feature/configuration order within adapter", fontsize=12
    )
    ax.set_ylabel("Comparison error / tolerance (log scale)", fontsize=12)
    ax.set_title("IBSI numerical error relative to tolerance")
    ax.grid(True, which="both", color="#cccccc", linestyle=":", linewidth=0.75)
    ax.legend(fontsize=12, frameon=True, edgecolor="#333333", ncol=2)
    _style_axes(ax)
    fig.tight_layout()
    files = _save_figure(fig, output / "figure_error_tolerance_ratio")
    plt.close(fig)
    is_ibsi1 = _is_ibsi1_digital_phantom(records)
    is_ibsi2_phase2 = _is_ibsi2_phase2(records)
    denominator_context = (
        " The 169 standardized checks come from 174 workbook instances representing "
        "172 Table 2 definitions."
        if is_ibsi1
        else (
            " IBSI 2 Phase 2 has 396 designed configuration-statistic checks, "
            "324 published reference rows, 323 standardized checks, and 73 "
            "designed checks without a standardized tolerance."
            if is_ibsi2_phase2
            else ""
        )
    )
    phase2_results = "; ".join(
        (
            f"{adapter.title()} {software_version}".strip()
            + f": {len(selected)} evaluated ratios, maximum {max(selected):.3g}"
        )
        for adapter, software_version in adapters
        for selected in [
            [
                float(record.error_tolerance_ratio)
                for record in evaluated
                if record.adapter == adapter
                and record.software_version == software_version
            ]
        ]
        if selected
    )
    return {
        "id": "figure_error_tolerance_ratio",
        "files": files,
        "caption": (
            "Evaluated numerical errors divided by their source tolerance. Marker shape and colour identify "
            "adapter versions; the dashed ratio-one line is the pass boundary. Exact zero-error checks with "
            "zero tolerance use the lower display floor; failed zero-tolerance checks have no finite ratio "
            "and remain explicit in the outcome figure and detail table."
            + denominator_context
        ),
        "alt_text": (
            "Log scatter compares error-to-tolerance ratios for 169 standardized checks from 174 "
            "workbook instances representing 172 definitions; shapes identify adapters and a dashed line marks one."
            if is_ibsi1
            else (
                "Log-scale scatter plot of IBSI 2 Phase 2 error-to-tolerance ratios; "
                + phase2_results
                + ". Shapes identify adapters and a dashed horizontal line marks ratio one. "
                "There are 323 standardized checks among 324 published rows and 396 designed checks."
                if is_ibsi2_phase2
                else "Log-scale scatter plot compares adapter-version error-to-tolerance ratios; distinct marker shapes identify adapters and a dashed horizontal line marks ratio one."
            )
        ),
    }


def _markdown_summary(
    summary: Sequence[Mapping[str, Any]],
    records: Sequence[ComparisonRecord] = (),
) -> str:
    digital_phantom_only = all(
        row.get("specification") == "IBSI 1"
        and row.get("profile") == "digital_phantom_3d_merged"
        for row in summary
    )
    ibsi2_phase2_only = all(
        row.get("specification") == "IBSI 2" and row.get("phase") == "phase2"
        for row in summary
    )
    if ibsi2_phase2_only:
        row_complete = _ibsi2_phase2_row_complete(records)
        b_summary = _ibsi2_phase2_b_summary(records)
        lines = [
            "# IBSI 2 Phase 2 compliance summary",
            "",
            (
                "**Publication execution status: complete exhaustive native-support "
                "design; every native-supported configuration was processed.**"
                if row_complete
                else "**Publication execution status: incomplete; one or more native-supported "
                "configurations were not supplied or failed, or native support was not "
                "exhaustively declared. Do not report adapter "
                "accuracy as a complete IBSI 2 result.**"
            ),
            "",
            (
                "Phase 2 defines 396 configuration-statistic checks. The published "
                "analysis table contains 324 rows, of which 323 have standardized "
                "consensus values and tolerances. Thus, 73 designed checks lack a "
                "standardized tolerance: 72 unpublished design cells and one "
                "published nonstandardized row. Rates are never reported without "
                "their numerator and denominator."
            ),
            "",
            (
                "Scope: this is a component-level native-filter/statistic evaluation "
                "after protocol-controlled common preprocessing. It does not validate "
                "each adapter's native image-import or resampling implementation. "
                "IBSI defines A as the slice-wise 2D pathway and B as the volumetric "
                "3D pathway; configuration 5B is the specified exception and averages "
                "native 2D Gabor responses over three orthogonal planes."
            ),
            "",
            "## Primary volumetric Configuration-B result",
            "",
            (
                "The 3D-focused result retains all 11 designed B configurations for "
                "native/API coverage and uses the 9 published B configurations for "
                "numerical comparison. Those 9 configurations contain 162 published "
                "statistic rows and 161 standardized tolerances; `8.B/stat_qcod` is "
                "the single published row without a standardized reference."
            ),
            "",
            (
                "| Adapter | Version | Native B / 11 designed | Native B / 9 published | "
                "Finite / 162 published | Passed / evaluated | Passed / 161 standardized |"
            ),
            "| --- | --- | ---: | ---: | ---: | ---: | ---: |",
        ]
        for row in b_summary:
            lines.append(
                f"| {row['adapter']} | {row['software_version']} | "
                f"{row['native_supported_b_designed_filters']}/{row['defined_b_filters']} | "
                f"{row['native_supported_b_published_filters']}/{row['published_b_filters']} | "
                f"{row['finite_published_b_checks']}/{row['published_b_reference_rows']} | "
                f"{row['passed_standardized_checks']}/{row['evaluated_standardized_checks']} | "
                f"{row['passed_standardized_checks']}/{row['standardized_b_checks']} |"
            )
        lines.extend(
            [
                "",
                (
                    "Configuration 5.B belongs to the volumetric B workflow but is "
                    "not a native 3D Gabor kernel: IBSI specifies 2D Gabor responses "
                    "averaged over three orthogonal planes, followed by statistics "
                    "over the complete 3D ROI."
                ),
                "",
                "## Secondary full A/B design",
                "",
                (
                    "| Adapter | Version | Native filters / 22 | Finite / 396 designed | Finite / 324 published | "
                    "Passed / evaluated | Passed / 323 standardized |"
                ),
                "| --- | --- | ---: | ---: | ---: | ---: | ---: |",
            ]
        )
        for row in summary:
            lines.append(
                f"| {row['adapter']} | {row['software_version']} | "
                f"{row['native_supported_filters']}/{row['defined_filters']} | "
                f"{row['finite_designed_checks']}/{IBSI2_PHASE2_DEFINED_CHECKS} | "
                f"{row['finite_published_reference_rows']}/{IBSI2_PHASE2_PUBLISHED_REFERENCE_ROWS} | "
                f"{row['passed']}/{row['evaluated']} | "
                f"{row['passed']}/{IBSI2_PHASE2_STANDARDIZED_CHECKS} |"
            )
        lines.extend(
            [
                "",
                (
                    "Native filter coverage comes only from the exhaustive reviewed "
                    "support declarations. A finite check separately means that calculation "
                    "from a native-supported response map yielded a finite statistic."
                ),
                "",
                (
                    "Metric definitions: native support is reviewed supported "
                    "configurations/22; finite calculation coverage is finite results/396; "
                    "conditional accuracy is passed/evaluated standardized checks; overall "
                    "standardized success is passed/323, retaining unsupported and "
                    "unevaluated standardized checks in the denominator."
                ),
                "",
                (
                    "Unsupported or unsupplied configurations, missing or non-finite "
                    "values, ambiguous mappings, and nonstandardized rows remain in "
                    "the detail table and are never counted as passes."
                ),
                "",
            ]
        )
        return "\n".join(lines)
    instance_heading = (
        "Finite workbook instances"
        if digital_phantom_only
        else "Finite benchmark instances"
    )
    lines = [
        "# IBSI compliance summary",
        "",
        (
            "Rates are never reported without their numerator and denominator. "
            "For IBSI 1, definition coverage and workbook-instance coverage are "
            "reported separately: 172 Table 2 definitions produce 174 digital-phantom "
            "instances, of which 169 have standardized tolerance references."
            if digital_phantom_only
            else "Rates are never reported without their numerator and denominator."
        ),
        "",
        f"| Adapter | Version | Native semantic definitions | Finite semantic definitions | Distinct source outputs | Exact identities | {instance_heading} | Finite / supported checks | Passed / evaluated | Passed / referencable |",
        "| --- | --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
    ]
    for row in summary:
        lines.append(
            "| {adapter} | {software_version} | {native_supported_definitions}/{defined_features} | "
            "{finite_definitions}/{defined_features} | "
            "{distinct_source_outputs} | {documented_exact_identity_definitions} | "
            "{finite_benchmark_instances}/{defined_benchmark_instances} | {finite}/{supported_checks} | "
            "{passed}/{evaluated} | {passed}/{referencable} |".format(**row)
        )
    lines.extend(
        [
            "",
            "Unsupported, missing, non-finite, ambiguous, and not-standardized rows remain in the detail table and are never counted as passes.",
            "",
        ]
    )
    pyradiomics_supported = {
        record.semantic_key
        for record in records
        if record.adapter == "pyradiomics"
        and record.observed_supported
        and record.semantic_key
    }
    pyradiomics_native = {
        name.replace(" [documented exact alias]", "")
        for record in records
        if record.adapter == "pyradiomics" and record.observed_supported
        for name in record.native_feature_names.split(", ")
        if name
    }
    pyradiomics_aliases = {
        record.semantic_key
        for record in records
        if record.adapter == "pyradiomics"
        and "[documented exact alias]" in record.native_feature_names
        and record.semantic_key
    }
    if pyradiomics_aliases:
        lines.extend(
            [
                (
                    "PyRadiomics semantic coverage comprises "
                    f"{len(pyradiomics_native)} distinct named native outputs plus "
                    f"{len(pyradiomics_aliases)} documented exact identities "
                    f"({len(pyradiomics_supported)} definitions total). The aliases share "
                    "active source scalars and are not independent calculations."
                ),
                "",
            ]
        )
    return "\n".join(lines)


def _write_xlsx(
    output_path: Path,
    detail: Sequence[Mapping[str, Any]],
    summary: Sequence[Mapping[str, Any]],
    family: Sequence[Mapping[str, Any]],
    *,
    extra_tables: Sequence[tuple[str, Sequence[Mapping[str, Any]]]] = (),
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to generate the declared XLSX compliance tables"
        ) from exc
    workbook = Workbook()
    workbook.remove(workbook.active)

    def excel_value(value: Any) -> Any:
        """Convert structured audit fields to deterministic, readable cell text."""

        if isinstance(value, (Mapping, list, tuple)):
            return json.dumps(
                value,
                ensure_ascii=False,
                sort_keys=True,
                default=str,
            )
        return value

    tables = [
        ("Summary", summary),
        ("By family", family),
        *extra_tables,
        ("Detail", detail),
    ]
    for name, rows in tables:
        sheet = workbook.create_sheet(name)
        fields = list(rows[0]) if rows else []
        sheet.append(fields)
        for row in rows:
            sheet.append([excel_value(row.get(field)) for field in fields])
        for cell in sheet[1]:
            cell.font = Font(name="Arial", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        for column in sheet.columns:
            width = min(
                50, max(12, max(len(str(cell.value or "")) for cell in column) + 2)
            )
            sheet.column_dimensions[column[0].column_letter].width = width
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(output_path)
    finally:
        workbook.close()


def generate_compliance_report(
    records: Sequence[ComparisonRecord],
    output_dir: Path,
    *,
    source_metadata: Optional[Mapping[str, Any]] = None,
) -> dict[str, Any]:
    if not records:
        raise ValueError("At least one comparison record is required")
    output = Path(output_dir).expanduser().resolve()
    output.mkdir(parents=True, exist_ok=True)
    detail_rows = [record.to_dict() for record in records]
    summary = _overall_summary(records)
    family = _family_summary(records)
    is_ibsi1 = _is_ibsi1_digital_phantom(records)
    is_ibsi2_phase2 = _is_ibsi2_phase2(records)
    configuration = (
        _ibsi2_phase2_configuration_summary(records) if is_ibsi2_phase2 else []
    )
    b_summary = _ibsi2_phase2_b_summary(records) if is_ibsi2_phase2 else []
    b_configuration = (
        _ibsi2_phase2_b_configuration_summary(records) if is_ibsi2_phase2 else []
    )

    atomic_write_text(output / "compliance_detail.csv", comparison_to_csv(records))
    atomic_write_text(
        output / "compliance_summary.csv", _csv_text(summary, list(summary[0]))
    )
    atomic_write_text(
        output / "compliance_by_family.csv", _csv_text(family, list(family[0]))
    )
    if is_ibsi2_phase2:
        atomic_write_text(
            output / "compliance_by_configuration.csv",
            _csv_text(configuration, list(configuration[0])),
        )
        atomic_write_text(
            output / "compliance_phase2_b_summary.csv",
            _csv_text(b_summary, list(b_summary[0])),
        )
        atomic_write_text(
            output / "compliance_phase2_b_by_configuration.csv",
            _csv_text(b_configuration, list(b_configuration[0])),
        )
    atomic_write_text(
        output / "compliance_summary.md",
        _markdown_summary(summary, records),
    )
    dictionary_tables: list[tuple[str, Sequence[Mapping[str, Any]]]] = [
        ("compliance_summary.csv", summary),
        ("compliance_by_family.csv", family),
        ("compliance_detail.csv", detail_rows),
    ]
    if is_ibsi2_phase2:
        dictionary_tables[2:2] = [
            ("compliance_phase2_b_summary.csv", b_summary),
            ("compliance_phase2_b_by_configuration.csv", b_configuration),
            ("compliance_by_configuration.csv", configuration),
        ]
    data_dictionary = _data_dictionary_rows(dictionary_tables)
    atomic_write_text(
        output / "data_dictionary.csv",
        _csv_text(data_dictionary, ["table", "field", "definition"]),
    )
    extra_tables: list[tuple[str, Sequence[Mapping[str, Any]]]] = [
        ("Data dictionary", data_dictionary)
    ]
    if is_ibsi2_phase2:
        extra_tables[0:0] = [
            ("B summary", b_summary),
            ("B by configuration", b_configuration),
            ("By configuration", configuration),
        ]
    _write_xlsx(
        output / "compliance_tables.xlsx",
        detail_rows,
        summary,
        family,
        extra_tables=extra_tables,
    )

    figure_metadata = [
        _plot_support(summary, output),
        _plot_outcomes(summary, output),
    ]
    if is_ibsi1:
        figure_metadata.extend(
            [
                _plot_ibsi1_family_heatmap(family, output, kind="coverage"),
                _plot_ibsi1_family_heatmap(family, output, kind="accuracy"),
                _plot_ibsi1_family_heatmap(family, output, kind="overall"),
            ]
        )
    elif is_ibsi2_phase2:
        figure_metadata.extend(
            [
                _plot_ibsi2_phase2_b_matrix(records, output),
                _plot_ibsi2_phase2_matrix(records, output),
            ]
        )
    else:
        figure_metadata.append(_plot_matrix(records, output))
    ratio_figure = _plot_tolerance_ratios(records, output)
    if ratio_figure is not None:
        figure_metadata.append(ratio_figure)
    captions = ["# Figure captions", ""]
    if is_ibsi2_phase2:
        captions.extend(
            [
                "## Interpretation key",
                "",
                (
                    "These are component-level native-filter/statistic results after "
                    "protocol-controlled common preprocessing. N yes/no denotes the "
                    "reviewed native-support declaration, F denotes finite calculations "
                    "out of the designed denominator, and P denotes passed/evaluated "
                    "standardized checks in the full-design matrix. The primary B-only "
                    "matrix instead colours cells by passed/fixed standardized denominator; "
                    "the B-only overall result uses passed/161 and the full-design "
                    "secondary result uses passed/323. IBSI defines A as "
                    "the slice-wise 2D pathway and B as the volumetric 3D pathway; 5B "
                    "is the specified exception and averages native 2D Gabor responses "
                    "over three orthogonal planes."
                ),
                "",
            ]
        )
    for item in figure_metadata:
        captions.extend(
            [
                f"## {item['id']}",
                "",
                str(item["caption"]),
                "",
                f"Alt text: {item['alt_text']}",
                "",
            ]
        )
    atomic_write_text(output / "figure_captions.md", "\n".join(captions))
    table_artifacts = [
        ("compliance_detail.csv", "table"),
        ("compliance_summary.csv", "table"),
        ("compliance_by_family.csv", "table"),
        ("data_dictionary.csv", "data_dictionary"),
        ("compliance_tables.xlsx", "workbook"),
        ("compliance_summary.md", "narrative"),
        ("figure_captions.md", "captions"),
    ]
    if is_ibsi2_phase2:
        table_artifacts.extend(
            [
                ("compliance_phase2_b_summary.csv", "table"),
                ("compliance_phase2_b_by_configuration.csv", "table"),
                ("compliance_by_configuration.csv", "table"),
            ]
        )
    figure_artifacts = [
        (filename, "figure") for item in figure_metadata for filename in item["files"]
    ]
    manifest = {
        "schema_version": 1,
        "record_count": len(records),
        "source_metadata": dict(source_metadata or {}),
        "denominator_policy": {
            "feature_support": (
                "finite direct outputs or audited exact-identity semantic assignments / 172 definitions; "
                "direct-native source count and exact-alias provenance are retained separately; "
                "both workbook operating points are required for each parameterised IVH definition"
                if is_ibsi1
                else (
                    "Phase 2 field: finite configuration-statistic "
                    "results / 396 designed checks; do not interpret as native support"
                    if is_ibsi2_phase2
                    else "unique observed mapped semantic features / all unique IBSI-defined semantic features"
                )
            ),
            "ibsi1_table2_definitions": (
                IBSI1_TABLE2_DEFINITION_COUNT if is_ibsi1 else None
            ),
            "ibsi1_workbook_benchmark_instances": (
                IBSI1_PROFILE_ROWS if is_ibsi1 else None
            ),
            "workbook_instance_coverage": (
                "finite calculated digital-phantom workbook instances / 174 instances"
                if is_ibsi1
                else None
            ),
            "ibsi1_standardized_tolerance_checks": (
                IBSI1_PROFILE_STANDARDIZED if is_ibsi1 else None
            ),
            "ibsi2_phase2_defined_checks": (
                IBSI2_PHASE2_DEFINED_CHECKS if is_ibsi2_phase2 else None
            ),
            "ibsi2_phase2_published_reference_rows": (
                IBSI2_PHASE2_PUBLISHED_REFERENCE_ROWS if is_ibsi2_phase2 else None
            ),
            "ibsi2_phase2_standardized_checks": (
                IBSI2_PHASE2_STANDARDIZED_CHECKS if is_ibsi2_phase2 else None
            ),
            "ibsi2_phase2_nonstandardized_checks": (
                IBSI2_PHASE2_NONSTANDARDIZED_CHECKS if is_ibsi2_phase2 else None
            ),
            "ibsi2_phase2_b_defined_filters": (
                len(_PHASE2_B_DEFINED_FILTER_IDS) if is_ibsi2_phase2 else None
            ),
            "ibsi2_phase2_b_published_filters": (
                len(_PHASE2_B_PUBLISHED_FILTER_IDS) if is_ibsi2_phase2 else None
            ),
            "ibsi2_phase2_b_published_reference_rows": (
                _PHASE2_B_PUBLISHED_REFERENCE_ROWS if is_ibsi2_phase2 else None
            ),
            "ibsi2_phase2_b_standardized_checks": (
                _PHASE2_B_STANDARDIZED_CHECKS if is_ibsi2_phase2 else None
            ),
            "published_reference_coverage": (
                "finite results / 324 published Phase 2 rows"
                if is_ibsi2_phase2
                else None
            ),
            "phase2_b_native_filter_support": (
                "explicit reviewed native-supported declarations / 11 designed "
                "volumetric B configurations; the published subset is reported "
                "separately over 9 configurations"
                if is_ibsi2_phase2
                else None
            ),
            "phase2_b_finite_reference_coverage": (
                "finite results / 162 published volumetric B reference rows"
                if is_ibsi2_phase2
                else None
            ),
            "phase2_b_overall_referenced_success": (
                "passed / 161 standardized volumetric B checks"
                if is_ibsi2_phase2
                else None
            ),
            "native_filter_support": (
                "explicit reviewed native-supported filter declarations / 22 defined filters; "
                "never inferred from candidate-map availability or finite statistics"
                if is_ibsi2_phase2
                else None
            ),
            "finite_calculation_coverage": (
                "finite native-filter/statistic results / 396 designed checks after "
                "protocol-controlled common preprocessing"
                if is_ibsi2_phase2
                else None
            ),
            "evaluation_scope": (_PHASE2_COMPONENT_SCOPE if is_ibsi2_phase2 else None),
            "conditional_accuracy": "passed / evaluated",
            "overall_referenced_success": (
                "passed / 323 standardized Phase 2 checks"
                if is_ibsi2_phase2
                else "passed / referencable standardized checks"
            ),
            "unsupported_is_failure": False,
            "missing_is_pass": False,
            "exact_semantic_aliases_are_independent_calculations": False,
        },
        "accessibility": {
            "font_family": "DejaVu Sans",
            "minimum_label_size_points": 12,
            "redundant_cues": "direct labels, outlines, hatches, and marker shapes",
            "formats": ["PDF", "SVG", "PNG"],
        },
        "figures": figure_metadata,
        "tables": [
            "compliance_detail.csv",
            "compliance_summary.csv",
            "compliance_by_family.csv",
            *(
                [
                    "compliance_phase2_b_summary.csv",
                    "compliance_phase2_b_by_configuration.csv",
                    "compliance_by_configuration.csv",
                ]
                if is_ibsi2_phase2
                else []
            ),
            "data_dictionary.csv",
            "compliance_tables.xlsx",
            "compliance_summary.md",
        ],
        "data_dictionary": "data_dictionary.csv",
        "figure_captions": "figure_captions.md",
        "release_manifest": "release_manifest.json",
    }
    if is_ibsi2_phase2:
        manifest["evaluation_scope"] = _PHASE2_COMPONENT_SCOPE
        manifest["protocol_variant_policy"] = (
            "A is the IBSI slice-wise 2D pathway and B is the volumetric 3D "
            "pathway; 5B is the specified exception, using native 2D Gabor "
            "responses averaged over three orthogonal planes."
        )
        manifest["publication_execution_complete"] = _ibsi2_phase2_row_complete(records)
        manifest["native_supported_filter_configurations"] = {
            str(row["adapter"]): int(row["native_supported_filters"]) for row in summary
        }
        manifest["native_filter_denominator"] = len(IBSI2_PHASE2_DEFINED_FILTER_IDS)
        manifest["phase2_b_scope"] = {
            "workflow_dimensionality": 3,
            "defined_filter_configurations": list(_PHASE2_B_DEFINED_FILTER_IDS),
            "published_filter_configurations": list(_PHASE2_B_PUBLISHED_FILTER_IDS),
            "defined_filter_count": len(_PHASE2_B_DEFINED_FILTER_IDS),
            "published_filter_count": len(_PHASE2_B_PUBLISHED_FILTER_IDS),
            "published_reference_rows": _PHASE2_B_PUBLISHED_REFERENCE_ROWS,
            "standardized_checks": _PHASE2_B_STANDARDIZED_CHECKS,
            "nonstandardized_published_pair": {
                "configuration": IBSI2_PHASE2_NONSTANDARDIZED_PUBLISHED_PAIR[0],
                "feature_tag": IBSI2_PHASE2_NONSTANDARDIZED_PUBLISHED_PAIR[1],
            },
            "gabor_5b_exception": (
                "5.B belongs to the volumetric B workflow but is defined with "
                "native 2D Gabor kernels averaged over three orthogonal planes."
            ),
        }
        manifest["native_supported_b_designed_filter_configurations"] = {
            str(row["adapter"]): int(row["native_supported_b_designed_filters"])
            for row in b_summary
        }
        manifest["native_supported_b_published_filter_configurations"] = {
            str(row["adapter"]): int(row["native_supported_b_published_filters"])
            for row in b_summary
        }
    if not is_ibsi2_phase2:
        for key in (
            "ibsi2_phase2_defined_checks",
            "ibsi2_phase2_published_reference_rows",
            "ibsi2_phase2_standardized_checks",
            "ibsi2_phase2_nonstandardized_checks",
            "ibsi2_phase2_b_defined_filters",
            "ibsi2_phase2_b_published_filters",
            "ibsi2_phase2_b_published_reference_rows",
            "ibsi2_phase2_b_standardized_checks",
            "published_reference_coverage",
            "phase2_b_native_filter_support",
            "phase2_b_finite_reference_coverage",
            "phase2_b_overall_referenced_success",
            "native_filter_support",
            "finite_calculation_coverage",
            "evaluation_scope",
        ):
            manifest["denominator_policy"].pop(key)
    release_artifacts = [
        *table_artifacts,
        *figure_artifacts,
        ("report_manifest.json", "manifest"),
    ]
    manifest["release_artifact_count"] = len(
        {relative_name for relative_name, _ in release_artifacts}
    )
    atomic_write_json(output / "report_manifest.json", manifest)
    release_manifest = _write_release_manifest(output, release_artifacts)
    if len(release_manifest["artifacts"]) != manifest["release_artifact_count"]:
        raise RuntimeError("Release artifact count changed during manifesting")
    return manifest


def _phase1_candidate_supplied(row: Mapping[str, Any]) -> bool:
    return _bool(row.get("candidate_supplied"))


def _phase1_native_supported(row: Mapping[str, Any]) -> bool:
    return _bool(row.get("native_supported"))


def _ibsi2_phase1_row_complete(rows: Sequence[Mapping[str, Any]]) -> bool:
    """Require an exhaustive support grid and processed native-supported maps."""

    grouped: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[str(row.get("adapter", ""))].append(row)
    expected_ids = set(IBSI2_PHASE1_TEST_IDS)
    return bool(grouped) and all(
        {str(row.get("test_id", "")) for row in values} == expected_ids
        and all("native_supported" in row for row in values)
        and all(
            not _phase1_native_supported(row)
            or (
                _phase1_candidate_supplied(row)
                and (not _bool(row.get("standardized")) or _bool(row.get("evaluated")))
            )
            for row in values
        )
        for values in grouped.values()
    )


def _phase1_summary(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[str(row["adapter"])].append(row)
    output = []
    for adapter, values in sorted(grouped.items()):
        standardized = [row for row in values if _bool(row.get("standardized"))]
        evaluated = [row for row in standardized if _bool(row.get("evaluated"))]
        passed = sum(
            row.get("passed") is True or str(row.get("passed")).lower() == "true"
            for row in evaluated
        )
        output.append(
            {
                "adapter": adapter,
                "generator_distribution": ", ".join(
                    sorted(
                        {
                            str(row.get("generator_distribution", ""))
                            for row in values
                            if str(row.get("generator_distribution", ""))
                        }
                    )
                ),
                "generator_version": ", ".join(
                    sorted(
                        {
                            str(row.get("generator_version", ""))
                            for row in values
                            if str(row.get("generator_version", ""))
                        }
                    )
                ),
                "defined_filter_tests": len(IBSI2_PHASE1_TEST_IDS),
                "native_supported_filter_tests": sum(
                    _phase1_native_supported(row) for row in values
                ),
                "candidate_maps_supplied": sum(
                    _phase1_candidate_supplied(row) for row in values
                ),
                "standardized_reference_tests": len(IBSI2_PHASE1_STANDARDIZED_IDS),
                "referencable": sum(
                    _bool(row.get("referencable")) for row in standardized
                ),
                "evaluated": len(evaluated),
                "passed": passed,
                "failed": len(evaluated) - passed,
                "conditional_accuracy": passed / len(evaluated) if evaluated else None,
                "overall_referenced_success": (
                    passed / len(IBSI2_PHASE1_STANDARDIZED_IDS)
                ),
            }
        )
    return output


def _phase1_strict_3d_summary(
    rows: Sequence[Mapping[str, Any]],
) -> list[dict[str, Any]]:
    """Summarize the 27 IBSI-defined 3D tests and 24 reference maps."""

    strict_ids = set(_PHASE1_STRICT_3D_TEST_IDS)
    standardized_ids = set(_PHASE1_STRICT_3D_STANDARDIZED_IDS)
    grouped: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[str(row["adapter"])].append(row)

    output = []
    for adapter, adapter_rows in sorted(grouped.items()):
        values = [
            row for row in adapter_rows if str(row.get("test_id", "")) in strict_ids
        ]
        standardized = [
            row for row in values if str(row["test_id"]) in standardized_ids
        ]
        evaluated = [row for row in standardized if _bool(row.get("evaluated"))]
        passed = sum(
            row.get("passed") is True or str(row.get("passed")).lower() == "true"
            for row in evaluated
        )
        output.append(
            {
                "adapter": adapter,
                "generator_distribution": ", ".join(
                    sorted(
                        {
                            str(row.get("generator_distribution", ""))
                            for row in adapter_rows
                            if str(row.get("generator_distribution", ""))
                        }
                    )
                ),
                "generator_version": ", ".join(
                    sorted(
                        {
                            str(row.get("generator_version", ""))
                            for row in adapter_rows
                            if str(row.get("generator_version", ""))
                        }
                    )
                ),
                "defined_filter_tests": len(_PHASE1_STRICT_3D_TEST_IDS),
                "native_supported_filter_tests": sum(
                    _phase1_native_supported(row) for row in values
                ),
                "candidate_maps_supplied": sum(
                    _phase1_candidate_supplied(row) for row in values
                ),
                "standardized_reference_tests": len(_PHASE1_STRICT_3D_STANDARDIZED_IDS),
                "referencable": sum(
                    _bool(row.get("referencable")) for row in standardized
                ),
                "evaluated": len(evaluated),
                "passed": passed,
                "failed": len(evaluated) - passed,
                "native_supported_nonstandardized_tests": sum(
                    _phase1_native_supported(row)
                    for row in values
                    if str(row["test_id"]) not in standardized_ids
                ),
                "conditional_accuracy": (
                    passed / len(evaluated) if evaluated else None
                ),
                "overall_referenced_success": (
                    passed / len(_PHASE1_STRICT_3D_STANDARDIZED_IDS)
                ),
            }
        )
    return output


def _phase1_strict_3d_by_filter(
    rows: Sequence[Mapping[str, Any]],
) -> list[dict[str, Any]]:
    """Summarize strict-3D support and conformity by filter family."""

    strict_ids = set(_PHASE1_STRICT_3D_TEST_IDS)
    standardized_ids = set(_PHASE1_STRICT_3D_STANDARDIZED_IDS)
    output = []
    adapters = sorted({str(row["adapter"]) for row in rows})
    for adapter in adapters:
        adapter_rows = [row for row in rows if str(row["adapter"]) == adapter]
        for family in _PHASE1_STRICT_3D_FILTER_ORDER:
            expected_ids = {
                test_id
                for test_id in strict_ids
                if str(PHASE1_FILTER_SPECS_BY_ID[test_id].parameters["filter"])
                == family
            }
            values = [
                row
                for row in adapter_rows
                if str(row.get("test_id", "")) in expected_ids
            ]
            expected_standardized = expected_ids.intersection(standardized_ids)
            standardized = [
                row for row in values if str(row["test_id"]) in standardized_ids
            ]
            evaluated = [row for row in standardized if _bool(row.get("evaluated"))]
            passed = sum(
                row.get("passed") is True or str(row.get("passed")).lower() == "true"
                for row in evaluated
            )
            output.append(
                {
                    "adapter": adapter,
                    "generator_version": ", ".join(
                        sorted(
                            {
                                str(row.get("generator_version", ""))
                                for row in adapter_rows
                                if str(row.get("generator_version", ""))
                            }
                        )
                    ),
                    "filter_family": family,
                    "strict_3d_scope": True,
                    "defined_filter_tests": len(expected_ids),
                    "native_supported_filter_tests": sum(
                        _phase1_native_supported(row) for row in values
                    ),
                    "candidate_maps_supplied": sum(
                        _phase1_candidate_supplied(row) for row in values
                    ),
                    "standardized_reference_tests": len(expected_standardized),
                    "evaluated": len(evaluated),
                    "passed": passed,
                    "failed": len(evaluated) - passed,
                    "conditional_accuracy": (
                        passed / len(evaluated) if evaluated else None
                    ),
                    "overall_referenced_success": (
                        passed / len(expected_standardized)
                        if expected_standardized
                        else None
                    ),
                }
            )
    return output


def _phase1_plot_strict_3d_matrix(
    rows: Sequence[Mapping[str, Any]], output: Path
) -> dict[str, Any]:
    """Render support and fixed-denominator accuracy for strict 3D filters."""

    import numpy as np

    plt = _pyplot()
    family_rows = _phase1_strict_3d_by_filter(rows)
    adapters = sorted({str(row["adapter"]) for row in family_rows})
    lookup = {
        (str(row["adapter"]), str(row["filter_family"])): row for row in family_rows
    }
    matrix = np.zeros((len(adapters), len(_PHASE1_STRICT_3D_FILTER_ORDER)), dtype=float)
    labels = np.empty(matrix.shape, dtype=object)
    for row_index, adapter in enumerate(adapters):
        for column_index, family in enumerate(_PHASE1_STRICT_3D_FILTER_ORDER):
            row = lookup[(adapter, family)]
            standardized = int(row["standardized_reference_tests"])
            passed = int(row["passed"])
            matrix[row_index, column_index] = (
                passed / standardized if standardized else 0.0
            )
            labels[row_index, column_index] = (
                f"N {int(row['native_supported_filter_tests'])}/"
                f"{int(row['defined_filter_tests'])}\n"
                f"M {int(row['candidate_maps_supplied'])}/"
                f"{int(row['defined_filter_tests'])}\n"
                f"E {int(row['evaluated'])}/{standardized}\n"
                f"P {passed}/{standardized}"
            )

    column_labels = []
    for family in _PHASE1_STRICT_3D_FILTER_ORDER:
        sample = next(row for row in family_rows if str(row["filter_family"]) == family)
        column_labels.append(
            f"{_FILTER_LABELS[family]}\n"
            f"(defined {sample['defined_filter_tests']}; "
            f"standardized {sample['standardized_reference_tests']})"
        )
    fig, ax = plt.subplots(figsize=(17.0, max(5.8, 1.08 * len(adapters) + 2.8)))
    image = ax.imshow(matrix, cmap="cividis", vmin=0.0, vmax=1.0, aspect="auto")
    ax.set_xticks(
        range(len(_PHASE1_STRICT_3D_FILTER_ORDER)),
        column_labels,
        rotation=25,
        ha="right",
    )
    ax.set_yticks(range(len(adapters)), [adapter.title() for adapter in adapters])
    ax.set_xticks(
        np.arange(len(_PHASE1_STRICT_3D_FILTER_ORDER) + 1) - 0.5,
        minor=True,
    )
    ax.set_yticks(np.arange(len(adapters) + 1) - 0.5, minor=True)
    ax.grid(which="minor", color="#222222", linewidth=1.0)
    ax.tick_params(which="minor", bottom=False, left=False)
    for row_index in range(len(adapters)):
        for column_index in range(len(_PHASE1_STRICT_3D_FILTER_ORDER)):
            value = matrix[row_index, column_index]
            ax.text(
                column_index,
                row_index,
                labels[row_index, column_index],
                ha="center",
                va="center",
                fontsize=11,
                fontweight="bold",
                color="white" if value < 0.45 else "#111111",
            )
    colorbar = fig.colorbar(image, ax=ax, pad=0.015)
    colorbar.set_label("Passed / fixed strict-3D standardized denominator")
    colorbar.ax.tick_params(labelsize=12)
    ax.set_xlabel("IBSI 2 Phase 1 strict-3D filter family", fontsize=12)
    ax.set_ylabel("Adapter", fontsize=12)
    ax.set_title("IBSI 2 Phase 1 strict-3D filter conformity")
    _style_axes(ax)
    fig.tight_layout()
    files = _save_figure(fig, output / "figure_phase1_strict_3d_filter_matrix")
    plt.close(fig)

    result_text = "; ".join(
        (
            f"{row['adapter'].title()} {row['generator_version']}: "
            f"{row['native_supported_filter_tests']}/27 native, "
            f"{row['candidate_maps_supplied']}/27 maps, "
            f"{row['passed']}/{row['evaluated']} passed/evaluated, "
            f"{row['passed']}/24 overall"
        )
        for row in _phase1_strict_3d_summary(rows)
    )
    return {
        "id": "figure_phase1_strict_3d_filter_matrix",
        "files": files,
        "caption": (
            "Primary IBSI 2 Phase 1 strict-3D result. The 27 IBSI-defined "
            "three-dimensional tests include 24 standardized response maps. "
            "Cells state N reviewed native support, M provenance-attested maps "
            "supplied, E evaluated standardized maps, and P passes over each "
            "fixed family reference denominator. Cell colour uses the fixed "
            "denominator, retaining unsupported and unevaluated maps. The nine "
            "IBSI-defined 2D tests, including all four Gabor tests, are excluded. "
            "Results: " + result_text + "."
        ),
        "alt_text": (
            "Adapter-by-filter-family matrix for 27 strict three-dimensional "
            "Phase 1 tests and 24 standardized maps. Direct labels report native "
            "support, supplied maps, evaluations, and passes. " + result_text + "."
        ),
    }


def _phase1_plot_matrix(
    rows: Sequence[Mapping[str, Any]], output: Path
) -> dict[str, Any]:
    import numpy as np

    plt = _pyplot()

    adapters = sorted({str(row["adapter"]) for row in rows})
    groups = [str(number) for number in range(1, 11)]
    matrix = np.full((len(adapters), len(groups)), np.nan)
    labels = np.empty(matrix.shape, dtype=object)
    for row_index, adapter in enumerate(adapters):
        for column_index, group in enumerate(groups):
            selected = [
                row
                for row in rows
                if str(row["adapter"]) == adapter
                and str(row["test_id"]).split(".", 1)[0] == group
            ]
            native_supported = sum(_phase1_native_supported(row) for row in selected)
            evaluated = [row for row in selected if _bool(row.get("evaluated"))]
            passed = sum(
                row.get("passed") is True or str(row.get("passed")).lower() == "true"
                for row in evaluated
            )
            matrix[row_index, column_index] = (
                passed / len(evaluated) if evaluated else np.nan
            )
            labels[row_index, column_index] = (
                f"N {native_supported}/{len(selected)}\nP {passed}/{len(evaluated)}"
            )
    fig, ax = plt.subplots(figsize=(14.5, max(5.0, 0.85 * len(adapters) + 2.5)))
    image = ax.imshow(
        np.ma.masked_invalid(matrix), cmap="cividis", vmin=0.0, vmax=1.0, aspect="auto"
    )
    ax.set_facecolor("#eeeeee")
    ax.set_xticks(range(len(groups)), [f"Filter {group}" for group in groups])
    ax.set_yticks(range(len(adapters)), [adapter.title() for adapter in adapters])
    ax.set_xticks(np.arange(len(groups) + 1) - 0.5, minor=True)
    ax.set_yticks(np.arange(len(adapters) + 1) - 0.5, minor=True)
    ax.grid(which="minor", color="#222222", linewidth=1.0)
    ax.tick_params(which="minor", bottom=False, left=False)
    for row_index in range(len(adapters)):
        for column_index in range(len(groups)):
            value = matrix[row_index, column_index]
            color = "white" if math.isfinite(value) and value < 0.45 else "#111111"
            ax.text(
                column_index,
                row_index,
                labels[row_index, column_index],
                ha="center",
                va="center",
                fontsize=12,
                fontweight="bold",
                color=color,
            )
    colorbar = fig.colorbar(image, ax=ax, pad=0.02)
    colorbar.set_label("Passed / evaluated response maps", fontsize=12)
    colorbar.ax.tick_params(labelsize=12)
    ax.set_xlabel("IBSI 2 filter-test group", fontsize=12)
    ax.set_ylabel("Adapter", fontsize=12)
    ax.set_title("IBSI 2 Phase 1 native filter support and voxel accuracy")
    _style_axes(ax)
    fig.tight_layout()
    files = _save_figure(fig, output / "figure_phase1_filter_matrix")
    plt.close(fig)
    return {
        "id": "figure_phase1_filter_matrix",
        "files": files,
        "caption": (
            "IBSI 2 Phase 1 native filter support and voxel-level tolerance "
            "accuracy grouped by filter family. Cells directly show N explicitly "
            "reviewed native-supported tests/defined and P passed/evaluated."
        ),
        "alt_text": (
            "Outlined matrix groups IBSI 2 filter tests; cells label reviewed native "
            "support over defined tests and passed over evaluated maps. "
            + "; ".join(
                f"{adapter.title()}: "
                f"{sum(_phase1_native_supported(row) for row in rows if str(row['adapter']) == adapter)}/36 native-supported, "
                f"{sum(row.get('passed') is True or str(row.get('passed')).lower() == 'true' for row in rows if str(row['adapter']) == adapter and _bool(row.get('evaluated')))}/"
                f"{sum(_bool(row.get('evaluated')) for row in rows if str(row['adapter']) == adapter)} passed/evaluated"
                for adapter in adapters
            )
            + "."
        ),
    }


def _phase1_plot_support(
    summary: Sequence[Mapping[str, Any]], output: Path
) -> dict[str, Any]:
    plt = _pyplot()

    adapters = [str(row["adapter"]) for row in summary]
    fig, ax = plt.subplots(figsize=(9.5, max(4.5, 0.72 * len(adapters) + 2.0)))
    for position, row in enumerate(summary):
        adapter = str(row["adapter"])
        supported = int(row["native_supported_filter_tests"])
        defined = int(row["defined_filter_tests"])
        color, _, hatch = ADAPTER_STYLES.get(adapter, ("#555555", "o", "//"))
        ax.barh(
            position,
            supported,
            height=0.62,
            color=color,
            edgecolor="#111111",
            linewidth=1.2,
            hatch=hatch,
        )
        ax.text(
            supported + 0.4,
            position,
            f"{supported}/{defined}",
            va="center",
            fontsize=12,
        )
    ax.set_yticks(range(len(adapters)), [adapter.title() for adapter in adapters])
    ax.invert_yaxis()
    ax.set_xlim(0, len(IBSI2_PHASE1_TEST_IDS) * 1.14)
    ax.set_xlabel("Explicitly reviewed native-supported filter tests", fontsize=12)
    ax.set_title("IBSI 2 Phase 1 native filter support")
    ax.grid(axis="x", color="#cccccc", linestyle=":", linewidth=0.8)
    _style_axes(ax)
    fig.tight_layout()
    files = _save_figure(fig, output / "figure_phase1_filter_support")
    plt.close(fig)
    return {
        "id": "figure_phase1_filter_support",
        "files": files,
        "caption": (
            "Explicitly reviewed native IBSI 2 Phase 1 filter support. Direct labels "
            "show native-supported/36 defined tests; outlines and hatches provide "
            "redundant cues."
        ),
        "alt_text": (
            "Horizontal hatched bars compare reviewed native IBSI 2 Phase 1 filter "
            "support: "
            + "; ".join(
                f"{str(row['adapter']).title()} "
                f"{int(row['native_supported_filter_tests'])}/36"
                for row in summary
            )
            + "."
        ),
    }


def _phase1_zero_tolerance_ratio_omissions(
    rows: Sequence[Mapping[str, Any]],
) -> list[dict[str, str]]:
    """Return evaluated rows whose error/tolerance ratio is undefined."""

    return [
        {
            "adapter": str(row.get("adapter", "")),
            "test_id": str(row.get("test_id", "")),
        }
        for row in rows
        if _bool(row.get("evaluated"))
        and _optional_float(row.get("voxel_tolerance")) == 0.0
    ]


def _phase1_plot_ratios(
    rows: Sequence[Mapping[str, Any]], output: Path
) -> Optional[dict[str, Any]]:
    plt = _pyplot()

    selected = []
    zero_tolerance_omissions = _phase1_zero_tolerance_ratio_omissions(rows)
    for row in rows:
        tolerance = _optional_float(row.get("voxel_tolerance"))
        error = _optional_float(row.get("max_abs_error"))
        if (
            not _bool(row.get("evaluated"))
            or tolerance is None
            or error is None
            or tolerance <= 0.0
        ):
            continue
        selected.append((str(row["adapter"]), str(row["test_id"]), error / tolerance))
    if not selected:
        return None
    adapters = sorted({item[0] for item in selected})
    fig, ax = plt.subplots(figsize=(11.0, 6.2))
    for adapter_index, adapter in enumerate(adapters):
        values = sorted(
            (test_id, ratio)
            for current, test_id, ratio in selected
            if current == adapter
        )
        color, marker, _ = ADAPTER_STYLES.get(adapter, ("#555555", "o", ""))
        ax.scatter(
            [index + adapter_index * 0.08 for index in range(len(values))],
            [max(ratio, 1e-8) for _, ratio in values],
            color=color,
            marker=marker,
            edgecolors="#111111",
            linewidths=0.45,
            s=42,
            label=adapter.title(),
        )
    ax.axhline(
        1.0,
        color="#111111",
        linestyle="--",
        linewidth=1.5,
        label="Voxel tolerance boundary",
    )
    ax.set_yscale("log")
    ax.set_xlabel("Deterministic standardized test order", fontsize=12)
    ax.set_ylabel("Maximum voxel error / allowed voxel error", fontsize=12)
    ax.set_title("IBSI 2 Phase 1 worst-voxel error relative to tolerance")
    ax.grid(True, which="both", color="#cccccc", linestyle=":", linewidth=0.75)
    ax.legend(fontsize=12, frameon=True, edgecolor="#333333", ncol=2)
    _style_axes(ax)
    fig.tight_layout()
    files = _save_figure(fig, output / "figure_phase1_voxel_tolerance_ratio")
    plt.close(fig)
    omission_text = (
        f" {len(zero_tolerance_omissions)} evaluated zero-tolerance row(s) are "
        "omitted because division by zero has no finite ratio; they remain in the "
        "detail and pass/fail results, and pass only when maximum absolute error is zero."
        if zero_tolerance_omissions
        else " No evaluated zero-tolerance rows were omitted."
    )
    return {
        "id": "figure_phase1_voxel_tolerance_ratio",
        "files": files,
        "caption": (
            "Maximum response-map voxel error divided by the one-percent "
            "reference-map-range tolerance. "
            "Marker shape identifies adapters and the dashed ratio-one line marks the boundary."
            + omission_text
        ),
        "alt_text": (
            "Log scatter plot shows worst-voxel error ratios by adapter; "
            + "; ".join(
                f"{adapter.title()} maximum ratio "
                f"{max(ratio for current, _, ratio in selected if current == adapter):.3g}"
                for adapter in adapters
            )
            + ". Marker shapes identify adapters and a dashed line marks the pass boundary."
            + omission_text
        ),
        "zero_tolerance_ratio_omissions": zero_tolerance_omissions,
    }


def generate_ibsi2_phase1_report(
    rows: Sequence[Mapping[str, Any]],
    output_dir: Path,
    *,
    source_metadata: Optional[Mapping[str, Any]] = None,
) -> dict[str, Any]:
    if not rows:
        raise ValueError("At least one IBSI 2 Phase 1 row is required")
    output = Path(output_dir).expanduser().resolve()
    output.mkdir(parents=True, exist_ok=True)
    detail_rows = []
    for row in rows:
        detail = dict(row)
        detail.setdefault("candidate_supplied", False)
        test_id = str(detail.get("test_id", ""))
        if test_id in PHASE1_FILTER_SPECS_BY_ID:
            parameters = PHASE1_FILTER_SPECS_BY_ID[test_id].parameters
            detail["filter_family"] = str(parameters["filter"])
            detail["kernel_dimensionality"] = int(parameters["dimensionality"])
            detail["strict_3d_scope"] = test_id in _PHASE1_STRICT_3D_TEST_IDS
        candidate_path = str(detail.get("candidate_path", "")).strip()
        if candidate_path:
            detail["candidate_path"] = _public_candidate_path(candidate_path)
        detail_rows.append(detail)
    detail_fields = list(dict.fromkeys(key for row in detail_rows for key in row))
    summary = _phase1_summary(rows)
    strict_3d_summary = _phase1_strict_3d_summary(rows)
    strict_3d_by_filter = _phase1_strict_3d_by_filter(rows)
    zero_tolerance_omissions = _phase1_zero_tolerance_ratio_omissions(rows)
    publication_complete = _ibsi2_phase1_row_complete(rows)
    if source_metadata is not None and isinstance(
        source_metadata.get("publication_complete"), bool
    ):
        publication_complete = publication_complete and bool(
            source_metadata["publication_complete"]
        )
    atomic_write_text(
        output / "phase1_detail.csv", _csv_text(detail_rows, detail_fields)
    )
    atomic_write_text(
        output / "phase1_summary.csv", _csv_text(summary, list(summary[0]))
    )
    atomic_write_text(
        output / "phase1_strict_3d_summary.csv",
        _csv_text(strict_3d_summary, list(strict_3d_summary[0])),
    )
    atomic_write_text(
        output / "phase1_strict_3d_by_filter.csv",
        _csv_text(strict_3d_by_filter, list(strict_3d_by_filter[0])),
    )
    lines = [
        "# IBSI 2 Phase 1 compliance summary",
        "",
        (
            "**Publication execution status: complete exhaustive native-support "
            "design; every native-supported response map was processed.**"
            if publication_complete
            else "**Publication execution status: incomplete; native support is not "
            "exhaustively declared or one or more native-supported response maps "
            "were not processed. Do not report adapter accuracy as a complete IBSI 2 result.**"
        ),
        "",
        (
            "Native support comes from the exhaustive reviewed support declarations; "
            "candidate-map availability is reported separately. Voxel "
            "accuracy uses the all-voxel boundary of 1% of the reference-map range. "
            "The primary strict-3D result reports both conditional passed/evaluated "
            "accuracy and passed/24 standardized maps."
        ),
        "",
        (
            "Ratio-plot disclosure: "
            f"{len(zero_tolerance_omissions)} evaluated zero-tolerance row(s) are "
            "omitted from the error/tolerance scatter because their ratio is "
            "undefined. They remain in the detail and pass/fail results; a "
            "zero-tolerance row passes only when maximum absolute error is zero."
        ),
        "",
        "## Primary strict-3D result",
        "",
        (
            "IBSI Phase 1 contains 27 filter tests whose official kernel "
            "dimensionality is 3D; 24 have standardized response maps. The nine "
            "officially 2D tests are excluded from this headline, including all "
            "four Gabor tests."
        ),
        "",
        "| Adapter | Generator version | Native supported / 27 defined | Candidate maps supplied / 27 defined | Evaluated / 24 referenced | Passed / evaluated | Passed / 24 standardized |",
        "| --- | --- | ---: | ---: | ---: | ---: | ---: |",
    ]
    for row in strict_3d_summary:
        lines.append(
            f"| {row['adapter']} | {row['generator_version']} | "
            f"{row['native_supported_filter_tests']}/{row['defined_filter_tests']} | "
            f"{row['candidate_maps_supplied']}/{row['defined_filter_tests']} | "
            f"{row['evaluated']}/{row['standardized_reference_tests']} | "
            f"{row['passed']}/{row['evaluated']} | "
            f"{row['passed']}/{len(_PHASE1_STRICT_3D_STANDARDIZED_IDS)} |"
        )
    lines.extend(
        [
            "",
            "## Secondary full 36-test design",
            "",
            "| Adapter | Generator version | Native supported / 36 defined | Candidate maps supplied / 36 defined | Evaluated / 33 referenced | Passed / evaluated | Passed / 33 standardized |",
            "| --- | --- | ---: | ---: | ---: | ---: | ---: |",
        ]
    )
    for row in summary:
        lines.append(
            f"| {row['adapter']} | {row['generator_version']} | "
            f"{row['native_supported_filter_tests']}/{row['defined_filter_tests']} | "
            f"{row['candidate_maps_supplied']}/{row['defined_filter_tests']} | "
            f"{row['evaluated']}/{row['standardized_reference_tests']} | "
            f"{row['passed']}/{row['evaluated']} | "
            f"{row['passed']}/{len(IBSI2_PHASE1_STANDARDIZED_IDS)} |"
        )
    atomic_write_text(output / "phase1_summary.md", "\n".join(lines) + "\n")
    data_dictionary = _data_dictionary_rows(
        [
            ("phase1_strict_3d_summary.csv", strict_3d_summary),
            ("phase1_strict_3d_by_filter.csv", strict_3d_by_filter),
            ("phase1_summary.csv", summary),
            ("phase1_detail.csv", detail_rows),
        ]
    )
    atomic_write_text(
        output / "data_dictionary.csv",
        _csv_text(data_dictionary, ["table", "field", "definition"]),
    )
    _write_xlsx(
        output / "phase1_tables.xlsx",
        detail_rows,
        summary,
        [],
        extra_tables=(
            ("Strict 3D summary", strict_3d_summary),
            ("Strict 3D by filter", strict_3d_by_filter),
            ("Data dictionary", data_dictionary),
        ),
    )
    figures = [
        _phase1_plot_strict_3d_matrix(rows, output),
        _phase1_plot_support(summary, output),
        _phase1_plot_matrix(rows, output),
    ]
    ratio = _phase1_plot_ratios(rows, output)
    if ratio is not None:
        figures.append(ratio)
    manifest = {
        "schema_version": 1,
        "source_metadata": dict(source_metadata or {}),
        "defined_tests": len(IBSI2_PHASE1_TEST_IDS),
        "standardized_reference_tests": len(IBSI2_PHASE1_STANDARDIZED_IDS),
        "nonstandardized_tests": list(IBSI2_PHASE1_NONSTANDARDIZED_IDS),
        "strict_3d_defined_tests": len(_PHASE1_STRICT_3D_TEST_IDS),
        "strict_3d_standardized_reference_tests": len(
            _PHASE1_STRICT_3D_STANDARDIZED_IDS
        ),
        "strict_3d_test_ids": list(_PHASE1_STRICT_3D_TEST_IDS),
        "excluded_2d_test_ids": [
            test_id
            for test_id in IBSI2_PHASE1_TEST_IDS
            if test_id not in _PHASE1_STRICT_3D_TEST_IDS
        ],
        "comparison_policy": IBSI2_PHASE1_COMPARISON_RULE,
        "ratio_plot_policy": {
            "positive_tolerance_rows": (
                "maximum absolute voxel error / voxel tolerance"
            ),
            "zero_tolerance_rows": (
                "ratio undefined and omitted from scatter only; retained in detail "
                "and pass/fail results; pass requires zero maximum absolute error"
            ),
            "zero_tolerance_ratio_omission_count": len(zero_tolerance_omissions),
            "zero_tolerance_ratio_omissions": zero_tolerance_omissions,
        },
        "denominator_policy": {
            "strict_3d_native_filter_support": (
                "explicit reviewed native-supported declarations / 27 "
                "IBSI-defined three-dimensional filter tests"
            ),
            "strict_3d_candidate_map_availability": (
                "provenance-attested candidate response maps supplied / 27 "
                "IBSI-defined three-dimensional filter tests"
            ),
            "strict_3d_conditional_accuracy": (
                "passed / evaluated among 24 strict-3D standardized response maps"
            ),
            "strict_3d_overall_referenced_success": (
                "passed / 24 strict-3D standardized response maps"
            ),
            "native_filter_support": (
                "explicit reviewed native-supported declarations / 36 defined filter tests"
            ),
            "candidate_map_availability": (
                "provenance-attested candidate response maps supplied / 36 defined filter tests; "
                "availability is not interpreted as native filter support"
            ),
            "conditional_accuracy": "passed / evaluated standardized response maps",
            "overall_referenced_success": "passed / 33 standardized reference tests",
        },
        "accessibility": {
            "font_family": "DejaVu Sans",
            "minimum_label_size_points": 12,
            "redundant_cues": "direct labels, outlines, hatches, and marker shapes",
            "formats": ["PDF", "SVG", "PNG"],
        },
        "figures": figures,
        "tables": [
            "phase1_detail.csv",
            "phase1_strict_3d_summary.csv",
            "phase1_strict_3d_by_filter.csv",
            "phase1_summary.csv",
            "phase1_summary.md",
            "data_dictionary.csv",
            "phase1_tables.xlsx",
        ],
        "data_dictionary": "data_dictionary.csv",
        "figure_captions": "phase1_figure_captions.md",
        "publication_execution_complete": publication_complete,
        "native_supported_filter_tests": {
            str(row["adapter"]): int(row["native_supported_filter_tests"])
            for row in summary
        },
        "native_filter_denominator": len(IBSI2_PHASE1_TEST_IDS),
        "strict_3d_native_supported_filter_tests": {
            str(row["adapter"]): int(row["native_supported_filter_tests"])
            for row in strict_3d_summary
        },
        "strict_3d_native_filter_denominator": len(_PHASE1_STRICT_3D_TEST_IDS),
    }
    captions = [
        "# Figure captions",
        "",
        "## Interpretation key",
        "",
        (
            "N denotes explicitly reviewed native filter support, M a supplied "
            "provenance-attested map, E evaluated standardized maps, and P passes "
            "over the stated denominator. The tolerance-ratio scatter excludes "
            f"{len(zero_tolerance_omissions)} evaluated zero-tolerance row(s) because "
            "division by zero is undefined; those rows remain explicit in the detail "
            "and pass/fail results."
        ),
        "",
    ]
    for figure in figures:
        captions.extend(
            [
                f"## {figure['id']}",
                "",
                figure["caption"],
                "",
                f"Alt text: {figure['alt_text']}",
                "",
            ]
        )
    atomic_write_text(output / "phase1_figure_captions.md", "\n".join(captions))
    release_artifacts = [
        ("phase1_detail.csv", "table"),
        ("phase1_strict_3d_summary.csv", "table"),
        ("phase1_strict_3d_by_filter.csv", "table"),
        ("phase1_summary.csv", "table"),
        ("phase1_summary.md", "narrative"),
        ("data_dictionary.csv", "data_dictionary"),
        ("phase1_tables.xlsx", "workbook"),
        ("phase1_figure_captions.md", "captions"),
        *[(filename, "figure") for figure in figures for filename in figure["files"]],
    ]
    release_artifacts.append(("phase1_report_manifest.json", "manifest"))
    manifest["release_manifest"] = "release_manifest.json"
    manifest["release_artifact_count"] = len(
        {relative_name for relative_name, _ in release_artifacts}
    )
    atomic_write_json(output / "phase1_report_manifest.json", manifest)
    release_manifest = _write_release_manifest(output, release_artifacts)
    if len(release_manifest["artifacts"]) != manifest["release_artifact_count"]:
        raise RuntimeError("Phase 1 release artifact count changed during manifesting")
    return manifest
