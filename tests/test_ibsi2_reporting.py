from __future__ import annotations

import csv
import hashlib
import json
import tempfile
import unittest
from dataclasses import replace
from pathlib import Path

from bench.compliance.models import ComparisonRecord
from bench.compliance.references import (
    IBSI2_PHASE1_NONSTANDARDIZED_IDS,
    IBSI2_PHASE1_STANDARDIZED_IDS,
    IBSI2_PHASE1_TEST_IDS,
    IBSI2_PHASE2_DEFINED_FILTER_IDS,
    IBSI2_PHASE2_NONSTANDARDIZED_PUBLISHED_PAIR,
    IBSI2_PHASE2_PUBLISHED_FILTER_IDS,
    IBSI2_PHASE2_TAGS,
)
from bench.compliance.report import (
    _ibsi2_phase2_row_complete,
    _markdown_summary,
    _overall_summary,
    generate_compliance_report,
    generate_ibsi2_phase1_report,
)


def _phase2_records() -> list[ComparisonRecord]:
    published = set(IBSI2_PHASE2_PUBLISHED_FILTER_IDS)
    records = []
    for configuration in IBSI2_PHASE2_DEFINED_FILTER_IDS:
        for feature_tag in IBSI2_PHASE2_TAGS:
            standardized = (
                configuration in published
                and (configuration, feature_tag)
                != IBSI2_PHASE2_NONSTANDARDIZED_PUBLISHED_PAIR
            )
            records.append(
                ComparisonRecord(
                    specification="IBSI 2",
                    phase="phase2",
                    adapter="pictologics",
                    software_version="0.5.1",
                    configuration=configuration,
                    profile="phase2_defined_filters",
                    aggregation="not_applicable",
                    family="intensity",
                    feature_name=feature_tag,
                    feature_tag=feature_tag,
                    semantic_key=f"intensity_statistics.{feature_tag}",
                    ibsi_code=feature_tag,
                    standardized=standardized,
                    observed_supported=True,
                    mapped=True,
                    attempted=True,
                    finite=True,
                    referencable=standardized,
                    evaluated=standardized,
                    passed=True if standardized else None,
                    status="pass" if standardized else "not_standardized",
                    native_feature_names=feature_tag,
                    value=1.0,
                    reference_value=1.0 if standardized else None,
                    tolerance=0.1 if standardized else None,
                    raw_abs_error=0.0 if standardized else None,
                    comparison_error=0.0 if standardized else None,
                    error_tolerance_ratio=0.0 if standardized else None,
                    comparison_policy="raw absolute error <= source tolerance",
                    detail=(
                        "reviewed native-supported filter declaration: unit-test; "
                        "evidence: unit-test capability audit"
                    ),
                )
            )
    return records


class IBSI2ReportingTests(unittest.TestCase):
    def test_phase2_report_uses_full_design_and_published_denominators(self) -> None:
        try:
            import matplotlib  # noqa: F401
            import openpyxl  # noqa: F401
        except ImportError:
            self.skipTest("matplotlib and openpyxl are required")

        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary)
            manifest = generate_compliance_report(_phase2_records(), output)
            with (output / "compliance_summary.csv").open(
                encoding="utf-8", newline=""
            ) as stream:
                summary = next(csv.DictReader(stream))
            self.assertEqual(summary["defined_features"], "396")
            self.assertEqual(summary["supported_features"], "396")
            self.assertEqual(summary["finite_designed_checks"], "396")
            self.assertEqual(summary["published_reference_rows"], "324")
            self.assertEqual(summary["finite_published_reference_rows"], "324")
            self.assertEqual(summary["defined_standardized_checks"], "323")
            self.assertEqual(summary["nonstandardized_defined_checks"], "73")
            self.assertEqual(summary["native_supported_filters"], "22")
            self.assertEqual(summary["defined_filters"], "22")
            self.assertEqual(summary["finite_designed_checks"], "396")
            self.assertEqual(summary["finite_calculation_denominator"], "396")
            self.assertEqual(summary["conditional_accuracy_numerator"], "323")
            self.assertEqual(summary["conditional_accuracy_denominator"], "323")
            self.assertEqual(summary["overall_standardized_success_denominator"], "323")
            self.assertEqual(
                summary["filter_support_inferred_from_candidate_maps"], "False"
            )

            with (output / "compliance_phase2_b_summary.csv").open(
                encoding="utf-8", newline=""
            ) as stream:
                b_summary = next(csv.DictReader(stream))
            self.assertEqual(b_summary["defined_b_filters"], "11")
            self.assertEqual(b_summary["native_supported_b_designed_filters"], "11")
            self.assertEqual(b_summary["published_b_filters"], "9")
            self.assertEqual(b_summary["native_supported_b_published_filters"], "9")
            self.assertEqual(b_summary["published_b_reference_rows"], "162")
            self.assertEqual(b_summary["finite_published_b_checks"], "162")
            self.assertEqual(b_summary["standardized_b_checks"], "161")
            self.assertEqual(b_summary["evaluated_standardized_checks"], "161")
            self.assertEqual(b_summary["passed_standardized_checks"], "161")
            self.assertEqual(
                b_summary["overall_standardized_success_denominator"], "161"
            )

            with (output / "compliance_phase2_b_by_configuration.csv").open(
                encoding="utf-8", newline=""
            ) as stream:
                b_configurations = list(csv.DictReader(stream))
            self.assertEqual(len(b_configurations), 11)
            self.assertEqual(
                {row["protocol_variant"] for row in b_configurations}, {"B"}
            )
            self.assertEqual(
                sum(
                    row["published_configuration"] == "True" for row in b_configurations
                ),
                9,
            )
            b_gabor = next(
                row for row in b_configurations if row["configuration"] == "5.B"
            )
            self.assertEqual(b_gabor["workflow_dimensionality"], "3")
            self.assertEqual(b_gabor["kernel_dimensionality"], "2")
            self.assertEqual(b_gabor["three_plane_gabor_exception"], "True")
            self.assertEqual(b_gabor["published_configuration"], "True")
            self.assertEqual(
                {
                    row["configuration"]
                    for row in b_configurations
                    if row["published_configuration"] == "False"
                },
                {"10.B", "11.B"},
            )

            with (output / "compliance_by_configuration.csv").open(
                encoding="utf-8", newline=""
            ) as stream:
                configurations = list(csv.DictReader(stream))
            self.assertEqual(len(configurations), 22)
            self.assertEqual(
                {row["protocol_variant"] for row in configurations}, {"A", "B"}
            )
            self.assertTrue(
                all(row["native_supported"] == "True" for row in configurations)
            )
            gabor_5b = next(
                row for row in configurations if row["configuration"] == "5.B"
            )
            self.assertEqual(gabor_5b["defined_checks"], "18")
            self.assertEqual(gabor_5b["finite_checks"], "18")
            self.assertEqual(gabor_5b["overall_standardized_success_denominator"], "18")

            from openpyxl import load_workbook

            workbook = load_workbook(output / "compliance_tables.xlsx", read_only=True)
            self.assertIn("B summary", workbook.sheetnames)
            self.assertIn("B by configuration", workbook.sheetnames)
            self.assertIn("By configuration", workbook.sheetnames)
            self.assertIn("Data dictionary", workbook.sheetnames)
            workbook.close()

            markdown = (output / "compliance_summary.md").read_text(encoding="utf-8")
            for denominator in ("11", "9", "162", "161", "396", "324", "323", "73"):
                self.assertIn(denominator, markdown)
            self.assertNotIn("Table 2", markdown)
            self.assertIn("Primary volumetric Configuration-B result", markdown)
            self.assertIn("Secondary full A/B design", markdown)
            self.assertIn("8.B/stat_qcod", markdown)
            self.assertIn("not a native 3D Gabor kernel", markdown)
            self.assertIn("complete exhaustive native-support design", markdown)
            self.assertIn("Native filters / 22", markdown)
            self.assertIn("component-level", markdown)
            self.assertIn("protocol-controlled common preprocessing", markdown)
            self.assertIn("A as the slice-wise 2D pathway", markdown)
            self.assertIn("B as the volumetric 3D pathway", markdown)
            self.assertIn("overall standardized success", markdown)
            self.assertTrue(manifest["publication_execution_complete"])
            self.assertEqual(
                manifest["native_supported_filter_configurations"]["pictologics"],
                22,
            )
            self.assertEqual(manifest["native_filter_denominator"], 22)
            self.assertEqual(
                manifest["native_supported_b_designed_filter_configurations"][
                    "pictologics"
                ],
                11,
            )
            self.assertEqual(
                manifest["native_supported_b_published_filter_configurations"][
                    "pictologics"
                ],
                9,
            )
            self.assertEqual(
                manifest["phase2_b_scope"]["published_reference_rows"], 162
            )
            self.assertEqual(manifest["phase2_b_scope"]["standardized_checks"], 161)
            self.assertEqual(
                manifest["phase2_b_scope"]["nonstandardized_published_pair"],
                {"configuration": "8.B", "feature_tag": "stat_qcod"},
            )

            captions = (output / "figure_captions.md").read_text(encoding="utf-8")
            self.assertNotIn("Table 2", captions)
            self.assertIn("Pictologics 0.5.1: 396/396 finite", captions)
            self.assertIn("Configurations 10A, 10B, 11A", captions)
            self.assertIn("N yes/no", captions)
            self.assertIn("component-level", captions)
            self.assertIn("slice-wise 2D A pathway", captions)
            self.assertIn("volumetric 3D B pathway", captions)
            self.assertIn("5B is the specified exception", captions)
            self.assertIn("5.B (dagger)", captions)
            self.assertIn("8.B has 17", captions)
            self.assertEqual(
                manifest["evaluation_scope"],
                (
                    "Component-level filter/statistic evaluation after "
                    "protocol-controlled common preprocessing; adapter-native image "
                    "import and resampling are outside this comparison."
                ),
            )

            dictionary = (output / "data_dictionary.csv").read_text(encoding="utf-8")
            self.assertIn("overall_standardized_success", dictionary)
            self.assertIn("unsupported or unevaluated", dictionary)
            self.assertIn("protocol_variant", dictionary)

            release = json.loads(
                (output / "release_manifest.json").read_text(encoding="utf-8")
            )
            release_paths = {entry["path"] for entry in release["artifacts"]}
            self.assertIn("compliance_phase2_b_summary.csv", release_paths)
            self.assertIn("compliance_phase2_b_by_configuration.csv", release_paths)
            self.assertIn("compliance_by_configuration.csv", release_paths)
            self.assertIn("compliance_tables.xlsx", release_paths)
            self.assertIn("figure_captions.md", release_paths)
            self.assertIn("figure_phase2_b_filter_matrix.svg", release_paths)
            self.assertIn("figure_support_accuracy_matrix.svg", release_paths)
            self.assertIn("report_manifest.json", release_paths)
            for entry in release["artifacts"]:
                self.assertFalse(Path(entry["path"]).is_absolute())
                self.assertNotIn("..", Path(entry["path"]).parts)
                self.assertEqual(
                    entry["sha256"],
                    hashlib.sha256((output / entry["path"]).read_bytes()).hexdigest(),
                )
            policy = manifest["denominator_policy"]
            self.assertEqual(policy["ibsi2_phase2_defined_checks"], 396)
            self.assertEqual(policy["ibsi2_phase2_published_reference_rows"], 324)
            self.assertEqual(policy["ibsi2_phase2_standardized_checks"], 323)
            self.assertEqual(policy["ibsi2_phase2_nonstandardized_checks"], 73)
            self.assertEqual(policy["ibsi2_phase2_b_defined_filters"], 11)
            self.assertEqual(policy["ibsi2_phase2_b_published_filters"], 9)
            self.assertEqual(policy["ibsi2_phase2_b_published_reference_rows"], 162)
            self.assertEqual(policy["ibsi2_phase2_b_standardized_checks"], 161)
            self.assertIn("never inferred", policy["native_filter_support"])

    def test_phase2_report_marks_unsupplied_candidate_design_incomplete(self) -> None:
        records = _phase2_records()
        records[0] = replace(
            records[0],
            observed_supported=False,
            attempted=False,
            finite=False,
            evaluated=False,
            passed=None,
            status="candidate_not_supplied",
        )
        self.assertFalse(_ibsi2_phase2_row_complete(records))
        markdown = _markdown_summary(_overall_summary(records), records)
        self.assertIn("Publication execution status: incomplete", markdown)

    def test_phase2_reviewed_native_unsupported_filter_is_complete(self) -> None:
        records = [
            replace(
                record,
                observed_supported=False,
                mapped=False,
                attempted=False,
                finite=False,
                referencable=record.referencable,
                evaluated=False,
                passed=None,
                status="native_unsupported",
                detail=(
                    "reviewed native-support declaration: unit-test unsupported; "
                    "evidence: unit-test capability audit"
                ),
            )
            if record.configuration == "11.B"
            else record
            for record in _phase2_records()
        ]
        self.assertTrue(_ibsi2_phase2_row_complete(records))
        summary = _overall_summary(records)[0]
        self.assertEqual(summary["native_supported_filters"], 21)
        self.assertEqual(summary["finite_designed_checks"], 378)

    def test_phase1_report_describes_supplied_maps_and_both_accuracy_rates(
        self,
    ) -> None:
        try:
            import matplotlib  # noqa: F401
            import openpyxl  # noqa: F401
        except ImportError:
            self.skipTest("matplotlib and openpyxl are required")

        standardized = set(IBSI2_PHASE1_STANDARDIZED_IDS)
        rows = []
        for test_id in IBSI2_PHASE1_TEST_IDS:
            is_standardized = test_id in standardized
            rows.append(
                {
                    "adapter": "pictologics",
                    "test_id": test_id,
                    "standardized": is_standardized,
                    "native_supported": True,
                    "candidate_supplied": True,
                    "referencable": is_standardized,
                    "evaluated": is_standardized,
                    "passed": True if is_standardized else None,
                    "status": "pass" if is_standardized else "not_standardized",
                    "generator_distribution": "pictologics",
                    "generator_version": "0.5.0",
                    "voxel_tolerance": (
                        0.0
                        if test_id == IBSI2_PHASE1_STANDARDIZED_IDS[0]
                        else (1.0 if is_standardized else None)
                    ),
                    "max_abs_error": 0.0
                    if test_id == IBSI2_PHASE1_STANDARDIZED_IDS[0]
                    else (0.1 if is_standardized else None),
                    "candidate_path": f"/private/candidate/maps/{test_id}.npy",
                }
            )

        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary)
            manifest = generate_ibsi2_phase1_report(rows, output)
            markdown = (output / "phase1_summary.md").read_text(encoding="utf-8")
            self.assertIn("Candidate maps supplied / 36 defined", markdown)
            self.assertIn("Native supported / 36", markdown)
            self.assertIn("Primary strict-3D result", markdown)
            self.assertIn("27 filter tests", markdown)
            self.assertIn("24 have standardized", markdown)
            self.assertIn("Secondary full 36-test design", markdown)
            self.assertIn("complete exhaustive native-support design", markdown)
            self.assertIn("Passed / evaluated", markdown)
            self.assertIn("Passed / 33 standardized", markdown)
            self.assertIn("1% of the reference-map range", markdown)
            self.assertIn("1 evaluated zero-tolerance row(s)", markdown)
            self.assertIn("omitted from the error/tolerance scatter", markdown)
            self.assertNotIn("Supported / 36 defined", markdown)
            self.assertEqual(
                manifest["nonstandardized_tests"],
                list(IBSI2_PHASE1_NONSTANDARDIZED_IDS),
            )
            detail_header = (
                (output / "phase1_detail.csv")
                .read_text(encoding="utf-8")
                .splitlines()[0]
            )
            self.assertIn("candidate_supplied", detail_header)
            self.assertNotIn("supported", detail_header.split(","))
            self.assertIn("native_supported", detail_header.split(","))
            detail = (output / "phase1_detail.csv").read_text(encoding="utf-8")
            self.assertNotIn("/private/candidate/maps", detail)
            self.assertIn("filter_family", detail_header.split(","))
            self.assertIn("kernel_dimensionality", detail_header.split(","))
            self.assertIn("strict_3d_scope", detail_header.split(","))

            with (output / "phase1_strict_3d_summary.csv").open(
                encoding="utf-8", newline=""
            ) as stream:
                strict_summary = next(csv.DictReader(stream))
            self.assertEqual(strict_summary["defined_filter_tests"], "27")
            self.assertEqual(strict_summary["native_supported_filter_tests"], "27")
            self.assertEqual(strict_summary["candidate_maps_supplied"], "27")
            self.assertEqual(strict_summary["standardized_reference_tests"], "24")
            self.assertEqual(strict_summary["evaluated"], "24")
            self.assertEqual(strict_summary["passed"], "24")

            with (output / "phase1_strict_3d_by_filter.csv").open(
                encoding="utf-8", newline=""
            ) as stream:
                strict_by_filter = list(csv.DictReader(stream))
            self.assertEqual(len(strict_by_filter), 7)
            self.assertEqual(
                sum(int(row["defined_filter_tests"]) for row in strict_by_filter),
                27,
            )
            self.assertEqual(
                sum(
                    int(row["standardized_reference_tests"]) for row in strict_by_filter
                ),
                24,
            )
            figure_manifest = json.loads(
                (output / "phase1_report_manifest.json").read_text(encoding="utf-8")
            )
            captions = " ".join(
                figure["caption"] for figure in figure_manifest["figures"]
            )
            self.assertIn("one-percent reference-map-range", captions)
            self.assertIn("1 evaluated zero-tolerance row(s)", captions)
            self.assertEqual(
                manifest["ratio_plot_policy"]["zero_tolerance_ratio_omission_count"],
                1,
            )
            self.assertEqual(manifest["release_manifest"], "release_manifest.json")
            self.assertEqual(manifest["strict_3d_defined_tests"], 27)
            self.assertEqual(manifest["strict_3d_standardized_reference_tests"], 24)
            self.assertEqual(
                manifest["strict_3d_native_supported_filter_tests"]["pictologics"],
                27,
            )
            release = json.loads(
                (output / "release_manifest.json").read_text(encoding="utf-8")
            )
            release_paths = {entry["path"] for entry in release["artifacts"]}
            self.assertIn("phase1_figure_captions.md", release_paths)
            self.assertIn("phase1_strict_3d_summary.csv", release_paths)
            self.assertIn("phase1_strict_3d_by_filter.csv", release_paths)
            self.assertIn("figure_phase1_strict_3d_filter_matrix.svg", release_paths)
            self.assertIn("phase1_report_manifest.json", release_paths)
            for entry in release["artifacts"]:
                self.assertEqual(
                    entry["sha256"],
                    hashlib.sha256((output / entry["path"]).read_bytes()).hexdigest(),
                )

    def test_phase1_report_visibly_warns_when_support_grid_is_incomplete(self) -> None:
        rows = [
            {
                "adapter": "pictologics",
                "test_id": "1.b.1",
                "standardized": True,
                "candidate_supplied": True,
                "referencable": True,
                "evaluated": True,
                "passed": True,
                "status": "pass",
            }
        ]
        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary)
            generate_ibsi2_phase1_report(rows, output)
            markdown = (output / "phase1_summary.md").read_text(encoding="utf-8")
            self.assertIn("Publication execution status: incomplete", markdown)


if __name__ == "__main__":
    unittest.main()
