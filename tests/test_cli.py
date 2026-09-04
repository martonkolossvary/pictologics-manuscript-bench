from __future__ import annotations

import io
import json
import tempfile
import unittest
from pathlib import Path
from unittest import mock

from bench.cli import _parser, _run_argv, main
from bench.compliance.references import (
    IBSI2_PHASE1_STANDARDIZED_IDS,
    IBSI2_PHASE1_TEST_IDS,
)


class CliForwardingTests(unittest.TestCase):
    def test_run_defaults_to_grouped_native_workloads(self) -> None:
        args = _parser().parse_args(["run", "--dataset-dir", "dataset"])
        self.assertEqual(args.workloads, "all")
        forwarded = _run_argv(args)
        self.assertEqual(forwarded[forwarded.index("--workloads") + 1], "all")

    def test_resume_and_guardrail_policy_are_forwarded(self) -> None:
        args = _parser().parse_args(
            [
                "run",
                "--dataset-dir",
                "dataset",
                "--guardrail-skip-ratio",
                "100",
                "--guardrail-min-observations",
                "2",
                "--checkpoint-interval",
                "3",
                "--resume",
                "--dry-run",
            ]
        )
        forwarded = _run_argv(args)
        for flag in (
            "--guardrail-skip-ratio",
            "--guardrail-min-observations",
            "--checkpoint-interval",
            "--resume",
            "--dry-run",
        ):
            self.assertIn(flag, forwarded)

    def test_machine_identity_and_cpu_metadata_are_forwarded(self) -> None:
        args = _parser().parse_args(
            [
                "run",
                "--dataset-dir",
                "dataset",
                "--machine-id",
                "windows-lab-01",
                "--machine-label",
                "Windows workstation",
                "--cpu-model",
                "Example CPU",
                "--cpu-base-ghz",
                "3.2",
                "--host-profile-id",
                "windows-lab-01",
                "--host-profile-sha256",
                "a" * 64,
                "--host-settings-json",
                '{"power_mode":"best_performance"}',
            ]
        )
        forwarded = _run_argv(args)
        expected = {
            "--machine-id": "windows-lab-01",
            "--machine-label": "Windows workstation",
            "--cpu-model": "Example CPU",
            "--cpu-base-ghz": "3.2",
            "--host-profile-id": "windows-lab-01",
            "--host-profile-sha256": "a" * 64,
            "--host-settings-json": '{"power_mode":"best_performance"}',
        }
        for flag, value in expected.items():
            self.assertEqual(forwarded[forwarded.index(flag) + 1], value)

    def test_compliance_runs_require_reference_manifests(self) -> None:
        ibsi1 = _parser().parse_args(
            [
                "compliance",
                "run-ibsi1",
                "--image",
                "image.nii.gz",
                "--mask",
                "mask.nii.gz",
                "--references",
                "references.csv",
                "--reference-manifest",
                "references.json",
                "--output-dir",
                "output",
            ]
        )
        self.assertEqual(ibsi1.reference_manifest, "references.json")

        phase2 = _parser().parse_args(
            [
                "compliance",
                "run-ibsi2-phase2",
                "--candidate-manifest",
                "candidates.json",
                "--references",
                "references.csv",
                "--reference-manifest",
                "references.json",
                "--output-dir",
                "output",
            ]
        )
        self.assertEqual(phase2.reference_manifest, "references.json")

    def test_ibsi1_a_to_e_executor_is_not_exposed_before_preprocessing_review(
        self,
    ) -> None:
        compliance = _parser()._subparsers._group_actions[0].choices["compliance"]
        command_action = compliance._subparsers._group_actions[0]
        self.assertNotIn("run-ibsi1-configurations", command_action.choices)

    def test_phase1_cli_report_serializes_shape_audit_fields_to_xlsx(self) -> None:
        try:
            import matplotlib  # noqa: F401
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("matplotlib and openpyxl are required")

        standardized = set(IBSI2_PHASE1_STANDARDIZED_IDS)
        rows = []
        support_declarations = {}
        for test_id in IBSI2_PHASE1_TEST_IDS:
            is_standardized = test_id in standardized
            rows.append(
                {
                    "adapter": "pictologics",
                    "test_id": test_id,
                    "standardized": is_standardized,
                    "candidate_supplied": True,
                    "referencable": is_standardized,
                    "evaluated": is_standardized,
                    "passed": True if is_standardized else None,
                    "status": "pass" if is_standardized else "not_standardized",
                    "generator_distribution": "pictologics",
                    "generator_version": "0.5.0",
                    "reference_shape": [64, 64, 64],
                    "candidate_shape": [64, 64, 64],
                    "voxel_tolerance": 1.0 if is_standardized else None,
                    "max_abs_error": 0.1 if is_standardized else None,
                }
            )
            support_declarations[("pictologics", test_id)] = {
                "adapter": "pictologics",
                "test_id": test_id,
                "native_supported": True,
                "reason": "unit-test native support",
                "evidence": "unit-test package API audit",
            }

        class CandidateEntries(list):
            pass

        candidate_entries = CandidateEntries()
        candidate_entries.adapters = ("pictologics",)
        candidate_entries.support_declarations = support_declarations

        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            reference_manifest = root / "reference-manifest.json"
            candidate_manifest = root / "candidate-manifest.json"
            output = root / "results"
            reference_manifest.write_text("{}\n", encoding="utf-8")
            candidate_manifest.write_text("{}\n", encoding="utf-8")

            with (
                mock.patch(
                    "bench.compliance.run.load_ibsi2_phase1_candidate_manifest",
                    return_value=candidate_entries,
                ),
                mock.patch(
                    "bench.compliance.run.configured_adapter_profiles",
                    return_value={
                        "pictologics": {
                            "distribution": "pictologics",
                            "version": "0.5.0",
                            "distribution_metadata_version": "0.5.0",
                            "python": "3.12",
                        }
                    },
                ),
                mock.patch(
                    "bench.compliance.evaluate.evaluate_ibsi2_phase1_candidates",
                    return_value=rows,
                ),
                mock.patch("sys.stdout", new=io.StringIO()),
            ):
                exit_code = main(
                    [
                        "compliance",
                        "evaluate-ibsi2-phase1",
                        "--reference-manifest",
                        str(reference_manifest),
                        "--reference-dir",
                        str(root),
                        "--candidate-manifest",
                        str(candidate_manifest),
                        "--adapters",
                        "pictologics",
                        "--output-dir",
                        str(output),
                    ]
                )

            self.assertEqual(exit_code, 0)
            workbook_path = output / "report" / "phase1_tables.xlsx"
            self.assertTrue(workbook_path.is_file())
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
            detail = workbook["Detail"]
            headings = [
                cell.value for cell in next(detail.iter_rows(min_row=1, max_row=1))
            ]
            first_row = [
                cell.value for cell in next(detail.iter_rows(min_row=2, max_row=2))
            ]
            self.assertEqual(
                first_row[headings.index("reference_shape")],
                json.dumps([64, 64, 64]),
            )
            self.assertEqual(
                first_row[headings.index("candidate_shape")],
                json.dumps([64, 64, 64]),
            )
            workbook.close()


if __name__ == "__main__":
    unittest.main()
